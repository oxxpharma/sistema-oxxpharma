"""Helper para gerar arquivos XLSX a partir de listas de dicts.

Usado pelos endpoints de export (invoices, commissions, points, withdrawals, etc).
"""
from __future__ import annotations
import io
from typing import List, Dict, Any, Optional, Sequence
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def make_xlsx(
    rows: Sequence[Dict[str, Any]],
    columns: Sequence[Dict[str, Any]],
    sheet_name: str = "Export",
    brand_color: str = "E8731A",
) -> bytes:
    """Gera XLSX em memoria.

    columns: [{key, label, width?, type? ('money'|'int'|'text')}]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    headers = [c.get("label") or c["key"] for c in columns]
    ws.append(headers)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor=brand_color)
    for i, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i)
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        out = []
        for col in columns:
            v = row.get(col["key"])
            t = col.get("type", "text")
            if v is None:
                out.append("")
            elif t == "money":
                try:
                    out.append(round(float(v), 2))
                except (TypeError, ValueError):
                    out.append(0)
            elif t == "int":
                try:
                    out.append(int(v))
                except (TypeError, ValueError):
                    out.append(0)
            elif t == "float":
                try:
                    out.append(float(v))
                except (TypeError, ValueError):
                    out.append(0)
            else:
                out.append(str(v))
        ws.append(out)

    # larguras
    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else f"A{chr(64 + i - 26)}"].width = col.get("width") or 18

    # formato moeda
    for i, col in enumerate(columns, start=1):
        if col.get("type") == "money":
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=i).number_format = '"R$" #,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_csv(
    rows: Sequence[Dict[str, Any]],
    columns: Sequence[Dict[str, Any]],
) -> bytes:
    import csv
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([c.get("label") or c["key"] for c in columns])
    for row in rows:
        line = []
        for col in columns:
            v = row.get(col["key"])
            t = col.get("type", "text")
            if v is None:
                line.append("")
            elif t == "money":
                try:
                    line.append(f"{float(v):.2f}")
                except (TypeError, ValueError):
                    line.append("0.00")
            else:
                line.append(str(v))
        writer.writerow(line)
    return buf.getvalue().encode("utf-8-sig")


def xlsx_response(content: bytes, filename: str):
    from fastapi.responses import Response as FastAPIResponse
    return FastAPIResponse(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def csv_response(content: bytes, filename: str):
    from fastapi.responses import Response as FastAPIResponse
    return FastAPIResponse(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
