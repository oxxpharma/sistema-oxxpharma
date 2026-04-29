"""Iteration 18 tests: XLSX export, MP credentials in DB, Correios shipping."""
import os
import io
import requests
import pytest
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://oxx-franchise-system.preview.emergentagent.com").rstrip("/")
ADMIN_EMAIL = "admin@oxxpharma.com"
ADMIN_PASSWORD = "admin123"


@pytest.fixture(scope="module")
def admin_token():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}, timeout=20)
    assert r.status_code == 200, f"Admin login failed: {r.status_code} {r.text}"
    data = r.json()
    tok = data.get("access_token") or data.get("token")
    assert tok, f"No token in login response: {data}"
    return tok


@pytest.fixture
def admin_client(admin_token):
    s = requests.Session()
    s.headers.update({"Authorization": f"Bearer {admin_token}", "Content-Type": "application/json"})
    return s


# ==================== Payments config in DB ====================

class TestPaymentsConfig:
    def test_get_admin_payments_config_masked(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/admin/payments-config", timeout=20)
        assert r.status_code == 200, r.text
        d = r.json()
        for key in ["mp_environment", "test_public_key", "test_access_token_masked", "test_configured",
                    "prod_public_key", "prod_access_token_masked", "webhook_secret_masked"]:
            assert key in d, f"missing {key}"
        # If test_configured, masked must contain '...' or be * - i.e., not full token
        if d["test_configured"] and d["test_access_token_masked"]:
            assert "..." in d["test_access_token_masked"] or "*" in d["test_access_token_masked"]

    def test_put_payments_config_persists_and_db_takes_precedence(self, admin_client):
        # Save original test public key & access token to restore
        orig = admin_client.get(f"{BASE_URL}/api/admin/payments-config").json()
        new_pub = "TEST-PUBKEY-DBOVERRIDE-12345"
        new_tok = "TEST-DB-OVERRIDE-ACCESS-TOKEN-ABCDEFGHIJKLMNOP"
        try:
            r = admin_client.put(f"{BASE_URL}/api/admin/payments-config", json={
                "mp_test_public_key": new_pub,
                "mp_test_access_token": new_tok,
                "mp_environment": "test",
            }, timeout=20)
            assert r.status_code == 200, r.text
            d = r.json()
            assert d["test_public_key"] == new_pub
            assert d["test_configured"] is True
            assert "..." in d["test_access_token_masked"]
            # Public config endpoint should reflect DB value
            pc = requests.get(f"{BASE_URL}/api/payments/config", timeout=20).json()
            assert pc.get("public_key") == new_pub
        finally:
            # Restore by clearing DB values to fall back to .env
            admin_client.put(f"{BASE_URL}/api/admin/payments-config", json={
                "mp_test_public_key": "",
                "mp_test_access_token": "",
            }, timeout=20)

    def test_put_invalid_environment(self, admin_client):
        r = admin_client.put(f"{BASE_URL}/api/admin/payments-config", json={"mp_environment": "staging"}, timeout=20)
        assert r.status_code == 400


# ==================== Correios config ====================

class TestCorreiosConfig:
    def test_get_default_config(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/admin/correios-config", timeout=20)
        assert r.status_code == 200
        d = r.json()
        for key in ["correios_enabled", "correios_origin_cep", "correios_services",
                    "correios_pickup_enabled", "correios_pickup_label", "correios_pickup_price",
                    "correios_default_length_cm", "correios_default_width_cm", "correios_default_height_cm",
                    "correios_min_weight_kg"]:
            assert key in d, f"missing {key}"
        assert isinstance(d["correios_services"], list)

    def test_update_config(self, admin_client):
        payload = {
            "correios_enabled": True,
            "correios_origin_cep": "01310100",
            "correios_services": [{"code": "04510", "label": "PAC"}, {"code": "04014", "label": "SEDEX"}],
            "correios_pickup_enabled": True,
            "correios_pickup_label": "Retirada Loja Centro",
            "correios_pickup_address": "Rua Teste, 100",
            "correios_pickup_price": 0.0,
            "correios_default_length_cm": 20,
            "correios_default_width_cm": 15,
            "correios_default_height_cm": 8,
        }
        r = admin_client.put(f"{BASE_URL}/api/admin/correios-config", json=payload, timeout=20)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["correios_enabled"] is True
        assert d["correios_origin_cep"] == "01310100"
        assert d["correios_pickup_enabled"] is True
        assert d["correios_pickup_label"] == "Retirada Loja Centro"
        assert d["correios_default_length_cm"] == 20

    def test_admin_correios_test(self, admin_client):
        r = admin_client.post(f"{BASE_URL}/api/admin/correios-test", json={"cep_destination": "20040002", "weight": 0.5}, timeout=30)
        assert r.status_code == 200, r.text
        d = r.json()
        assert "options" in d
        assert "package" in d

    def test_admin_correios_logs(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/admin/correios-logs", timeout=20)
        assert r.status_code == 200
        d = r.json()
        assert "logs" in d
        assert isinstance(d["logs"], list)


# ==================== Public shipping calculate ====================

class TestShippingCalculate:
    def test_calculate_with_pickup(self, admin_client):
        # Ensure pickup enabled (set in TestCorreiosConfig.test_update_config)
        admin_client.put(f"{BASE_URL}/api/admin/correios-config", json={
            "correios_enabled": True,
            "correios_pickup_enabled": True,
            "correios_pickup_label": "Retirada Loja",
            "correios_pickup_price": 0.0,
            "correios_origin_cep": "01310100",
        }, timeout=20)
        r = requests.post(f"{BASE_URL}/api/shipping/calculate",
                          json={"cep_destination": "20040002", "items": [{"weight": 0.4, "quantity": 1}]},
                          timeout=30)
        assert r.status_code == 200, r.text
        d = r.json()
        assert "options" in d
        codes = [o.get("code") for o in d["options"]]
        assert "PICKUP" in codes
        assert d["package"]["weight_kg"] >= 0.3

    def test_calculate_invalid_cep(self):
        r = requests.post(f"{BASE_URL}/api/shipping/calculate",
                          json={"cep_destination": "123", "items": [{"weight": 0.4, "quantity": 1}]},
                          timeout=20)
        assert r.status_code == 200
        d = r.json()
        assert "error" in d

    def test_calculate_disabled_no_pickup(self, admin_client):
        admin_client.put(f"{BASE_URL}/api/admin/correios-config", json={
            "correios_enabled": False, "correios_pickup_enabled": False,
        }, timeout=20)
        try:
            r = requests.post(f"{BASE_URL}/api/shipping/calculate",
                              json={"cep_destination": "20040002", "items": [{"weight": 0.5, "quantity": 1}]},
                              timeout=20)
            assert r.status_code == 200
            d = r.json()
            assert d["options"] == []
            assert d["package"] is not None
        finally:
            # restore
            admin_client.put(f"{BASE_URL}/api/admin/correios-config", json={
                "correios_enabled": True, "correios_pickup_enabled": True,
                "correios_origin_cep": "01310100",
            }, timeout=20)


# ==================== Product dimensions ====================

class TestProductDimensions:
    def test_product_create_with_dimensions(self, admin_client):
        payload = {
            "name": "TEST_it18 Produto Frete",
            "description": "Teste dimensoes",
            "price": 50.0,
            "stock": 10,
            "category": "test",
            "weight": 0.7,
            "length_cm": 25,
            "width_cm": 15,
            "height_cm": 10,
        }
        r = admin_client.post(f"{BASE_URL}/api/admin/products", json=payload, timeout=20)
        assert r.status_code in (200, 201), r.text
        prod = r.json()
        pid = prod.get("product_id") or prod.get("id")
        assert pid
        try:
            r2 = requests.get(f"{BASE_URL}/api/products/{pid}", timeout=20)
            assert r2.status_code == 200
            body = r2.json()
            p = body.get("product") if isinstance(body, dict) and "product" in body else body
            assert p.get("weight") == 0.7
            assert p.get("length_cm") == 25
            assert p.get("width_cm") == 15
            assert p.get("height_cm") == 10
        finally:
            admin_client.delete(f"{BASE_URL}/api/admin/products/{pid}", timeout=20)


# ==================== Points XLSX export ====================

class TestPointsXlsx:
    def test_export_xlsx(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/admin/points-report/export.xlsx", timeout=30)
        assert r.status_code == 200, r.text
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml.sheet" in ct, f"unexpected content-type: {ct}"
        # Load with openpyxl
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        assert headers == ["Data/Hora", "ID", "Nome", "Pontos totais"]
        # Freeze panes on row 2 (header fixed)
        assert ws.freeze_panes in ("A2", "$A$2")
        # auto_filter applied if data exists
        if ws.max_row > 1:
            assert ws.auto_filter.ref is not None
