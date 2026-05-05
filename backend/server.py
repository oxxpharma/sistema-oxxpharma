"""
OxxPharma - E-commerce MVP (Fase 1)
Backend API: Auth, Products, Categories, Cart, Checkout, Orders, Addresses, Admin
"""

import os
import uuid
import asyncio
import bcrypt
import logging
from datetime import datetime, timezone, timedelta
from typing import Optional, List, Dict
from contextlib import asynccontextmanager

from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, HTTPException, Depends, Request, Response, Query, BackgroundTasks, Header
from fastapi.responses import RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr
from motor.motor_asyncio import AsyncIOMotorClient
import jwt

import email_service
import card_service
import payments_service
import correios_service
import maxx_service
import melhorenvio_service
import store_extras

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

MONGO_URL = os.environ.get("MONGO_URL")
DB_NAME = os.environ.get("DB_NAME")
JWT_SECRET = os.environ.get("JWT_SECRET")
JWT_ALGORITHM = "HS256"

# ==================== HELPERS ====================

def gen_id(prefix=""):
    return f"{prefix}{uuid.uuid4().hex[:12]}"

def gen_referral_code() -> str:
    return uuid.uuid4().hex[:8].upper()

# Comissao por afiliado sobre vendas via link de indicacao (default - pode ser sobrescrito via settings)
AFFILIATE_COMMISSION_RATE = 0.08

# Tipos de rede MMN
NETWORK_CUSTOMER = "customer"      # Cliente comum - so 8% afiliado, sem MMN
NETWORK_1 = "network_1"            # Rede 1: importada do sistema externo
NETWORK_2 = "network_2"            # Rede 2: Propagandista promovido organicamente

DEFAULT_SETTINGS = {
    "affiliate_commission_rate": 0.08,
    # Percentuais por geracao em % (ex: 5.0 = 5%)
    "network1_generations": [5.0, 3.0, 2.0, 1.0, 1.0, 0.5],
    "network2_generations": [5.0, 3.0, 2.0, 1.0, 1.0, 0.5],
    "propaganda_threshold_referrals": 5,       # minimo de indicacoes/mes para virar "em alta"
    "propaganda_threshold_period_days": 30,     # janela de analise
    "withdrawal_enabled": False,                # saques ativos/desativados
    "withdrawal_min_amount": 50.0,              # valor minimo de saque
    "withdrawal_release_days": 15,              # dias apos pagamento do pedido para liberar para saque
    # Dados da empresa (para notas de faturamento)
    "company_name": "OxxPharma Farmacia Ltda",
    "company_cnpj": "",
    "company_address": "",
    "company_city": "",
    "company_state": "",
    "company_zip": "",
    "company_phone": "",
    "company_email": "contato@oxxpharma.com",
    "invoice_prefix": "OXX",
    "invoice_counter": 0,                       # incrementado a cada emissao
    # Email (Resend)
    "email_enabled": False,
    "resend_api_key": "",
    "email_from": "OxxPharma <onboarding@resend.dev>",
    "email_admin_recipients": "",               # lista de emails (virgula) que recebem alertas admin
    # Gatilhos on/off (admin controla granularmente)
    "email_trigger_order_created": True,
    "email_trigger_order_paid": True,
    "email_trigger_order_shipped": True,
    "email_trigger_order_delivered": True,
    "email_trigger_commission_earned": True,
    "email_trigger_admin_new_candidate": True,
    "email_trigger_admin_new_order": True,
    # Email que recebe a fatura detalhada de cada pedido PAGO (vazio = desabilitado)
    "order_invoice_email_to": "",
    "email_trigger_welcome": True,
    # Webhook inbound Rede 1
    "external_webhook_token": "",               # gerado no seed
}

async def get_settings(db):
    s = await db.settings.find_one({"_id": "global"})
    if not s:
        s = {"_id": "global", **DEFAULT_SETTINGS, "updated_at": now_iso()}
        await db.settings.insert_one(s)
    # Garantir campos novos (merge com defaults)
    merged = {**DEFAULT_SETTINGS, **{k: v for k, v in s.items() if k != "_id"}}
    return merged


def get_app_url():
    """URL pública do app (usado em emails, links, redirects)."""
    url = os.environ.get("APP_URL") or os.environ.get("BACKEND_URL") or os.environ.get("REACT_APP_BACKEND_URL") or "http://localhost:3000"
    return url.rstrip("/")


def order_ctx(order: dict, user: dict):
    oid = order.get("order_id", "")
    app_url = get_app_url()
    return {
        "order": order,
        "user": user,
        "order_short_id": (oid[-8:].upper() if oid else ""),
        "order_link": f"{app_url}/pedido/{oid}",
        "referral_link": f"{app_url}/?ref={user.get('referral_code', '')}" if user.get('referral_code') else app_url,
    }


async def admin_recipients(db) -> List[str]:
    settings = await get_settings(db)
    raw = (settings.get("email_admin_recipients") or "").strip()
    if not raw:
        # Fallback: todos os usuarios com role=admin
        admins = await db.users.find({"role": "admin"}, {"_id": 0, "email": 1}).to_list(20)
        return [a.get("email") for a in admins if a.get("email")]
    return [x.strip() for x in raw.split(",") if x.strip()]

def hash_pw(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_pw(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_token(user_id: str, email: str, role: str = "customer") -> str:
    return jwt.encode({
        "sub": user_id, "email": email, "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "access"
    }, JWT_SECRET, algorithm=JWT_ALGORITHM)

def now_iso():
    return datetime.now(timezone.utc).isoformat()

async def get_current_user(request: Request):
    # Prioridade: header Authorization > cookie. Isso eh critico para impersonation,
    # onde o frontend troca o token no localStorage mas o cookie do admin permanece.
    auth = request.headers.get("Authorization", "")
    token = None
    if auth.startswith("Bearer "):
        token = auth[7:]
    if not token:
        token = request.cookies.get("access_token")
    if not token:
        raise HTTPException(status_code=401, detail="Nao autenticado")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await request.app.db.users.find_one({"user_id": payload["sub"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario nao encontrado")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalido")

async def get_optional_user(request: Request):
    try:
        return await get_current_user(request)
    except Exception:
        return None

def require_admin():
    """Admin-level: qualquer role com acesso ao /backoffice (super_admin, admin, financeiro, comercial)."""
    async def dep(user: dict = Depends(get_current_user)):
        if not _is_admin_level(user):
            raise HTTPException(status_code=403, detail="Acesso negado")
        return user
    return dep


ADMIN_ROLES = {"admin", "super_admin", "financeiro", "comercial"}


def _is_admin_level(user: dict) -> bool:
    """Retorna True se o user tem acesso ao backoffice (qualquer role admin)."""
    r = user.get("role")
    return r in ADMIN_ROLES or user.get("access_level", 99) <= 1


def _role_of(user: dict) -> str:
    """Normaliza a role do usuario. Legacy 'admin' continua como super_admin a menos
    que tenha role explicito diferente."""
    return user.get("role") or "customer"


def require_role(*allowed_roles):
    """Restringe a um conjunto especifico de roles. Super_admin sempre passa."""
    async def dep(user: dict = Depends(get_current_user)):
        r = _role_of(user)
        if r == "super_admin" or r == "admin" or r in allowed_roles:
            # 'admin' legado = super_admin
            if r in ("super_admin", "admin") or r in allowed_roles:
                return user
        raise HTTPException(status_code=403, detail="Acesso negado para este papel")
    return dep


def require_super_admin():
    """Strict: apenas role='super_admin'. Admin novo NAO entra."""
    async def dep(user: dict = Depends(get_current_user)):
        r = _role_of(user)
        if r != "super_admin":
            raise HTTPException(status_code=403, detail="Apenas Super Admin")
        return user
    return dep

def set_cookie(response: Response, token: str):
    response.set_cookie(key="access_token", value=token, httponly=True, secure=False, samesite="lax", max_age=7*24*60*60, path="/")

# ==================== MODELS ====================

class AuthRegister(BaseModel):
    email: EmailStr
    password: str
    name: str
    phone: Optional[str] = None
    cpf: Optional[str] = None
    sponsor_code: Optional[str] = None  # Codigo de indicacao (referral do afiliado)

class AuthLogin(BaseModel):
    email: EmailStr
    password: str

class AddressCreate(BaseModel):
    label: Optional[str] = "Casa"
    street: str
    number: str
    complement: Optional[str] = None
    neighborhood: str
    city: str
    state: str
    zip_code: str
    is_default: bool = False

class ProductCreate(BaseModel):
    name: str
    description: str
    price: float
    discount_price: Optional[float] = None
    category: str
    subcategory: Optional[str] = None
    images: List[str] = []
    stock: int = 0
    active: bool = True
    featured: bool = False
    brand: Optional[str] = None
    weight: Optional[float] = None
    length_cm: Optional[float] = None
    width_cm: Optional[float] = None
    height_cm: Optional[float] = None
    points_value: float = 0  # Pontos atribuidos por unidade comprada (manual pelo admin)
    pricing_tiers: List[Dict] = []  # [{type:'guest'|'logged'|'category', user_category_id?, price, label?}]

class CategoryCreate(BaseModel):
    name: str
    description: Optional[str] = None
    image_url: Optional[str] = None
    parent: Optional[str] = None
    order: int = 0
    active: bool = True

class CartItemAdd(BaseModel):
    product_id: str
    quantity: int = 1

class CheckoutData(BaseModel):
    address_id: str
    payment_method: str = "pix"  # pix | credit_card | boleto (MVP: mock)
    notes: Optional[str] = None
    ref_code: Optional[str] = None  # Codigo de indicacao do afiliado (URL ?ref=XXX)
    coupon_code: Optional[str] = None  # Cupom aplicado no checkout
    shipping_price: Optional[float] = None   # valor escolhido na cotacao
    shipping_service_name: Optional[str] = None
    shipping_carrier: Optional[str] = None
    shipping_service_id: Optional[str] = None
    shipping_delivery_days: Optional[int] = None
    voucher_amount: Optional[float] = None  # Iter 36: usa saldo de voucher pre-pago (parcial ou total)

class WithdrawalCreate(BaseModel):
    amount: float
    pix_key: str
    pix_key_type: str  # cpf | email | phone | random
    pix_name: str
    pix_cpf: str
    notes: Optional[str] = None

# ==================== LIFESPAN ====================

async def seed_admin(db):
    email = os.environ.get("ADMIN_EMAIL", "admin@oxxpharma.com")
    pw = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": email})
    if not existing:
        await db.users.insert_one({
            "user_id": gen_id("user_"), "email": email, "password_hash": hash_pw(pw),
            "name": "Administrador OxxPharma", "phone": None, "role": "super_admin",
            "access_level": 0, "status": "active", "addresses": [],
            "referral_code": gen_referral_code(),
            "sponsor_id": None, "sponsor_code": None,
            "network_type": NETWORK_CUSTOMER,
            "network_sponsor_id": None,
            "external_id": None,
            "created_at": now_iso(),
        })
        logger.info(f"Admin criado: {email}")
    else:
        update = {}
        if not verify_pw(pw, existing.get("password_hash", "")):
            update["password_hash"] = hash_pw(pw)
        if not existing.get("referral_code"):
            update["referral_code"] = gen_referral_code()
        # Iter 38: admin seed sempre garante role=super_admin.
        # Usuarios com role legado 'admin' sao promovidos para super_admin.
        if existing.get("role") not in ("super_admin",):
            update["role"] = "super_admin"
            update["access_level"] = 0
        if update:
            await db.users.update_one({"email": email}, {"$set": update})

async def seed_categories(db):
    count = await db.categories.count_documents({})
    if count == 0:
        cats = [
            {"category_id": gen_id("cat_"), "name": "Medicamentos", "description": "Medicamentos em geral", "image_url": "", "parent": None, "order": 1, "active": True, "created_at": now_iso()},
            {"category_id": gen_id("cat_"), "name": "Dermocosmeticos", "description": "Cuidados com a pele", "image_url": "", "parent": None, "order": 2, "active": True, "created_at": now_iso()},
            {"category_id": gen_id("cat_"), "name": "Vitaminas e Suplementos", "description": "Vitaminas e suplementos alimentares", "image_url": "", "parent": None, "order": 3, "active": True, "created_at": now_iso()},
            {"category_id": gen_id("cat_"), "name": "Higiene Pessoal", "description": "Produtos de higiene", "image_url": "", "parent": None, "order": 4, "active": True, "created_at": now_iso()},
            {"category_id": gen_id("cat_"), "name": "Infantil", "description": "Produtos infantis", "image_url": "", "parent": None, "order": 5, "active": True, "created_at": now_iso()},
            {"category_id": gen_id("cat_"), "name": "Bem-estar", "description": "Produtos para bem-estar", "image_url": "", "parent": None, "order": 6, "active": True, "created_at": now_iso()},
        ]
        await db.categories.insert_many(cats)
        logger.info("Categorias padrao criadas")

async def seed_products(db):
    count = await db.products.count_documents({})
    if count == 0:
        prods = [
            {"product_id": gen_id("prod_"), "name": "Vitamina C 1000mg", "description": "Vitamina C efervescente com 10 comprimidos. Reforco para imunidade.", "price": 29.90, "discount_price": 22.90, "category": "Vitaminas e Suplementos", "images": ["https://images.unsplash.com/photo-1584308666744-24d5c474f2ae?w=400"], "stock": 150, "active": True, "featured": True, "brand": "OxxVita", "created_at": now_iso()},
            {"product_id": gen_id("prod_"), "name": "Protetor Solar FPS 50", "description": "Protetor solar facial e corporal com alta protecao. Textura leve.", "price": 89.90, "discount_price": 69.90, "category": "Dermocosmeticos", "images": ["https://images.unsplash.com/photo-1556228578-0d85b1a4d571?w=400"], "stock": 80, "active": True, "featured": True, "brand": "OxxDerm", "created_at": now_iso()},
            {"product_id": gen_id("prod_"), "name": "Omega 3 EPA DHA", "description": "Oleo de peixe concentrado com 60 capsulas. Saude cardiovascular.", "price": 59.90, "discount_price": 44.90, "category": "Vitaminas e Suplementos", "images": ["https://images.unsplash.com/photo-1550572017-edd951aa8f72?w=400"], "stock": 200, "active": True, "featured": True, "brand": "OxxVita", "created_at": now_iso()},
            {"product_id": gen_id("prod_"), "name": "Shampoo Anticaspa 200ml", "description": "Shampoo anticaspa de uso diario. Controle eficaz.", "price": 34.90, "discount_price": None, "category": "Higiene Pessoal", "images": ["https://images.unsplash.com/photo-1631729371254-42c2892f0e6e?w=400"], "stock": 120, "active": True, "featured": False, "brand": "OxxCare", "created_at": now_iso()},
            {"product_id": gen_id("prod_"), "name": "Creme Hidratante Corporal", "description": "Hidratante corporal com vitamina E. Pele macia e sedosa.", "price": 45.90, "discount_price": 35.90, "category": "Dermocosmeticos", "images": ["https://images.unsplash.com/photo-1608248543803-ba4f8c70ae0b?w=400"], "stock": 90, "active": True, "featured": True, "brand": "OxxDerm", "created_at": now_iso()},
            {"product_id": gen_id("prod_"), "name": "Multivitaminico A-Z", "description": "Complexo multivitaminico completo com 30 comprimidos.", "price": 39.90, "discount_price": 29.90, "category": "Vitaminas e Suplementos", "images": ["https://images.unsplash.com/photo-1559757175-7cb057fba93c?w=400"], "stock": 250, "active": True, "featured": False, "brand": "OxxVita", "created_at": now_iso()},
            {"product_id": gen_id("prod_"), "name": "Fralda Infantil P c/40", "description": "Fraldas descartaveis tamanho P com 40 unidades.", "price": 49.90, "discount_price": 39.90, "category": "Infantil", "images": ["https://images.unsplash.com/photo-1515488042361-ee00e0ddd4e4?w=400"], "stock": 60, "active": True, "featured": False, "brand": "OxxBaby", "created_at": now_iso()},
            {"product_id": gen_id("prod_"), "name": "Colageno Hidrolisado", "description": "Colageno em po com 300g. Pele, cabelos e unhas.", "price": 79.90, "discount_price": 59.90, "category": "Bem-estar", "images": ["https://images.unsplash.com/photo-1556228724-4e756cee32a3?w=400"], "stock": 100, "active": True, "featured": True, "brand": "OxxVita", "created_at": now_iso()},
        ]
        await db.products.insert_many(prods)
        logger.info("Produtos demo criados")

@asynccontextmanager
async def lifespan(app: FastAPI):
    app.mongodb_client = AsyncIOMotorClient(MONGO_URL)
    app.db = app.mongodb_client[DB_NAME]
    logger.info("Conectado ao MongoDB")
    await app.db.users.create_index("email", unique=True)
    await app.db.users.create_index("user_id", unique=True)
    await app.db.users.create_index("referral_code", unique=True, partialFilterExpression={"referral_code": {"$type": "string"}})
    await app.db.products.create_index("product_id", unique=True)
    await app.db.products.create_index("category")
    await app.db.orders.create_index("order_id", unique=True)
    await app.db.orders.create_index("user_id")
    await app.db.orders.create_index("invoice_number", unique=True, sparse=True)
    await app.db.categories.create_index("category_id", unique=True)
    await app.db.carts.create_index("user_id", unique=True)
    await app.db.commissions.create_index("commission_id", unique=True)
    await app.db.commissions.create_index("user_id")
    await app.db.commissions.create_index([("user_id", 1), ("status", 1)])
    await app.db.commissions.create_index("withdrawal_id")
    await app.db.users.create_index("network_type")
    await app.db.users.create_index("network_sponsor_id")
    await app.db.users.create_index("external_id", sparse=True)
    await app.db.users.create_index("leader_external_id", sparse=True)
    await app.db.users.create_index("cpf_digits", sparse=True)
    await app.db.withdrawals.create_index("withdrawal_id", unique=True)
    await app.db.withdrawals.create_index("user_id")
    await app.db.withdrawals.create_index([("status", 1), ("created_at", -1)])
    await app.db.email_templates.create_index("slug", unique=True)
    await app.db.email_templates.create_index("template_id", unique=True)
    await app.db.email_logs.create_index("log_id", unique=True)
    await app.db.email_logs.create_index("created_at")
    await app.db.webhook_logs.create_index("log_id", unique=True)
    await app.db.webhook_logs.create_index("created_at")
    await email_service.seed_default_templates(app.db)
    await app.db.card_batches.create_index("batch_id", unique=True)
    await app.db.card_batches.create_index("created_at")
    await app.db.card_api_logs.create_index("log_id", unique=True)
    await app.db.card_api_logs.create_index("created_at")
    # Iter 39: cliques no link de indicacao
    await app.db.referral_clicks.create_index("owner_user_id")
    await app.db.referral_clicks.create_index("created_at")
    # Migracao: se ainda nao rodou, reseta referral_codes -> None e marca referral_program_active=False
    migration = await app.db.migrations.find_one({"_id": "reset_referrals_for_card_program"})
    if not migration:
        await app.db.users.update_many(
            {"role": {"$ne": "admin"}},
            {
                "$unset": {"referral_code": ""},
                "$set": {
                    "referral_program_active": False,
                    "referral_enrollment": None,
                    "referral_enrolled_at": None,
                },
            },
        )
        # Admin continua com codigo (pode indicar normal)
        await app.db.users.update_many(
            {"role": "admin", "referral_program_active": {"$exists": False}},
            {"$set": {"referral_program_active": True}},
        )
        await app.db.migrations.insert_one({"_id": "reset_referrals_for_card_program", "ran_at": now_iso()})
        logger.info("Migracao reset_referrals_for_card_program aplicada")
    # Inicia scheduler do cartao
    card_service.start_scheduler(lambda: app.db)
    # Garantir token de webhook
    settings_doc = await app.db.settings.find_one({"_id": "global"}) or {}
    if not settings_doc.get("external_webhook_token"):
        await app.db.settings.update_one(
            {"_id": "global"},
            {"$set": {"external_webhook_token": gen_referral_code() + gen_referral_code()}},
            upsert=True,
        )
    await seed_admin(app.db)
    await seed_categories(app.db)
    await seed_products(app.db)
    yield
    app.mongodb_client.close()

app = FastAPI(title="OxxPharma E-commerce API", lifespan=lifespan)
# CORS: allow_origins=["*"] eh incompativel com allow_credentials=True (regra do navegador).
# Solucao: allow_origin_regex=".*" -> o middleware reflete o Origin do request na resposta.
app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=".*",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== AUTH ====================

@app.post("/api/auth/register")
async def register(request: Request, response: Response, data: AuthRegister):
    db = request.app.db
    if await db.users.find_one({"email": data.email.lower()}):
        raise HTTPException(status_code=400, detail="Email ja cadastrado")

    # Processar sponsor_code (indicacao) se fornecido
    sponsor_id = None
    sponsor_code_norm = None
    if data.sponsor_code:
        code = data.sponsor_code.strip().upper()
        sponsor = await db.users.find_one({"referral_code": code}, {"_id": 0})
        if sponsor:
            sponsor_id = sponsor["user_id"]
            sponsor_code_norm = code

    # referral_code so eh gerado quando usuario adere ao programa (nao mais no cadastro)
    user = {
        "user_id": gen_id("user_"), "email": data.email.lower(),
        "password_hash": hash_pw(data.password), "name": data.name,
        "phone": data.phone, "cpf": data.cpf, "cpf_digits": _clean_cpf(data.cpf),
        "role": "customer", "access_level": 99,
        "status": "active", "addresses": [],
        "referral_code": None,
        "referral_program_active": False,
        "referral_enrollment": None,
        "referral_enrolled_at": None,
        "sponsor_id": sponsor_id, "sponsor_code": sponsor_code_norm,
        "network_type": NETWORK_CUSTOMER,
        "network_sponsor_id": None,
        "external_id": None,
        "created_at": now_iso(),
    }
    await db.users.insert_one(user)
    token = create_token(user["user_id"], user["email"], "customer")
    set_cookie(response, token)
    u = await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0, "password_hash": 0})
    # Email de boas-vindas (async, nao bloqueia)
    app_url = get_app_url()
    asyncio.create_task(email_service.trigger(db, "welcome", u["email"], {
        "user": u,
        "referral_link": f"{app_url}/?ref={u.get('referral_code','')}",
    }))
    # Se tem sponsor customer, checa se virou candidato a Propagandista
    if sponsor_id:
        sponsor = await db.users.find_one({"user_id": sponsor_id}, {"_id": 0, "password_hash": 0})
        if sponsor and sponsor.get("network_type") == NETWORK_CUSTOMER:
            settings = await get_settings(db)
            threshold = int(settings.get("propaganda_threshold_referrals") or 5)
            period_days = int(settings.get("propaganda_threshold_period_days") or 30)
            since = (datetime.now(timezone.utc) - timedelta(days=period_days)).isoformat()
            count = await db.users.count_documents({"sponsor_id": sponsor_id, "created_at": {"$gte": since}})
            # Notificar apenas no momento em que CRUZA o threshold (evita spam)
            if count == threshold:
                admins = await admin_recipients(db)
                if admins:
                    url = f"{app_url}/backoffice/candidatos"
                    asyncio.create_task(email_service.trigger(db, "admin_new_candidate", admins, {
                        "candidate": {**sponsor, "referrals_in_period": count},
                        "period_days": period_days,
                        "admin_link": url,
                    }))
    return {"token": token, "user": u}

@app.post("/api/auth/login")
async def login(request: Request, response: Response, data: AuthLogin):
    db = request.app.db
    user = await db.users.find_one({"email": data.email.lower()}, {"_id": 0})
    if not user or not verify_pw(data.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Credenciais invalidas")
    if user.get("status") in ("cancelled", "inactive", "deleted"):
        raise HTTPException(status_code=401, detail="Conta desativada")
    role = user.get("role", "customer")
    if user.get("access_level", 99) <= 1 and role not in ADMIN_ROLES:
        role = "super_admin"
    token = create_token(user["user_id"], user["email"], role)
    set_cookie(response, token)
    user.pop("password_hash", None)
    return {"token": token, "user": user}

@app.get("/api/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    u = dict(user)
    u.pop("password_hash", None)
    return u

@app.post("/api/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"message": "Deslogado"}

# ==================== USER PROFILE & ADDRESSES ====================

@app.put("/api/users/me")
async def update_profile(request: Request, user: dict = Depends(get_current_user)):
    db = request.app.db
    body = await request.json()
    update = {}
    for field in ["name", "phone", "cpf", "pix_key", "pix_key_type"]:
        if field in body:
            update[field] = body[field]
    if update:
        await db.users.update_one({"user_id": user["user_id"]}, {"$set": update})
    u = await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0, "password_hash": 0})
    return u

@app.get("/api/users/me/addresses")
async def list_addresses(user: dict = Depends(get_current_user)):
    return {"addresses": user.get("addresses", [])}

@app.post("/api/users/me/addresses")
async def add_address(request: Request, data: AddressCreate, user: dict = Depends(get_current_user)):
    db = request.app.db
    addr = data.model_dump()
    addr["address_id"] = gen_id("addr_")
    addrs = user.get("addresses", [])
    if data.is_default:
        for a in addrs:
            a["is_default"] = False
    addrs.append(addr)
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"addresses": addrs}})
    return {"addresses": addrs}

@app.put("/api/users/me/addresses/{address_id}")
async def update_address(request: Request, address_id: str, data: AddressCreate, user: dict = Depends(get_current_user)):
    db = request.app.db
    addrs = user.get("addresses", [])
    found = False
    for i, a in enumerate(addrs):
        if a.get("address_id") == address_id:
            new_addr = data.model_dump()
            new_addr["address_id"] = address_id
            if data.is_default:
                for x in addrs:
                    x["is_default"] = False
            addrs[i] = new_addr
            found = True
            break
    if not found:
        raise HTTPException(status_code=404, detail="Endereco nao encontrado")
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"addresses": addrs}})
    return {"addresses": addrs}

@app.delete("/api/users/me/addresses/{address_id}")
async def delete_address(request: Request, address_id: str, user: dict = Depends(get_current_user)):
    db = request.app.db
    addrs = [a for a in user.get("addresses", []) if a.get("address_id") != address_id]
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"addresses": addrs}})
    return {"addresses": addrs}

# ==================== CATEGORIES ====================

@app.get("/api/categories")
async def list_categories(request: Request):
    db = request.app.db
    cats = await db.categories.find({"active": True}, {"_id": 0}).sort("order", 1).to_list(100)
    return {"categories": cats}

@app.post("/api/admin/categories")
async def create_category(request: Request, data: CategoryCreate, user: dict = Depends(require_admin())):
    db = request.app.db
    cat = {"category_id": gen_id("cat_"), **data.model_dump(), "created_at": now_iso()}
    await db.categories.insert_one(cat)
    return await db.categories.find_one({"category_id": cat["category_id"]}, {"_id": 0})

@app.put("/api/admin/categories/{category_id}")
async def update_category(request: Request, category_id: str, data: CategoryCreate, user: dict = Depends(require_admin())):
    db = request.app.db
    update = data.model_dump()
    update["updated_at"] = now_iso()
    r = await db.categories.update_one({"category_id": category_id}, {"$set": update})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Categoria nao encontrada")
    return await db.categories.find_one({"category_id": category_id}, {"_id": 0})

@app.delete("/api/admin/categories/{category_id}")
async def delete_category(request: Request, category_id: str, user: dict = Depends(require_admin())):
    db = request.app.db
    r = await db.categories.delete_one({"category_id": category_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Categoria nao encontrada")
    return {"message": "Categoria removida"}

# ==================== PRODUCTS (PUBLIC) ====================

@app.get("/api/products")
async def list_products(request: Request, category: Optional[str] = None, search: Optional[str] = None, featured: Optional[bool] = None, page: int = 1, limit: int = 20):
    db = request.app.db
    q = {"active": True}
    if category:
        q["category"] = {"$regex": f"^{category}$", "$options": "i"}
    if featured:
        q["featured"] = True
    if search:
        q["$or"] = [{"name": {"$regex": search, "$options": "i"}}, {"description": {"$regex": search, "$options": "i"}}, {"brand": {"$regex": search, "$options": "i"}}]
    total = await db.products.count_documents(q)
    products = await db.products.find(q, {"_id": 0}).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    user = await get_optional_user(request)
    products = [store_extras.apply_pricing_to_product(p, user) for p in products]
    return {"products": products, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit)}

@app.get("/api/products/featured")
async def featured_products(request: Request, limit: int = 8):
    db = request.app.db
    prods = await db.products.find({"active": True, "featured": True}, {"_id": 0}).limit(limit).to_list(limit)
    if len(prods) < limit:
        more = await db.products.find({"active": True, "featured": {"$ne": True}}, {"_id": 0}).limit(limit - len(prods)).to_list(limit - len(prods))
        prods.extend(more)
    user = await get_optional_user(request)
    prods = [store_extras.apply_pricing_to_product(p, user) for p in prods]
    return {"products": prods}

@app.get("/api/products/{product_id}")
async def get_product(request: Request, product_id: str):
    db = request.app.db
    p = await db.products.find_one({"product_id": product_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Produto nao encontrado")
    related = await db.products.find({"category": p.get("category"), "product_id": {"$ne": product_id}, "active": True}, {"_id": 0}).limit(4).to_list(4)
    user = await get_optional_user(request)
    p = store_extras.apply_pricing_to_product(p, user)
    related = [store_extras.apply_pricing_to_product(r, user) for r in related]
    return {"product": p, "related": related}

# ==================== PRODUCTS (ADMIN) ====================

@app.get("/api/admin/products")
async def admin_list_products(request: Request, category: Optional[str] = None, search: Optional[str] = None, page: int = 1, limit: int = 20, user: dict = Depends(require_admin())):
    db = request.app.db
    q = {}
    if category:
        q["category"] = category
    if search:
        q["$or"] = [{"name": {"$regex": search, "$options": "i"}}, {"description": {"$regex": search, "$options": "i"}}]
    total = await db.products.count_documents(q)
    products = await db.products.find(q, {"_id": 0}).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    return {"products": products, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit)}

@app.post("/api/admin/products")
async def create_product(request: Request, data: ProductCreate, user: dict = Depends(require_admin())):
    db = request.app.db
    prod = {"product_id": gen_id("prod_"), **data.model_dump(), "created_at": now_iso()}
    await db.products.insert_one(prod)
    return await db.products.find_one({"product_id": prod["product_id"]}, {"_id": 0})

@app.put("/api/admin/products/{product_id}")
async def update_product(request: Request, product_id: str, data: ProductCreate, user: dict = Depends(require_admin())):
    db = request.app.db
    update = data.model_dump()
    update["updated_at"] = now_iso()
    r = await db.products.update_one({"product_id": product_id}, {"$set": update})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Produto nao encontrado")
    return await db.products.find_one({"product_id": product_id}, {"_id": 0})

@app.delete("/api/admin/products/{product_id}")
async def delete_product(request: Request, product_id: str, user: dict = Depends(require_admin())):
    db = request.app.db
    r = await db.products.delete_one({"product_id": product_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Produto nao encontrado")
    return {"message": "Produto removido"}

# ==================== CART ====================

@app.get("/api/cart")
async def get_cart(request: Request, user: dict = Depends(get_current_user)):
    db = request.app.db
    cart = await db.carts.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not cart:
        return {"items": [], "subtotal": 0, "count": 0}
    items = cart.get("items", [])
    subtotal = 0
    enriched = []
    for item in items:
        prod = await db.products.find_one({"product_id": item["product_id"]}, {"_id": 0})
        if prod:
            price_info = store_extras.effective_price(prod, user)
            price = float(price_info["price"])
            original = float(price_info["original_price"])
            total = price * item["quantity"]
            subtotal += total
            enriched.append({**item, "name": prod["name"], "price": price, "original_price": original, "tier_applied": price_info.get("applied_tier"), "image": (prod.get("images") or [None])[0], "total": total, "stock": prod.get("stock", 0)})
    return {"items": enriched, "subtotal": round(subtotal, 2), "count": len(enriched)}

@app.post("/api/cart/items")
async def add_to_cart(request: Request, data: CartItemAdd, user: dict = Depends(get_current_user)):
    db = request.app.db
    prod = await db.products.find_one({"product_id": data.product_id, "active": True}, {"_id": 0})
    if not prod:
        raise HTTPException(status_code=404, detail="Produto nao encontrado")
    if prod.get("stock", 0) < data.quantity:
        raise HTTPException(status_code=400, detail="Estoque insuficiente")
    cart = await db.carts.find_one({"user_id": user["user_id"]})
    if not cart:
        await db.carts.insert_one({"user_id": user["user_id"], "items": [{"product_id": data.product_id, "quantity": data.quantity}], "updated_at": now_iso()})
    else:
        items = cart.get("items", [])
        found = False
        for i in items:
            if i["product_id"] == data.product_id:
                i["quantity"] += data.quantity
                found = True
                break
        if not found:
            items.append({"product_id": data.product_id, "quantity": data.quantity})
        await db.carts.update_one({"user_id": user["user_id"]}, {"$set": {"items": items, "updated_at": now_iso()}})
    return await get_cart(request, user)

@app.put("/api/cart/items/{product_id}")
async def update_cart_item(request: Request, product_id: str, user: dict = Depends(get_current_user)):
    db = request.app.db
    body = await request.json()
    qty = body.get("quantity", 1)
    if qty <= 0:
        return await remove_from_cart(request, product_id, user)
    prod = await db.products.find_one({"product_id": product_id}, {"_id": 0})
    if prod and qty > prod.get("stock", 0):
        raise HTTPException(status_code=400, detail="Estoque insuficiente")
    cart = await db.carts.find_one({"user_id": user["user_id"]})
    if cart:
        items = cart.get("items", [])
        for i in items:
            if i["product_id"] == product_id:
                i["quantity"] = qty
                break
        await db.carts.update_one({"user_id": user["user_id"]}, {"$set": {"items": items, "updated_at": now_iso()}})
    return await get_cart(request, user)

@app.delete("/api/cart/items/{product_id}")
async def remove_from_cart(request: Request, product_id: str, user: dict = Depends(get_current_user)):
    db = request.app.db
    cart = await db.carts.find_one({"user_id": user["user_id"]})
    if cart:
        items = [i for i in cart.get("items", []) if i["product_id"] != product_id]
        await db.carts.update_one({"user_id": user["user_id"]}, {"$set": {"items": items, "updated_at": now_iso()}})
    return await get_cart(request, user)

# ==================== CHECKOUT & ORDERS ====================

@app.post("/api/checkout")
async def checkout(request: Request, data: CheckoutData, user: dict = Depends(get_current_user)):
    db = request.app.db
    # Get cart
    cart = await db.carts.find_one({"user_id": user["user_id"]})
    if not cart or not cart.get("items"):
        raise HTTPException(status_code=400, detail="Carrinho vazio")
    # Get address
    addrs = user.get("addresses", [])
    addr = next((a for a in addrs if a.get("address_id") == data.address_id), None)
    if not addr:
        raise HTTPException(status_code=400, detail="Endereco nao encontrado")
    # Build order items
    items = []
    subtotal = 0
    for ci in cart["items"]:
        prod = await db.products.find_one({"product_id": ci["product_id"], "active": True}, {"_id": 0})
        if not prod:
            continue
        if prod.get("stock", 0) < ci["quantity"]:
            raise HTTPException(status_code=400, detail=f"Estoque insuficiente: {prod['name']}")
        # preco efetivo considerando pricing_tiers do usuario
        price_info = store_extras.effective_price(prod, user)
        price = float(price_info["price"])
        total = round(price * ci["quantity"], 2)
        subtotal += total
        items.append({"product_id": prod["product_id"], "name": prod["name"], "price": price, "quantity": ci["quantity"], "total": total, "image": (prod.get("images") or [None])[0], "points_value": float(prod.get("points_value") or 0), "tier_applied": price_info.get("applied_tier")})
        # Decrement stock
        await db.products.update_one({"product_id": prod["product_id"]}, {"$inc": {"stock": -ci["quantity"]}})
    if not items:
        raise HTTPException(status_code=400, detail="Nenhum produto valido")
    # ============ Frete ============
    # Preferir o valor cotado pelo frontend (cliente escolheu uma opção).
    # Se nao vier, recotar server-side para nao cobrar zero indevidamente.
    shipping = None
    shipping_meta = {}
    if data.shipping_price is not None and data.shipping_price >= 0:
        shipping = float(data.shipping_price)
        shipping_meta = {
            "shipping_service_name": data.shipping_service_name,
            "shipping_carrier": data.shipping_carrier,
            "shipping_service_id": data.shipping_service_id,
            "shipping_delivery_days": data.shipping_delivery_days,
        }
    else:
        # Recotar no provider ativo usando CEP do endereco escolhido
        cep_dest = (addr.get("zip_code") or "").replace("-", "").replace(".", "").strip()
        calc_items = []
        for it in items:
            prod = await db.products.find_one({"product_id": it["product_id"]}, {"_id": 0})
            calc_items.append({
                "weight": prod.get("weight") or 0.3 if prod else 0.3,
                "length_cm": prod.get("length_cm") if prod else None,
                "width_cm": prod.get("width_cm") if prod else None,
                "height_cm": prod.get("height_cm") if prod else None,
                "quantity": it["quantity"],
            })
        fs_settings_tmp = await _get_site_settings(db)
        provider = (fs_settings_tmp.get("shipping_provider") or "correios").lower()
        try:
            if provider == "melhorenvio":
                me_items = [{
                    "weight_kg": i.get("weight") or 0.3,
                    "length_cm": i.get("length_cm"), "width_cm": i.get("width_cm"), "height_cm": i.get("height_cm"),
                    "quantity": i.get("quantity", 1),
                    "insurance_value": (subtotal / max(len(calc_items), 1)),
                } for i in calc_items]
                calc = await melhorenvio_service.calculate_shipping(db, cep_dest, me_items, insurance_value=subtotal)
            else:
                calc = await correios_service.calculate_freight(db, cep_dest, calc_items, subtotal)
            opts = (calc or {}).get("options") or []
            if opts:
                best = min(opts, key=lambda o: float(o.get("price") or 0))
                shipping = float(best.get("price") or 0)
                shipping_meta = {
                    "shipping_service_name": best.get("service_name") or best.get("name"),
                    "shipping_carrier": best.get("company_name") or best.get("carrier"),
                    "shipping_service_id": best.get("service_id") or best.get("code"),
                    "shipping_delivery_days": best.get("delivery_days"),
                }
        except Exception:
            pass
    if shipping is None:
        shipping = 0.0  # Em ultimo caso, evita cobrar valor fantasma

    # Aplica frete grátis se configurado em site_settings (espelha regras publicas)
    fs_settings = await _get_site_settings(db)
    fs_mode = (fs_settings.get("free_shipping_mode") or "off").lower()
    fs_min = float(fs_settings.get("free_shipping_min_subtotal") or 0)
    fs_audiences = fs_settings.get("free_shipping_audiences") or []
    user_net = user.get("network_type") or "customer"
    user_cats = user.get("category_ids") or []
    matches_audience = (
        user_net in fs_audiences
        or any(isinstance(t, str) and t.startswith("cat:") and t[4:] in user_cats for t in fs_audiences)
    )
    if fs_mode == "all":
        shipping = 0.0
    elif fs_mode == "above" and subtotal >= fs_min and fs_min > 0:
        shipping = 0.0
    elif fs_mode == "audiences" and matches_audience and (fs_min <= 0 or subtotal >= fs_min):
        shipping = 0.0

    # Cupom (opcional)
    discount_amount = 0.0
    coupon_code_applied = None
    coupon_id_applied = None
    if getattr(data, "coupon_code", None):
        result = await store_extras.validate_coupon(db, data.coupon_code, round(subtotal, 2), user)
        if not result.get("valid"):
            raise HTTPException(status_code=400, detail=result.get("reason") or "Cupom inválido")
        discount_amount = float(result.get("discount") or 0)
        coupon_code_applied = result["coupon"]["code"]
        coupon_id_applied = result["coupon"]["coupon_id"]

    settings = await get_settings(db)
    affiliate_rate = settings["affiliate_commission_rate"]

    # Identificar afiliado (prioridade: sponsor_id do usuario, senao ref_code do checkout)
    affiliate_id = None
    affiliate_code = None
    if user.get("sponsor_id"):
        aff = await db.users.find_one({"user_id": user["sponsor_id"]}, {"_id": 0})
        if aff:
            affiliate_id = aff["user_id"]
            affiliate_code = aff.get("referral_code")
    elif data.ref_code:
        code = data.ref_code.strip().upper()
        aff = await db.users.find_one({"referral_code": code}, {"_id": 0})
        if aff and aff["user_id"] != user["user_id"]:
            affiliate_id = aff["user_id"]
            affiliate_code = code

    commission_amount = round(subtotal * affiliate_rate, 2) if affiliate_id else 0

    # Iter 36: Voucher pre-pago (deduzido do total)
    voucher_used = 0.0
    user_fresh = await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0, "voucher_balance": 1})
    user_voucher = float((user_fresh or {}).get("voucher_balance") or 0)
    grand_total = round(subtotal + shipping - discount_amount, 2)
    requested_voucher = float(data.voucher_amount or 0)
    if requested_voucher > 0:
        # Limita ao saldo do user E ao total do pedido
        voucher_used = round(min(requested_voucher, user_voucher, grand_total), 2)
        if voucher_used <= 0:
            voucher_used = 0.0
    final_total = round(grand_total - voucher_used, 2)

    order = {
        "order_id": gen_id("ord_"), "user_id": user["user_id"],
        "customer_name": user.get("name"), "customer_email": user.get("email"),
        "items": items, "subtotal": round(subtotal, 2),
        "shipping_cost": shipping,
        **shipping_meta,
        "discount_amount": round(discount_amount, 2),
        "coupon_code": coupon_code_applied,
        "voucher_used": voucher_used,
        "voucher_balance_before": round(user_voucher, 2),
        "total": final_total,
        "total_before_voucher": grand_total,
        "shipping_address": addr, "payment_method": data.payment_method,
        "payment_status": "pending", "order_status": "pending",
        "payment_provider": "mock",  # Sera "mercadopago" quando integrado
        "payment_id": None, "payment_url": None,
        "affiliate_id": affiliate_id, "affiliate_code": affiliate_code,
        "affiliate_commission": commission_amount,
        "notes": data.notes, "created_at": now_iso(),
    }
    await db.orders.insert_one(order)

    # Iter 36: Debita o voucher do user agora (lock-in para evitar consumir 2x)
    if voucher_used > 0:
        await db.users.update_one(
            {"user_id": user["user_id"]},
            {"$inc": {"voucher_balance": -voucher_used},
             "$push": {"voucher_history": {
                 "delta": -voucher_used,
                 "source": "checkout",
                 "order_id": order["order_id"],
                 "received_at": now_iso(),
             }}},
        )

    # Atualiza usage_count do cupom (apos criar a ordem)
    if coupon_id_applied:
        await store_extras.increment_coupon_usage(db, coupon_id_applied)

    # ============ COMISSOES ============
    commissions_to_insert = []

    # 1) Comissao de afiliado (8% configuravel) - pago a quem indicou via link
    # Iter 40: sempre cria. Se afiliado nao esta inscrito no programa, status=pending_enrollment
    # e fica oculto ate ele se inscrever.
    if affiliate_id and commission_amount > 0:
        affiliate_user = await db.users.find_one({"user_id": affiliate_id}, {"_id": 0})
        aff_active = bool(affiliate_user and affiliate_user.get("referral_program_active"))
        commissions_to_insert.append({
            "commission_id": gen_id("com_"),
            "user_id": affiliate_id,
            "order_id": order["order_id"],
            "customer_id": user["user_id"],
            "customer_name": user.get("name"),
            "type": "affiliate",
            "network_type": None,
            "generation": 0,
            "amount": commission_amount,
            "rate": affiliate_rate,
            "order_subtotal": round(subtotal, 2),
            "status": "pending" if aff_active else "pending_enrollment",
            "program_active_at_creation": aff_active,
            "created_at": now_iso(),
        })

    # 2) Comissoes de rede MMN (ate 6 geracoes)
    # Iter 40: comissoes sao SEMPRE geradas (sem filtrar por programa de beneficios).
    # Beneficiarios nao inscritos ficam com status='pending_enrollment' e a comissao eh
    # liberada quando se inscrevem. A geracao incrementa a cada passo da cadeia,
    # mesmo quando o ancestral nao eh MMN (nao paga, mas "gasta" a geracao,
    # preservando a topologia da rede).
    network1_pcts = settings.get("network1_generations", []) or []
    network2_pcts = settings.get("network2_generations", []) or []
    current_id = user.get("sponsor_id") or user.get("network_sponsor_id")
    visited = {user["user_id"]}
    generation = 1
    while generation <= 6 and current_id and current_id not in visited:
        visited.add(current_id)
        ancestor = await db.users.find_one({"user_id": current_id}, {"_id": 0})
        if not ancestor:
            break
        a_net = ancestor.get("network_type")
        if a_net in (NETWORK_1, NETWORK_2):
            pcts = network1_pcts if a_net == NETWORK_1 else network2_pcts
            pct = pcts[generation - 1] if generation - 1 < len(pcts) else 0
            if pct > 0:
                amt = round(subtotal * pct / 100, 2)
                if amt > 0:
                    a_active = bool(ancestor.get("referral_program_active"))
                    commissions_to_insert.append({
                        "commission_id": gen_id("com_"),
                        "user_id": ancestor["user_id"],
                        "order_id": order["order_id"],
                        "customer_id": user["user_id"],
                        "customer_name": user.get("name"),
                        "type": "network_gen",
                        "network_type": a_net,
                        "generation": generation,
                        "amount": amt,
                        "rate": pct / 100,
                        "order_subtotal": round(subtotal, 2),
                        "status": "pending" if a_active else "pending_enrollment",
                        "program_active_at_creation": a_active,
                        "created_at": now_iso(),
                    })
        current_id = ancestor.get("sponsor_id") or ancestor.get("network_sponsor_id")
        generation += 1

    if commissions_to_insert:
        await db.commissions.insert_many(commissions_to_insert)

    # Clear cart
    await db.carts.delete_one({"user_id": user["user_id"]})
    final_order = await db.orders.find_one({"order_id": order["order_id"]}, {"_id": 0})

    # === EMAILS (async, nao bloqueia a response) ===
    ctx = order_ctx(final_order, user)
    # Pedido criado -> cliente
    asyncio.create_task(email_service.trigger(db, "order_created", user["email"], ctx))
    # Novo pedido -> admins
    admins = await admin_recipients(db)
    if admins:
        app_url = get_app_url()
        asyncio.create_task(email_service.trigger(db, "admin_new_order", admins, {
            **ctx, "items_count": len(items),
            "admin_link": f"{app_url}/backoffice/pedidos",
        }))
    # Comissoes de afiliado -> cada beneficiario
    for comm in commissions_to_insert:
        if comm.get("type") != "affiliate":
            continue
        receiver = await db.users.find_one({"user_id": comm["user_id"]}, {"_id": 0, "password_hash": 0})
        if not receiver or not receiver.get("email"):
            continue
        asyncio.create_task(email_service.trigger(db, "commission_earned", receiver["email"], {
            **order_ctx(final_order, receiver),
            "customer_name": user.get("name"),
            "commission": {
                "amount": f"{comm['amount']:.2f}",
                "rate_pct": f"{comm.get('rate', 0) * 100:.1f}",
            },
        }))

    return final_order

@app.get("/api/orders")
async def list_user_orders(request: Request, page: int = 1, limit: int = 10, user: dict = Depends(get_current_user)):
    db = request.app.db
    q = {"user_id": user["user_id"]}
    total = await db.orders.count_documents(q)
    orders = await db.orders.find(q, {"_id": 0}).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    return {"orders": orders, "total": total, "page": page}

@app.get("/api/orders/{order_id}")
async def get_order(request: Request, order_id: str, user: dict = Depends(get_current_user)):
    db = request.app.db
    o = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    if not o:
        raise HTTPException(status_code=404, detail="Pedido nao encontrado")
    is_admin = user.get("role") == "admin" or user.get("access_level", 99) <= 1
    if not is_admin and o.get("user_id") != user["user_id"]:
        raise HTTPException(status_code=403, detail="Acesso negado")
    return o

# ==================== ADMIN ORDERS ====================

@app.get("/api/admin/orders")
async def admin_list_orders(request: Request, status: Optional[str] = None, search: Optional[str] = None, page: int = 1, limit: int = 20, user: dict = Depends(require_admin())):
    db = request.app.db
    q = {}
    if status:
        q["order_status"] = status
    if search:
        q["$or"] = [{"order_id": {"$regex": search, "$options": "i"}}, {"customer_name": {"$regex": search, "$options": "i"}}, {"customer_email": {"$regex": search, "$options": "i"}}]
    total = await db.orders.count_documents(q)
    orders = await db.orders.find(q, {"_id": 0}).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    return {"orders": orders, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit)}

@app.put("/api/admin/orders/{order_id}/status")
async def admin_update_order_status(request: Request, order_id: str, user: dict = Depends(require_admin())):
    db = request.app.db
    body = await request.json()
    status = body.get("status")
    if not status:
        raise HTTPException(status_code=400, detail="Status obrigatorio")
    o = await db.orders.find_one({"order_id": order_id})
    if not o:
        raise HTTPException(status_code=404, detail="Pedido nao encontrado")
    update = {"order_status": status, "updated_at": now_iso()}
    if status == "paid":
        update["payment_status"] = "paid"
        update["paid_at"] = now_iso()
        # Auto-emitir nota de faturamento na 1a vez que pedido for marcado como pago
        if not o.get("invoice_number"):
            inv = await issue_invoice(db, o)
            update["invoice_number"] = inv["number"]
            update["invoice_issued_at"] = inv["issued_at"]
    elif status == "shipped":
        update["shipped_at"] = now_iso()
    elif status == "delivered":
        update["delivered_at"] = now_iso()
    elif status == "cancelled":
        update["cancelled_at"] = now_iso()
        # Restore stock
        for item in o.get("items", []):
            await db.products.update_one({"product_id": item["product_id"]}, {"$inc": {"stock": item["quantity"]}})
    await db.orders.update_one({"order_id": order_id}, {"$set": update})
    # Sincronizar status das comissoes do afiliado
    if status == "paid":
        await db.commissions.update_many(
            {"order_id": order_id, "status": "pending"},
            {"$set": {"status": "paid", "paid_at": now_iso()}},
        )
        # Registrar pontos
        await register_points_from_order(db, order_id)
    elif status == "cancelled":
        await db.commissions.update_many(
            {"order_id": order_id, "status": {"$in": ["pending", "paid"]}},
            {"$set": {"status": "cancelled"}},
        )
    final = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    # Gatilhos de email
    order_user = await db.users.find_one({"user_id": final.get("user_id")}, {"_id": 0, "password_hash": 0}) if final else None
    if final and order_user and order_user.get("email"):
        ctx = order_ctx(final, order_user)
        slug_map = {"paid": "order_paid", "shipped": "order_shipped", "delivered": "order_delivered"}
        slug = slug_map.get(status)
        if slug:
            asyncio.create_task(email_service.trigger(db, slug, order_user["email"], ctx))
    # Fatura detalhada para email configurado (apenas quando vira 'paid')
    if status == "paid" and final and order_user:
        asyncio.create_task(_send_admin_invoice_if_configured(db, final, order_user))
    return final

async def _send_admin_invoice_if_configured(db, order, order_user):
    """Envia copia da fatura detalhada para o email configurado em site_settings.
    Acionado quando order vira 'paid' (status de pagamento confirmado).
    """
    try:
        # BUGFIX Iter 38: a chave 'order_invoice_email_to' eh salva em settings {_id: "global"}
        # via PUT /api/admin/settings. _get_site_settings le de {_id: "site"} (aparencia/loja),
        # portanto nunca via a config. Usamos get_settings(db) para ler o doc global.
        s = await get_settings(db)
        to_addr = (s.get("order_invoice_email_to") or "").strip()
        if not to_addr:
            return
        ctx = order_ctx(order, order_user)
        # Pre-renderiza HTML com items, endereco, totais (o render_template nao suporta loops)
        items = order.get("items") or []
        items_rows = []
        for it in items:
            qty = int(it.get("quantity") or 1)
            unit = float(it.get("price") or 0)
            total = float(it.get("total") or (qty * unit))
            items_rows.append(
                f"<tr>"
                f"<td style='padding:6px 8px;border-bottom:1px solid #EEE;'>{(it.get('name') or it.get('product_name') or '')[:80]}</td>"
                f"<td style='padding:6px 8px;border-bottom:1px solid #EEE;text-align:center;'>{qty}</td>"
                f"<td style='padding:6px 8px;border-bottom:1px solid #EEE;text-align:right;'>R$ {unit:.2f}</td>"
                f"<td style='padding:6px 8px;border-bottom:1px solid #EEE;text-align:right;font-weight:bold;'>R$ {total:.2f}</td>"
                f"</tr>"
            )
        items_table = (
            "<table width='100%' cellspacing='0' cellpadding='0' style='border-collapse:collapse;font-size:13px;'>"
            "<thead><tr style='background:#F7F7F7;'>"
            "<th style='padding:6px 8px;text-align:left;'>Produto</th>"
            "<th style='padding:6px 8px;text-align:center;'>Qtd</th>"
            "<th style='padding:6px 8px;text-align:right;'>Unit.</th>"
            "<th style='padding:6px 8px;text-align:right;'>Total</th>"
            "</tr></thead><tbody>" + "".join(items_rows) + "</tbody></table>"
        )
        addr = order.get("shipping_address") or {}
        address_html = (
            f"{addr.get('name') or order_user.get('name') or ''}<br>"
            f"{addr.get('street') or ''}, {addr.get('number') or ''} {addr.get('complement') or ''}<br>"
            f"{addr.get('neighborhood') or ''} - {addr.get('city') or ''}/{addr.get('state') or ''}<br>"
            f"CEP {addr.get('zip_code') or ''}"
        )
        ctx["invoice"] = {
            "items_table_html": items_table,
            "address_html": address_html,
            "subtotal_fmt": f"R$ {float(order.get('subtotal') or 0):.2f}",
            "shipping_fmt": f"R$ {float(order.get('shipping_cost') or 0):.2f}",
            "discount_fmt": f"R$ {float(order.get('discount_amount') or 0):.2f}",
            "total_fmt": f"R$ {float(order.get('total') or 0):.2f}",
            "shipping_service": order.get("shipping_service_name") or "-",
            "shipping_carrier": order.get("shipping_carrier") or "-",
            "payment_method": (order.get("payment_method") or "-").upper(),
            "paid_at_fmt": (order.get("paid_at") or "")[:19].replace("T", " "),
            "invoice_number": order.get("invoice_number") or "-",
            "coupon_code": order.get("coupon_code") or "-",
        }
        await email_service.trigger(db, "invoice_admin_paid", to_addr, ctx, meta={"order_id": order.get("order_id")})
    except Exception as e:
        logger.warning(f"Falha enviando fatura admin: {e}")


@app.delete("/api/admin/orders/{order_id}")
async def admin_delete_order(
    request: Request,
    order_id: str,
    restore_stock: bool = True,
    user: dict = Depends(require_admin()),
):
    """Delecao em cascata de um pedido (util para limpeza de testes).

    Remove:
      - o pedido (`orders`)
      - comissoes associadas (`commissions`)
      - pontos associados (`points_log`)
      - logs de webhook de pagamento (`payment_webhook_logs`)
      - logs de webhook MMN/Maxx (`webhook_logs`)
    Reverte:
      - estoque dos produtos do pedido (se `restore_stock=true`, default true)
      - usage_count do cupom (se o pedido usou cupom)
    """
    db = request.app.db
    order = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Pedido nao encontrado")

    summary = {
        "order_id": order_id,
        "commissions_deleted": 0,
        "points_deleted": 0,
        "payment_webhook_logs_deleted": 0,
        "webhook_logs_deleted": 0,
        "stock_restored": False,
        "coupon_reverted": None,
    }

    # 1) Comissoes geradas pelo pedido
    r = await db.commissions.delete_many({"order_id": order_id})
    summary["commissions_deleted"] = r.deleted_count

    # 2) Pontos gerados
    r = await db.points_log.delete_many({"order_id": order_id})
    summary["points_deleted"] = r.deleted_count

    # 3) Logs de webhooks
    r = await db.payment_webhook_logs.delete_many(
        {"$or": [{"order_id": order_id}, {"external_reference": order_id}]}
    )
    summary["payment_webhook_logs_deleted"] = r.deleted_count

    r = await db.webhook_logs.delete_many({"order_id": order_id})
    summary["webhook_logs_deleted"] = r.deleted_count

    # 4) Reverter estoque (apenas se pedido tinha items abatidos e admin pediu)
    if restore_stock and order.get("items"):
        # Se o pedido ja foi cancelado antes, o estoque pode ter sido restaurado;
        # evitamos dupla restauracao checando o status.
        if order.get("order_status") not in ("cancelled",):
            for it in order["items"]:
                await db.products.update_one(
                    {"product_id": it["product_id"]},
                    {"$inc": {"stock": int(it.get("quantity", 0))}},
                )
            summary["stock_restored"] = True

    # 5) Reverter usage_count do cupom (se o pedido usou)
    if order.get("coupon_code"):
        coupon = await db.coupons.find_one({"code": order["coupon_code"]})
        if coupon and (coupon.get("usage_count") or 0) > 0:
            await db.coupons.update_one(
                {"code": order["coupon_code"]},
                {"$inc": {"usage_count": -1}},
            )
            summary["coupon_reverted"] = order["coupon_code"]

    # 6) Deleta o pedido
    await db.orders.delete_one({"order_id": order_id})

    logger.info(f"Admin {user['user_id']} deletou pedido {order_id}: {summary}")
    return {"ok": True, "summary": summary}


# ==================== ADMIN USERS ====================

@app.get("/api/admin/users")
async def admin_list_users(request: Request, search: Optional[str] = None, role: Optional[str] = None, page: int = 1, limit: int = 20, user: dict = Depends(require_admin())):
    db = request.app.db
    q = {}
    if role:
        q["role"] = role
    if search:
        q["$or"] = [{"name": {"$regex": search, "$options": "i"}}, {"email": {"$regex": search, "$options": "i"}}]
    total = await db.users.count_documents(q)
    users = await db.users.find(q, {"_id": 0, "password_hash": 0}).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    return {"users": users, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit)}

@app.get("/api/admin/users/{user_id}")
async def admin_get_user(request: Request, user_id: str, user: dict = Depends(require_admin())):
    db = request.app.db
    u = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    if not u:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    order_count = await db.orders.count_documents({"user_id": user_id})
    u["total_orders"] = order_count
    return u


# ============ Deteccao e fusao de usuarios duplicados ============
def _user_short(u: Dict) -> Dict:
    return {
        "user_id": u.get("user_id"),
        "name": u.get("name") or "",
        "email": u.get("email") or "",
        "phone": u.get("phone") or "",
        "cpf": u.get("cpf") or "",
        "external_id": u.get("external_id") or "",
        "leader_external_id": u.get("leader_external_id") or "",
        "network_type": u.get("network_type") or "",
        "created_at": u.get("created_at"),
        "addresses": u.get("addresses") or [],
        "has_orders": False,
        "has_points": False,
    }


@app.get("/api/admin/duplicate-users")
async def admin_duplicate_users(request: Request, user: dict = Depends(require_admin())):
    """Cruza usuarios por CPF, email e telefone (normalizados) e retorna grupos
    com 2+ usuarios. Usa Mongo aggregate para performance."""
    db = request.app.db

    async def _groups_by(field):
        pipeline = [
            {"$match": {field: {"$exists": True, "$ne": None, "$nin": ["", None]}}},
            {"$group": {"_id": f"${field}", "user_ids": {"$addToSet": "$user_id"}, "count": {"$sum": 1}}},
            {"$match": {"count": {"$gte": 2}}},
        ]
        return await db.users.aggregate(pipeline).to_list(5000)

    groups = []
    seen_pairs = set()  # (uid_a, uid_b) ordenado para nao duplicar grupo

    # Pre-popular phone_digits para users que nao tem (lazy migration)
    async for udoc in db.users.find({"phone": {"$exists": True, "$ne": ""}, "phone_digits": {"$exists": False}}, {"_id": 0, "user_id": 1, "phone": 1}):
        pd = _clean_phone(udoc.get("phone"))
        if pd:
            await db.users.update_one({"user_id": udoc["user_id"]}, {"$set": {"phone_digits": pd}})

    # Coleta groups por cpf_digits
    cpf_groups = await _groups_by("cpf_digits")
    email_groups = await _groups_by("email")
    phone_groups = await _groups_by("phone_digits")

    raw = []
    for g in cpf_groups:
        raw.append({"match_field": "cpf", "match_value": g["_id"], "user_ids": g["user_ids"]})
    for g in email_groups:
        raw.append({"match_field": "email", "match_value": g["_id"], "user_ids": g["user_ids"]})
    for g in phone_groups:
        raw.append({"match_field": "phone", "match_value": g["_id"], "user_ids": g["user_ids"]})

    # Hidratar com dados do usuario + flags
    for grp in raw:
        users_full = await db.users.find({"user_id": {"$in": grp["user_ids"]}}, {"_id": 0, "password_hash": 0}).to_list(20)
        if len(users_full) < 2:
            continue
        # Skip se for o mesmo "par" ja visto via outro campo
        ids_sorted = tuple(sorted(grp["user_ids"]))
        if ids_sorted in seen_pairs and len(ids_sorted) == 2:
            # Adicionar o match_field como info adicional no grupo ja existente
            for existing in groups:
                ex_ids = tuple(sorted([u["user_id"] for u in existing["users"]]))
                if ex_ids == ids_sorted:
                    if grp["match_field"] not in existing["match_fields"]:
                        existing["match_fields"].append(grp["match_field"])
                    break
            continue
        seen_pairs.add(ids_sorted)

        # Ordenar: o mais recente com external_id primeiro (sugerido como "drop")
        # mas user com mais historico (orders/points) eh sugerido como "keep"
        for u in users_full:
            u["has_orders"] = (await db.orders.count_documents({"user_id": u["user_id"]})) > 0
            u["has_points"] = (await db.points_log.count_documents({"user_id": u["user_id"]})) > 0
            u["has_external_id"] = bool(u.get("external_id"))
        # Sugestao: keep = o que tem mais "historico", drop = o duplicado vindo da API ou sem historico
        users_full.sort(key=lambda u: (
            -1 if (u["has_orders"] or u["has_points"]) else 1,
            -1 if u.get("external_id") else 1,
            u.get("created_at") or "",
        ))

        groups.append({
            "match_fields": [grp["match_field"]],
            "match_value": grp["match_value"],
            "users": [_user_short({**u, "has_orders": u["has_orders"], "has_points": u["has_points"]}) for u in users_full],
            "suggested_keep": users_full[0]["user_id"],
            "suggested_drop": users_full[-1]["user_id"] if len(users_full) > 1 else None,
        })

    return {"groups": groups, "total_groups": len(groups)}


class _MergeBody(BaseModel):
    keep_user_id: str
    drop_user_id: str


@app.post("/api/admin/merge-users")
async def admin_merge_users(request: Request, data: _MergeBody, user: dict = Depends(require_admin())):
    """Funde 2 usuarios. O 'keep' permanece e absorve relacionamentos do 'drop'.

    Substitui dados cadastrais do keep pelos do drop SE o drop tiver valor preenchido
    (cobre o caso "user da Maxx tinha external_id mas faltavam dados que estavam na OxxPharma"
    e tambem o inverso).
    """
    db = request.app.db
    if data.keep_user_id == data.drop_user_id:
        raise HTTPException(status_code=400, detail="keep_user_id e drop_user_id sao iguais")
    keep = await db.users.find_one({"user_id": data.keep_user_id})
    drop = await db.users.find_one({"user_id": data.drop_user_id})
    if not keep or not drop:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")

    # 1. Atualizar dados cadastrais do keep com os do drop (se preenchido)
    overwrite_fields = ["name", "email", "phone", "cpf", "cpf_digits", "phone_digits",
                        "external_id", "leader_external_id", "network_type",
                        "rg", "birth_date", "mother_name"]
    update_set = {}
    for f in overwrite_fields:
        v_drop = drop.get(f)
        if v_drop and str(v_drop).strip():  # so substitui se drop tem valor
            update_set[f] = v_drop
    # Email: garantir lower e nao colidir
    if "email" in update_set:
        update_set["email"] = str(update_set["email"]).lower()
        # Se ja existe outro user com esse email diferente do keep, abortar
        coll = await db.users.find_one({"email": update_set["email"], "user_id": {"$nin": [data.keep_user_id, data.drop_user_id]}})
        if coll:
            raise HTTPException(status_code=409, detail=f"Email {update_set['email']} ja em uso por outro usuario")
    # Recalcula cpf_digits se cpf veio
    if "cpf" in update_set:
        update_set["cpf_digits"] = _clean_cpf(update_set["cpf"]) or update_set.get("cpf_digits")
    if "phone" in update_set:
        update_set["phone_digits"] = _clean_phone(update_set["phone"]) or update_set.get("phone_digits")

    # Mesclar enderecos: keep + drop, deduplicando por (zip+number)
    keep_addrs = keep.get("addresses") or []
    drop_addrs = drop.get("addresses") or []
    if drop_addrs:
        existing_keys = {(a.get("zip_code"), a.get("number"), a.get("street")) for a in keep_addrs}
        merged = list(keep_addrs)
        for a in drop_addrs:
            k = (a.get("zip_code"), a.get("number"), a.get("street"))
            if k not in existing_keys:
                merged.append(a)
                existing_keys.add(k)
        update_set["addresses"] = merged

    # PIX: se keep nao tem mas drop tem
    if not keep.get("pix_key") and drop.get("pix_key"):
        update_set["pix_key"] = drop["pix_key"]
        update_set["pix_key_type"] = drop.get("pix_key_type")

    # referral_program_active: ficar TRUE se algum dos dois tiver
    if drop.get("referral_program_active") and not keep.get("referral_program_active"):
        update_set["referral_program_active"] = True
        if drop.get("referral_code") and not keep.get("referral_code"):
            update_set["referral_code"] = drop.get("referral_code")
        if drop.get("referral_enrollment"):
            update_set["referral_enrollment"] = drop["referral_enrollment"]

    update_set["updated_at"] = now_iso()
    update_set["merged_from_user_ids"] = (keep.get("merged_from_user_ids") or []) + [data.drop_user_id]

    # Antes de aplicar o update no keep, precisamos liberar campos com indice unico
    # (email, cpf_digits, external_id) que possam estar sendo "transferidos" do drop.
    unset_on_drop = {}
    for f in ("email", "cpf_digits", "external_id"):
        if f in update_set and update_set.get(f) and drop.get(f) == update_set.get(f):
            unset_on_drop[f] = ""
    if unset_on_drop:
        await db.users.update_one({"user_id": data.drop_user_id}, {"$unset": unset_on_drop})

    await db.users.update_one({"user_id": data.keep_user_id}, {"$set": update_set})

    # 2. Migrar relacionamentos (compras, pontos, comissoes, saques, batches do cartao,
    #    referidos, indicacoes, e qualquer ref a user_id)
    moved = {}
    for coll_name, field in [
        ("orders", "user_id"),
        ("commissions", "user_id"),
        ("commissions", "from_user_id"),
        ("withdrawals", "user_id"),
        ("points_log", "user_id"),
        ("payment_webhook_logs", "user_id"),
        ("addresses", "user_id"),
    ]:
        coll = db[coll_name]
        try:
            r = await coll.update_many({field: data.drop_user_id}, {"$set": {field: data.keep_user_id}})
            moved[f"{coll_name}.{field}"] = r.modified_count
        except Exception as e:
            moved[f"{coll_name}.{field}"] = f"err: {e}"

    # 3. Atualizar referencias de SPONSOR/LEADER de outros users
    # Quem tinha drop como sponsor passa a ter keep
    r1 = await db.users.update_many({"sponsor_id": data.drop_user_id}, {"$set": {"sponsor_id": data.keep_user_id}})
    moved["users.sponsor_id"] = r1.modified_count
    r2 = await db.users.update_many({"network_sponsor_id": data.drop_user_id}, {"$set": {"network_sponsor_id": data.keep_user_id}})
    moved["users.network_sponsor_id"] = r2.modified_count

    # 4. Mover lines do card_batches (estrutura: lines[].user_id)
    try:
        r3 = await db.card_batches.update_many({"lines.user_id": data.drop_user_id}, {"$set": {"lines.$[elem].user_id": data.keep_user_id}}, array_filters=[{"elem.user_id": data.drop_user_id}])
        moved["card_batches.lines"] = r3.modified_count
    except Exception as e:
        moved["card_batches.lines"] = f"err: {e}"

    # 5. Audit log da fusao (rastreabilidade)
    try:
        # Filtra so campos cadastrais reais para o log (descarta updated_at, merged_from_user_ids etc)
        cadastral_fields_set = set(overwrite_fields) | {"addresses", "pix_key", "pix_key_type", "referral_program_active", "referral_code", "referral_enrollment"}
        fields_overwritten = sorted([k for k in update_set.keys() if k in cadastral_fields_set])
        await db.merge_audit_log.insert_one({
            "merge_id": str(uuid.uuid4()),
            "kept_user_id": data.keep_user_id,
            "deleted_user_id": data.drop_user_id,
            "performed_by_user_id": user.get("user_id"),
            "performed_by_email": user.get("email"),
            "performed_at": now_iso(),
            "kept_snapshot_before": {k: keep.get(k) for k in overwrite_fields if keep.get(k)},
            "drop_snapshot": {k: drop.get(k) for k in overwrite_fields if drop.get(k)},
            "fields_overwritten": fields_overwritten,
            "moved_counts": {k: v for k, v in moved.items() if isinstance(v, int)},
        })
    except Exception as e:
        logger.warning(f"merge_audit_log insert failed: {e}")

    # 6. Deletar o drop
    await db.users.delete_one({"user_id": data.drop_user_id})

    fresh = await db.users.find_one({"user_id": data.keep_user_id}, {"_id": 0, "password_hash": 0})
    return {"success": True, "kept": fresh, "deleted_user_id": data.drop_user_id, "moved": moved}


@app.get("/api/admin/merge-audit-log")
async def admin_merge_audit_log(request: Request, limit: int = 100, user: dict = Depends(require_admin())):
    """Historico de fusoes manuais de contas (auditoria)."""
    db = request.app.db
    rows = await db.merge_audit_log.find({}, {"_id": 0}).sort("performed_at", -1).limit(int(limit)).to_list(int(limit))
    return {"items": rows, "total": len(rows)}


@app.get("/api/admin/users/{user_id}/details")
async def admin_user_details(request: Request, user_id: str, user: dict = Depends(require_admin())):
    """Painel agregador: retorna user + KPIs + listas (commissions, orders, network,
    card balance/logs, points). Usado por /backoffice/usuarios/:id."""
    db = request.app.db
    u = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    if not u:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")

    settings = await get_settings(db)
    release_days = int(settings.get("withdrawal_release_days") or 0)
    cutoff = (datetime.now(timezone.utc) - timedelta(days=release_days)).isoformat()

    # ---- KPIs / Saldos ----
    available_agg = await db.commissions.aggregate([
        {"$match": {"user_id": user_id, "status": "paid", "withdrawal_id": {"$in": [None, ""]}, "paid_at": {"$lte": cutoff}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    quarantine_agg = await db.commissions.aggregate([
        {"$match": {"user_id": user_id, "status": "paid", "withdrawal_id": {"$in": [None, ""]}, "paid_at": {"$gt": cutoff}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    pending_agg = await db.commissions.aggregate([
        {"$match": {"user_id": user_id, "status": "pending"}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    earned_agg = await db.commissions.aggregate([
        {"$match": {"user_id": user_id, "status": {"$in": ["paid", "pending"]}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    withdrawn_agg = await db.withdrawals.aggregate([
        {"$match": {"user_id": user_id, "status": "paid_out"}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    spent_agg = await db.orders.aggregate([
        {"$match": {"user_id": user_id, "payment_status": "paid"}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}, "count": {"$sum": 1}}}
    ]).to_list(1)
    total_orders = await db.orders.count_documents({"user_id": user_id})
    last_order = await db.orders.find_one({"user_id": user_id}, {"_id": 0}, sort=[("created_at", -1)])

    # Cartao de beneficios: somar batches enviados deste user
    card_sent_agg = await db.card_batches.aggregate([
        {"$unwind": "$lines"},
        {"$match": {"lines.user_id": user_id}},
        {"$group": {"_id": None, "total": {"$sum": "$lines.amount"}}}
    ]).to_list(1)
    card_total = round(card_sent_agg[0]["total"] if card_sent_agg else 0, 2)

    # Pontos totais
    points_agg = await db.points_log.aggregate([
        {"$match": {"user_id": user_id}},
        {"$group": {"_id": None, "total": {"$sum": "$points_total"}}}
    ]).to_list(1)
    points_total = int(points_agg[0]["total"] if points_agg else 0)

    # Downline direto (1a geracao)
    direct_downline = await db.users.count_documents({"network_sponsor_id": user_id})
    direct_referrals = await db.users.count_documents({"sponsor_id": user_id})

    kpis = {
        "available": round(available_agg[0]["total"] if available_agg else 0, 2),
        "quarantine": round(quarantine_agg[0]["total"] if quarantine_agg else 0, 2),
        "pending_commissions": round(pending_agg[0]["total"] if pending_agg else 0, 2),
        "total_earned": round(earned_agg[0]["total"] if earned_agg else 0, 2),
        "total_withdrawn": round(withdrawn_agg[0]["total"] if withdrawn_agg else 0, 2),
        "total_spent": round(spent_agg[0]["total"] if spent_agg else 0, 2),
        "total_orders": total_orders,
        "paid_orders": int(spent_agg[0]["count"] if spent_agg else 0),
        "card_total_sent": card_total,
        "points_total": points_total,
        "direct_downline": direct_downline,
        "direct_referrals": direct_referrals,
        "last_order_at": last_order.get("created_at") if last_order else None,
    }

    # Iter 35: breakdown de comissoes por origem (Indicacao / Equipe 1 / Equipe 2)
    by_source_agg = await db.commissions.aggregate([
        {"$match": {"user_id": user_id}},
        {"$group": {
            "_id": {"type": "$type", "network_type": "$network_type", "status": "$status"},
            "total": {"$sum": "$amount"},
            "count": {"$sum": 1},
        }},
    ]).to_list(50)
    commissions_by_source = {
        "affiliate": {"pending": 0, "paid": 0, "paid_out": 0, "count": 0},
        "network_1": {"pending": 0, "paid": 0, "paid_out": 0, "count": 0},
        "network_2": {"pending": 0, "paid": 0, "paid_out": 0, "count": 0},
    }
    for s in by_source_agg:
        key = "affiliate" if s["_id"].get("type") == "affiliate" else (s["_id"].get("network_type") or "network_1")
        if key not in commissions_by_source:
            commissions_by_source[key] = {"pending": 0, "paid": 0, "paid_out": 0, "count": 0}
        st = s["_id"].get("status") or "pending"
        if st in ("pending", "paid", "paid_out"):
            commissions_by_source[key][st] = round(commissions_by_source[key].get(st, 0) + (s["total"] or 0), 2)
        commissions_by_source[key]["count"] += s["count"]
    kpis["commissions_by_source"] = commissions_by_source

    # ---- Comissoes (ate 200 mais recentes) ----
    commissions = await db.commissions.find({"user_id": user_id}, {"_id": 0}).sort("created_at", -1).limit(200).to_list(200)

    # ---- Pedidos (ate 100 mais recentes) ----
    orders = await db.orders.find({"user_id": user_id}, {"_id": 0}).sort("created_at", -1).limit(100).to_list(100)

    # ---- Rede MMN ----
    sponsor = None
    if u.get("sponsor_id"):
        sp = await db.users.find_one({"user_id": u["sponsor_id"]}, {"_id": 0, "user_id": 1, "name": 1, "email": 1, "referral_code": 1})
        if sp:
            sponsor = sp
    network_sponsor = None
    if u.get("network_sponsor_id"):
        nsp = await db.users.find_one({"user_id": u["network_sponsor_id"]}, {"_id": 0, "user_id": 1, "name": 1, "email": 1, "external_id": 1, "network_type": 1})
        if nsp:
            network_sponsor = nsp
    referrals = await db.users.find({"sponsor_id": user_id}, {"_id": 0, "user_id": 1, "name": 1, "email": 1, "created_at": 1, "referral_program_active": 1}).sort("created_at", -1).limit(100).to_list(100)
    network_downline = await db.users.find({"network_sponsor_id": user_id}, {"_id": 0, "user_id": 1, "name": 1, "email": 1, "external_id": 1, "network_type": 1, "created_at": 1}).sort("created_at", -1).limit(100).to_list(100)

    # Downline ate 6 geracoes (BFS)
    # Iter 31: agrega cadeia de afiliados (sponsor_id) + cadeia MMN herdada
    # (network_sponsor_id) - dedup por user_id. Reflete a regra de comissao MMN
    # que sobe por sponsor_id|network_sponsor_id.
    downline_by_generation = []
    current_ids = [user_id]
    seen = {user_id}
    for gen in range(1, 7):
        level_users = await db.users.find(
            {"$or": [
                {"sponsor_id": {"$in": current_ids}},
                {"network_sponsor_id": {"$in": current_ids}},
            ], "user_id": {"$nin": list(seen)}},
            {"_id": 0, "user_id": 1, "name": 1, "email": 1, "phone": 1, "external_id": 1, "network_type": 1, "created_at": 1, "referral_program_active": 1, "sponsor_id": 1, "network_sponsor_id": 1},
        ).sort("created_at", -1).limit(2000).to_list(2000)
        downline_by_generation.append({
            "generation": gen,
            "members_count": len(level_users),
            "members": level_users,
        })
        new_ids = [u["user_id"] for u in level_users]
        seen.update(new_ids)
        current_ids = new_ids
        if not current_ids:
            # preenche zeros ate 6
            for g in range(gen + 1, 7):
                downline_by_generation.append({"generation": g, "members_count": 0, "members": []})
            break

    # ---- Cartao: batches/lines deste user ----
    card_lines = await db.card_batches.aggregate([
        {"$unwind": "$lines"},
        {"$match": {"lines.user_id": user_id}},
        {"$project": {
            "_id": 0,
            "batch_id": "$batch_id",
            "batch_created_at": "$created_at",
            "batch_status": "$status",
            "amount": "$lines.amount",
            "commissions_count": "$lines.commissions_count",
            "sent_at": "$exported_at",
        }},
        {"$sort": {"batch_created_at": -1}},
        {"$limit": 100},
    ]).to_list(100)

    # ---- Pontos: ultimos 200 ----
    points_logs = await db.points_log.find({"user_id": user_id}, {"_id": 0}).sort("registered_at", -1).limit(200).to_list(200)

    return {
        "user": u,
        "kpis": kpis,
        "commissions": commissions,
        "orders": orders,
        "network": {
            "sponsor": sponsor,
            "network_sponsor": network_sponsor,
            "referrals": referrals,
            "downline": network_downline,
            "downline_by_generation": downline_by_generation,
        },
        "card": {
            "total_sent": card_total,
            "lines": card_lines,
        },
        "points": {
            "total": points_total,
            "logs": points_logs,
        },
    }


# ==================== ADMIN DASHBOARD ====================

# ==================== COMMISSION HELPER (reusable) ====================

async def compute_order_commissions(db, order, customer, settings, *, mark_retroactive=False, batch_id=None):
    """Calcula a lista de comissoes (afiliado + MMN ate 6 geracoes) para um pedido.

    Iter 40: Comissoes sao SEMPRE geradas para todos os usuarios na cadeia (afiliado +
    network_1/network_2 ate 6 geracoes). O programa de beneficios controla apenas a
    *visibilidade e o pagamento*: se o beneficiario nao esta com `referral_program_active=True`
    no momento da geracao, a comissao fica com status='pending_enrollment' e nao aparece
    para ele em `/api/users/me/commissions` ate ele se inscrever. Quando ele ativa o
    programa, todas as comissoes pendentes dele sao promovidas para 'pending'.

    Reproduz a regra do /api/checkout (Iter 31), removendo apenas o filtro de inscricao:
      - afiliado: sponsor_id direto do customer recebe affiliate_commission_rate (gen 0)
      - MMN: sobe pela cadeia sponsor_id -> network_sponsor_id ate 6 niveis. Em cada nivel,
        ancestor recebe se for network_1 ou network_2 (independente de inscricao).
      - Se mark_retroactive=True, marca cada doc com retroactive + recalc_batch_id.

    Retorna lista de dicts prontos para inserir em db.commissions (NAO insere).
    """
    subtotal = float(order.get("subtotal") or 0)
    order_id = order["order_id"]
    customer_id = customer["user_id"]
    customer_name = customer.get("name")
    affiliate_rate = float(settings.get("affiliate_commission_rate") or 0)

    out = []

    def _status_for(beneficiary: dict) -> tuple:
        """Retorna (status, program_active) baseado na inscricao do beneficiario."""
        active = bool(beneficiary.get("referral_program_active"))
        return ("pending" if active else "pending_enrollment", active)

    # 1) Afiliado
    sp_id = customer.get("sponsor_id")
    if sp_id and sp_id != customer_id:
        amt = round(subtotal * affiliate_rate, 2)
        if amt > 0:
            sponsor = await db.users.find_one({"user_id": sp_id}, {"_id": 0})
            status, program_active = _status_for(sponsor or {})
            doc = {
                "commission_id": gen_id("com_"),
                "user_id": sp_id,
                "order_id": order_id,
                "customer_id": customer_id,
                "customer_name": customer_name,
                "type": "affiliate",
                "network_type": None,
                "generation": 0,
                "amount": amt,
                "rate": affiliate_rate,
                "order_subtotal": round(subtotal, 2),
                "status": status,
                "program_active_at_creation": program_active,
                "created_at": now_iso(),
            }
            if mark_retroactive:
                doc["retroactive"] = True
                doc["recalc_batch_id"] = batch_id
            out.append(doc)

    # 2) MMN ate 6 geracoes (Iter 40: sem filtro de inscricao; cadeia em estrutura linear)
    # Importante: cliente regular no meio da cadeia GASTA geracao (so nao recebe comissao).
    # Isso preserva a topologia da rede MMN exatamente como cadastrada.
    network1_pcts = settings.get("network1_generations", []) or []
    network2_pcts = settings.get("network2_generations", []) or []
    current_id = customer.get("sponsor_id") or customer.get("network_sponsor_id")
    visited = {customer_id}
    generation = 1
    while generation <= 6 and current_id and current_id not in visited:
        visited.add(current_id)
        ancestor = await db.users.find_one({"user_id": current_id}, {"_id": 0})
        if not ancestor:
            break
        a_net = ancestor.get("network_type")
        if a_net in (NETWORK_1, NETWORK_2):
            pcts = network1_pcts if a_net == NETWORK_1 else network2_pcts
            pct = pcts[generation - 1] if generation - 1 < len(pcts) else 0
            if pct > 0:
                amt = round(subtotal * pct / 100, 2)
                if amt > 0:
                    status, program_active = _status_for(ancestor)
                    doc = {
                        "commission_id": gen_id("com_"),
                        "user_id": ancestor["user_id"],
                        "order_id": order_id,
                        "customer_id": customer_id,
                        "customer_name": customer_name,
                        "type": "network_gen",
                        "network_type": a_net,
                        "generation": generation,
                        "amount": amt,
                        "rate": pct / 100,
                        "order_subtotal": round(subtotal, 2),
                        "status": status,
                        "program_active_at_creation": program_active,
                        "created_at": now_iso(),
                    }
                    if mark_retroactive:
                        doc["retroactive"] = True
                        doc["recalc_batch_id"] = batch_id
                    out.append(doc)
        # Avanca para o proximo ancestral E para a proxima geracao,
        # mesmo quando o atual nao for MMN (nao quebra a cadeia)
        current_id = ancestor.get("sponsor_id") or ancestor.get("network_sponsor_id")
        generation += 1

    return out


# ==================== COMMISSION HELPER END ====================

@app.get("/api/admin/dashboard")
async def admin_dashboard(request: Request, user: dict = Depends(require_admin())):
    db = request.app.db
    n = datetime.now(timezone.utc)
    month_start = n.replace(day=1, hour=0, minute=0, second=0).isoformat()
    today_start = n.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = (today_start - timedelta(days=7)).isoformat()
    prev_week_start = (today_start - timedelta(days=14)).isoformat()
    last_30_start = (today_start - timedelta(days=29)).isoformat()

    total_users = await db.users.count_documents({"role": "customer"})
    total_orders = await db.orders.count_documents({})
    month_orders = await db.orders.count_documents({"created_at": {"$gte": month_start}})

    revenue_agg = await db.orders.aggregate([{"$match": {"payment_status": "paid"}}, {"$group": {"_id": None, "total": {"$sum": "$total"}, "count": {"$sum": 1}}}]).to_list(1)
    total_revenue = revenue_agg[0]["total"] if revenue_agg else 0
    paid_count = revenue_agg[0]["count"] if revenue_agg else 0
    month_rev_agg = await db.orders.aggregate([{"$match": {"payment_status": "paid", "created_at": {"$gte": month_start}}}, {"$group": {"_id": None, "total": {"$sum": "$total"}}}]).to_list(1)
    month_revenue = month_rev_agg[0]["total"] if month_rev_agg else 0
    pending_orders = await db.orders.count_documents({"order_status": "pending"})
    total_products = await db.products.count_documents({"active": True})
    recent_orders = await db.orders.find({}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)

    # Status breakdown com totais em R$
    status_agg = await db.orders.aggregate([
        {"$group": {"_id": "$order_status", "count": {"$sum": 1}, "total": {"$sum": "$total"}}}
    ]).to_list(20)
    by_status = [{"status": s["_id"] or "pending", "count": s["count"], "total": round(s["total"] or 0, 2)} for s in status_agg]

    # Comparativo semanal (semana atual vs anterior)
    cur_week = await db.orders.aggregate([
        {"$match": {"payment_status": "paid", "created_at": {"$gte": week_start}}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}, "count": {"$sum": 1}}}
    ]).to_list(1)
    prev_week = await db.orders.aggregate([
        {"$match": {"payment_status": "paid", "created_at": {"$gte": prev_week_start, "$lt": week_start}}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}, "count": {"$sum": 1}}}
    ]).to_list(1)
    cw_rev = cur_week[0]["total"] if cur_week else 0
    cw_count = cur_week[0]["count"] if cur_week else 0
    pw_rev = prev_week[0]["total"] if prev_week else 0
    pw_count = prev_week[0]["count"] if prev_week else 0
    def _pct(cur, prev):
        if not prev:
            return None  # sem base de comparacao
        return round((cur - prev) / prev * 100, 2)

    # Receita por dia (ultimos 30 dias)
    daily = await db.orders.aggregate([
        {"$match": {"payment_status": "paid", "created_at": {"$gte": last_30_start}}},
        {"$group": {
            "_id": {"$substr": ["$created_at", 0, 10]},
            "revenue": {"$sum": "$total"},
            "orders": {"$sum": 1},
        }},
        {"$sort": {"_id": 1}},
    ]).to_list(60)
    # Preenche dias faltantes com zero
    daily_map = {d["_id"]: {"revenue": round(d["revenue"], 2), "orders": d["orders"]} for d in daily}
    revenue_by_day = []
    for i in range(30):
        day = (today_start - timedelta(days=29 - i)).date().isoformat()
        d = daily_map.get(day, {"revenue": 0, "orders": 0})
        revenue_by_day.append({"date": day, "revenue": d["revenue"], "orders": d["orders"]})

    # Top 10 compradores (por valor pago)
    top_buyers_agg = await db.orders.aggregate([
        {"$match": {"payment_status": "paid", "user_id": {"$ne": None}}},
        {"$group": {"_id": "$user_id", "total": {"$sum": "$total"}, "orders": {"$sum": 1}}},
        {"$sort": {"total": -1}},
        {"$limit": 10},
    ]).to_list(10)
    top_buyer_ids = [b["_id"] for b in top_buyers_agg]
    buyer_users = {}
    if top_buyer_ids:
        async for u in db.users.find({"user_id": {"$in": top_buyer_ids}}, {"_id": 0, "user_id": 1, "name": 1, "email": 1, "network_type": 1}):
            buyer_users[u["user_id"]] = u
    top_buyers = []
    for b in top_buyers_agg:
        u = buyer_users.get(b["_id"], {})
        top_buyers.append({
            "user_id": b["_id"],
            "name": u.get("name") or "(sem nome)",
            "email": u.get("email"),
            "network_type": u.get("network_type") or "customer",
            "total": round(b["total"] or 0, 2),
            "orders": b["orders"],
        })

    # Top 10 indicadores (por comissoes ganhas, tipo affiliate ou network_gen)
    top_aff_agg = await db.commissions.aggregate([
        {"$match": {"status": {"$in": ["paid", "pending"]}}},
        {"$group": {
            "_id": "$user_id",
            "total": {"$sum": "$amount"},
            "count": {"$sum": 1},
            "paid": {"$sum": {"$cond": [{"$eq": ["$status", "paid"]}, "$amount", 0]}},
        }},
        {"$sort": {"total": -1}},
        {"$limit": 10},
    ]).to_list(10)
    top_aff_ids = [a["_id"] for a in top_aff_agg]
    aff_users = {}
    referral_counts = {}
    if top_aff_ids:
        async for u in db.users.find({"user_id": {"$in": top_aff_ids}}, {"_id": 0, "user_id": 1, "name": 1, "email": 1, "network_type": 1, "referral_code": 1}):
            aff_users[u["user_id"]] = u
        # Conta indicacoes diretas (sponsor_id)
        ref_agg = await db.users.aggregate([
            {"$match": {"sponsor_id": {"$in": top_aff_ids}}},
            {"$group": {"_id": "$sponsor_id", "n": {"$sum": 1}}},
        ]).to_list(50)
        referral_counts = {r["_id"]: r["n"] for r in ref_agg}
    top_affiliates = []
    for a in top_aff_agg:
        u = aff_users.get(a["_id"], {})
        top_affiliates.append({
            "user_id": a["_id"],
            "name": u.get("name") or "(sem nome)",
            "email": u.get("email"),
            "network_type": u.get("network_type") or "customer",
            "referral_code": u.get("referral_code"),
            "total_earned": round(a["total"] or 0, 2),
            "total_paid": round(a["paid"] or 0, 2),
            "commissions_count": a["count"],
            "referrals_count": referral_counts.get(a["_id"], 0),
        })

    # Comissoes consolidadas
    comm_agg = await db.commissions.aggregate([
        {"$group": {"_id": "$status", "total": {"$sum": "$amount"}}}
    ]).to_list(10)
    commissions_summary = {(s["_id"] or "pending"): round(s["total"] or 0, 2) for s in comm_agg}

    # Ticket medio (orders pagos)
    avg_ticket = round(total_revenue / paid_count, 2) if paid_count else 0

    return {
        "total_users": total_users, "total_orders": total_orders, "month_orders": month_orders,
        "total_revenue": round(total_revenue, 2), "month_revenue": round(month_revenue, 2),
        "pending_orders": pending_orders, "total_products": total_products,
        "paid_orders": paid_count, "avg_ticket": avg_ticket,
        "recent_orders": recent_orders, "orders_by_status": by_status,
        "revenue_by_day": revenue_by_day,
        "weekly_comparison": {
            "current_revenue": round(cw_rev, 2),
            "previous_revenue": round(pw_rev, 2),
            "revenue_pct": _pct(cw_rev, pw_rev),
            "current_orders": cw_count,
            "previous_orders": pw_count,
            "orders_pct": _pct(cw_count, pw_count),
        },
        "top_buyers": top_buyers,
        "top_affiliates": top_affiliates,
        "commissions_summary": commissions_summary,
    }


# ==================== ADMIN: RECALCULAR COMISSOES (RETROATIVO) ====================

class _RecalcBody(BaseModel):
    date_from: Optional[str] = None  # ISO date inclusive (YYYY-MM-DD)
    date_to: Optional[str] = None    # ISO date inclusive
    user_id: Optional[str] = None    # filtra por user_id (comprador)
    customer_email: Optional[str] = None  # filtra por email do comprador
    order_ids: Optional[List[str]] = None  # opcional, lista especifica
    force: bool = False              # Iter 41: se True, apaga comissoes existentes e recalcula


async def _list_orders_for_recalc(db, body: _RecalcBody):
    """Retorna orders pagos. Se body.force, retorna TODOS no filtro (mesmo que ja tenham
    comissoes — elas serao apagadas no apply). Caso contrario, apenas os que estao SEM
    comissao (comportamento original)."""
    q = {"payment_status": "paid"}
    if body.date_from:
        q.setdefault("created_at", {})["$gte"] = body.date_from + "T00:00:00"
    if body.date_to:
        q.setdefault("created_at", {})["$lte"] = body.date_to + "T23:59:59"
    if body.user_id:
        q["user_id"] = body.user_id
    if body.customer_email:
        cust = await db.users.find_one({"email": body.customer_email.strip().lower()}, {"_id": 0, "user_id": 1})
        if not cust:
            return []
        q["user_id"] = cust["user_id"]
    if body.order_ids:
        q["order_id"] = {"$in": body.order_ids}

    orders = await db.orders.find(q, {"_id": 0}).sort("created_at", -1).limit(2000).to_list(2000)
    if not orders:
        return []
    if body.force:
        return orders
    order_ids = [o["order_id"] for o in orders]
    have = await db.commissions.distinct("order_id", {"order_id": {"$in": order_ids}})
    have_set = set(have)
    return [o for o in orders if o["order_id"] not in have_set]


# Mantem nome antigo para compatibilidade com chamadas existentes
async def _list_orders_missing_commissions(db, body: _RecalcBody):
    return await _list_orders_for_recalc(db, body)


async def _build_recalc_preview(db, body: _RecalcBody):
    """Para cada order elegivel, simula compute_order_commissions e agrega o resultado."""
    settings = await get_settings(db)
    orders = await _list_orders_missing_commissions(db, body)
    by_user = {}  # user_id -> {name, email, total, count}
    affected_orders = []
    total_amount = 0.0
    total_commissions = 0
    for o in orders:
        cust = await db.users.find_one({"user_id": o.get("user_id")}, {"_id": 0}) if o.get("user_id") else None
        if not cust:
            affected_orders.append({
                "order_id": o["order_id"], "customer_name": o.get("customer_name"),
                "customer_email": o.get("customer_email"), "subtotal": o.get("subtotal"),
                "total": o.get("total"), "created_at": o.get("created_at"),
                "commissions_count": 0, "commissions_total": 0,
                "skipped_reason": "Comprador nao encontrado",
            })
            continue
        comms = await compute_order_commissions(db, o, cust, settings, mark_retroactive=False)
        order_total = round(sum(c["amount"] for c in comms), 2)
        affected_orders.append({
            "order_id": o["order_id"],
            "customer_name": o.get("customer_name") or cust.get("name"),
            "customer_email": o.get("customer_email") or cust.get("email"),
            "subtotal": o.get("subtotal"),
            "total": o.get("total"),
            "created_at": o.get("created_at"),
            "commissions_count": len(comms),
            "commissions_total": order_total,
            "commissions_breakdown": [
                {"user_id": c["user_id"], "type": c["type"], "generation": c["generation"],
                 "amount": c["amount"], "rate": c["rate"], "status": c.get("status"),
                 "program_active_at_creation": c.get("program_active_at_creation")}
                for c in comms
            ],
        })
        for c in comms:
            uid = c["user_id"]
            row = by_user.setdefault(uid, {"user_id": uid, "total": 0, "count": 0})
            row["total"] = round(row["total"] + c["amount"], 2)
            row["count"] += 1
        total_amount = round(total_amount + order_total, 2)
        total_commissions += len(comms)

    # Hidrata nomes/emails dos beneficiarios
    if by_user:
        async for u in db.users.find({"user_id": {"$in": list(by_user.keys())}}, {"_id": 0, "user_id": 1, "name": 1, "email": 1, "network_type": 1}):
            by_user[u["user_id"]]["name"] = u.get("name")
            by_user[u["user_id"]]["email"] = u.get("email")
            by_user[u["user_id"]]["network_type"] = u.get("network_type")

    by_user_list = sorted(by_user.values(), key=lambda x: -x["total"])

    # Iter 40: split por status (pending = inscrito; pending_enrollment = aguardando)
    pending_count = 0
    pending_enrollment_count = 0
    pending_amount = 0.0
    pending_enrollment_amount = 0.0
    for o in affected_orders:
        for c in (o.get("commissions_breakdown") or []):
            if c.get("status") == "pending_enrollment":
                pending_enrollment_count += 1
                pending_enrollment_amount = round(pending_enrollment_amount + (c.get("amount") or 0), 2)
            else:
                pending_count += 1
                pending_amount = round(pending_amount + (c.get("amount") or 0), 2)

    return {
        "orders_eligible": len(orders),
        "orders_with_commissions": len([o for o in affected_orders if o["commissions_count"] > 0]),
        "orders_without_eligible_chain": len([o for o in affected_orders if o["commissions_count"] == 0]),
        "total_commissions": total_commissions,
        "total_amount": total_amount,
        "commissions_pending": pending_count,
        "commissions_pending_amount": pending_amount,
        "commissions_pending_enrollment": pending_enrollment_count,
        "commissions_pending_enrollment_amount": pending_enrollment_amount,
        "affected_orders": affected_orders,
        "beneficiaries": by_user_list,
    }


@app.post("/api/admin/recalc-commissions/preview")
async def admin_recalc_preview(request: Request, body: _RecalcBody, user: dict = Depends(require_admin())):
    """Modo SIMULAR: nao grava nada. Retorna resumo das comissoes que seriam criadas."""
    db = request.app.db
    return await _build_recalc_preview(db, body)


@app.post("/api/admin/recalc-commissions/apply")
async def admin_recalc_apply(request: Request, body: _RecalcBody, user: dict = Depends(require_admin())):
    """Modo APLICAR: cria comissoes retroativas para orders pagos sem comissao,
    seguindo a mesma regra do checkout (afiliado + MMN ate 6 geracoes via cadeia
    sponsor_id->network_sponsor_id, MMN com programa ativo).

    Cada commission criada recebe retroactive=True e recalc_batch_id=<batch>.
    Audit log em db.recalc_audit_log.
    """
    db = request.app.db
    settings = await get_settings(db)
    orders = await _list_orders_missing_commissions(db, body)
    if not orders:
        return {"success": True, "batch_id": None, "orders_processed": 0, "commissions_created": 0, "total_amount": 0}

    batch_id = gen_id("recalc_")
    all_docs = []
    processed_ids = []
    deleted_count = 0
    for o in orders:
        cust = await db.users.find_one({"user_id": o.get("user_id")}, {"_id": 0}) if o.get("user_id") else None
        if not cust:
            continue
        if body.force:
            # Iter 41: apaga TODAS as comissoes existentes deste pedido para recriar com a regra atual.
            # Comissoes ja PAGAS sao preservadas (somente pending e pending_enrollment sao apagadas).
            del_res = await db.commissions.delete_many({
                "order_id": o["order_id"],
                "status": {"$in": ["pending", "pending_enrollment"]},
            })
            deleted_count += del_res.deleted_count
        else:
            if await db.commissions.count_documents({"order_id": o["order_id"]}) > 0:
                continue
        comms = await compute_order_commissions(db, o, cust, settings, mark_retroactive=True, batch_id=batch_id)
        if body.force:
            # Em modo force, nao recriar comissoes que ja foram pagas (preserva historico)
            paid_keys = set()
            async for paid in db.commissions.find(
                {"order_id": o["order_id"], "status": "paid"},
                {"_id": 0, "user_id": 1, "type": 1, "generation": 1}
            ):
                paid_keys.add((paid["user_id"], paid["type"], paid.get("generation")))
            comms = [c for c in comms if (c["user_id"], c["type"], c.get("generation")) not in paid_keys]
        if comms:
            all_docs.extend(comms)
            processed_ids.append(o["order_id"])

    total_amount = round(sum(d["amount"] for d in all_docs), 2)
    if all_docs:
        await db.commissions.insert_many(all_docs)

    # Audit log
    try:
        await db.recalc_audit_log.insert_one({
            "batch_id": batch_id,
            "performed_by_user_id": user.get("user_id"),
            "performed_by_email": user.get("email"),
            "performed_at": now_iso(),
            "filters": body.dict(),
            "orders_processed": len(processed_ids),
            "commissions_created": len(all_docs),
            "total_amount": total_amount,
            "order_ids": processed_ids,
        })
    except Exception as e:
        logger.warning(f"recalc_audit_log insert failed: {e}")

    return {
        "success": True,
        "batch_id": batch_id,
        "orders_processed": len(processed_ids),
        "commissions_created": len(all_docs),
        "total_amount": total_amount,
    }


@app.get("/api/admin/recalc-commissions/history")
async def admin_recalc_history(request: Request, limit: int = 50, user: dict = Depends(require_admin())):
    db = request.app.db
    rows = await db.recalc_audit_log.find({}, {"_id": 0}).sort("performed_at", -1).limit(int(limit)).to_list(int(limit))
    return {"items": rows, "total": len(rows)}

# ==================== AFFILIATE / REFERRALS ====================

@app.get("/api/referrals/validate/{code}")
async def validate_referral_code(request: Request, code: str):
    """Valida codigo de indicacao publico - usado pela loja antes do checkout.
    Iter 39: tambem registra um click na collection `referral_clicks` para
    estatisticas do indicador na pagina /indique-ganhe.
    """
    db = request.app.db
    code_norm = code.strip().upper()
    u = await db.users.find_one({"referral_code": code_norm, "status": "active"}, {"_id": 0, "password_hash": 0})
    if not u:
        return {"valid": False}
    # Registra clique (best-effort, nao bloqueia resposta)
    try:
        await db.referral_clicks.insert_one({
            "click_id": gen_id("clk_"),
            "code": code_norm,
            "owner_user_id": u.get("user_id"),
            "ip": (request.client.host if request.client else None),
            "user_agent": request.headers.get("user-agent", "")[:300],
            "referer": request.headers.get("referer", "")[:300],
            "created_at": now_iso(),
        })
    except Exception as e:
        logger.debug(f"referral_clicks insert error: {e}")
    return {"valid": True, "code": code_norm, "affiliate_name": u.get("name")}

@app.get("/api/users/me/referrals")
async def my_referrals_list(request: Request, page: int = 1, limit: int = 20, user: dict = Depends(get_current_user)):
    """Iter 39: lista quem entrou pelo link do usuario logado, com status agregado:
    cadastrou, comprou (qtd/total), aderiu ao programa de beneficios.
    Tambem retorna sumario de cliques (total + ultimos 30 dias).
    """
    db = request.app.db
    uid = user["user_id"]

    # Cliques no link
    clicks_total = await db.referral_clicks.count_documents({"owner_user_id": uid})
    from datetime import timedelta
    since_30 = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    clicks_30 = await db.referral_clicks.count_documents({
        "owner_user_id": uid, "created_at": {"$gte": since_30}
    })

    # Lista de pessoas indicadas (sponsor_id = uid). Ordena por mais recente.
    q = {"sponsor_id": uid}
    total = await db.users.count_documents(q)
    skip = max(0, (page - 1) * limit)
    cursor = db.users.find(
        q,
        {"_id": 0, "user_id": 1, "name": 1, "email": 1, "created_at": 1,
         "referral_program_active": 1, "referral_enrollment_status": 1}
    ).sort("created_at", -1).skip(skip).limit(limit)
    indicated = await cursor.to_list(limit)

    # Para cada usuario, agrega pedidos pagos e total gasto
    paid_filter = {"$in": ["paid", "shipped", "delivered"]}
    enriched = []
    enrolled_total = 0
    purchasers_total = 0
    for u in indicated:
        agg = await db.orders.aggregate([
            {"$match": {"user_id": u["user_id"], "payment_status": paid_filter}},
            {"$group": {"_id": None, "count": {"$sum": 1}, "total": {"$sum": "$total"}}},
        ]).to_list(1)
        orders_count = agg[0]["count"] if agg else 0
        orders_total = round(agg[0]["total"], 2) if agg else 0
        if orders_count > 0:
            purchasers_total += 1
        enrolled = bool(u.get("referral_program_active") or u.get("referral_enrollment_status") in ("approved", "pending_approval"))
        if enrolled:
            enrolled_total += 1
        # Mascara o email para privacidade (ex.: ma****@gmail.com)
        email = u.get("email") or ""
        masked_email = email
        if email and "@" in email:
            local, _, domain = email.partition("@")
            if len(local) > 3:
                masked_email = local[:2] + "*" * max(2, len(local) - 4) + local[-2:] + "@" + domain
        enriched.append({
            "user_id": u["user_id"],
            "name": (u.get("name") or "").split(" ")[0] + (" " + (u.get("name") or "").split(" ")[-1][:1] + "." if len((u.get("name") or "").split(" ")) > 1 else ""),
            "email_masked": masked_email,
            "registered_at": u.get("created_at"),
            "orders_count": orders_count,
            "orders_total": orders_total,
            "has_purchases": orders_count > 0,
            "enrolled_in_program": enrolled,
            "enrollment_status": u.get("referral_enrollment_status"),
        })

    # Totais agregados (independente da paginacao)
    enrolled_count_full = await db.users.count_documents({
        "sponsor_id": uid,
        "$or": [
            {"referral_program_active": True},
            {"referral_enrollment_status": {"$in": ["approved", "pending_approval"]}},
        ],
    })
    # Compradores: precisa de aggregation (orders join) - faz lookup compacto
    purchasers_pipeline = [
        {"$match": {"sponsor_id": uid}},
        {"$lookup": {
            "from": "orders",
            "let": {"uid": "$user_id"},
            "pipeline": [
                {"$match": {"$expr": {"$and": [
                    {"$eq": ["$user_id", "$$uid"]},
                    {"$in": ["$payment_status", ["paid", "shipped", "delivered"]]},
                ]}}},
                {"$limit": 1},
            ],
            "as": "paid_orders",
        }},
        {"$match": {"paid_orders.0": {"$exists": True}}},
        {"$count": "n"},
    ]
    pres = await db.users.aggregate(purchasers_pipeline).to_list(1)
    purchasers_count_full = pres[0]["n"] if pres else 0

    return {
        "summary": {
            "clicks_total": clicks_total,
            "clicks_30d": clicks_30,
            "registered_count": total,
            "purchasers_count": purchasers_count_full,
            "enrolled_in_program_count": enrolled_count_full,
        },
        "users": enriched,
        "page": page,
        "limit": limit,
        "total": total,
        "pages": max(1, (total + limit - 1) // limit),
    }


@app.get("/api/users/me/referral")
async def my_referral(request: Request, user: dict = Depends(get_current_user)):
    db = request.app.db
    # referral_code agora so eh gerado quando usuario adere ao programa (via /referral-enrollment)
    has_program = bool(user.get("referral_program_active") and user.get("referral_code"))
    # Estatisticas
    total_comm = await db.commissions.aggregate([
        {"$match": {"user_id": user["user_id"]}},
        {"$group": {"_id": "$status", "total": {"$sum": "$amount"}, "count": {"$sum": 1}}}
    ]).to_list(10)
    stats = {"pending": 0, "paid": 0, "cancelled": 0, "total_count": 0, "total_earned": 0}
    for s in total_comm:
        k = s["_id"] or "pending"
        stats[k] = round(s["total"], 2)
        stats["total_count"] += s["count"]
    stats["total_earned"] = round(stats.get("paid", 0), 2)
    # Total de pessoas indicadas
    referrals_count = await db.users.count_documents({"sponsor_id": user["user_id"]})
    # Saldos do novo modelo (cartao)
    account_agg = await db.commissions.aggregate([
        {"$match": {"user_id": user["user_id"], "status": "paid", "sent_to_card": {"$ne": True}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    sent_agg = await db.commissions.aggregate([
        {"$match": {"user_id": user["user_id"], "sent_to_card": True}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    # Historico detalhado de envios para o cartao (D+2)
    history_pipeline = [
        {"$match": {"user_id": user["user_id"], "sent_to_card": True}},
        {"$group": {
            "_id": {"batch_id": "$card_batch_id", "sent_at": "$sent_to_card_at"},
            "amount": {"$sum": "$amount"},
            "count": {"$sum": 1},
        }},
        {"$sort": {"_id.sent_at": -1}},
        {"$limit": 50},
    ]
    raw = await db.commissions.aggregate(history_pipeline).to_list(50)
    card_history = []
    for r in raw:
        sent_at = r["_id"].get("sent_at")
        if not sent_at:
            continue
        try:
            sent_dt = datetime.fromisoformat(sent_at.replace("Z", "+00:00")) if isinstance(sent_at, str) else sent_at
            available_dt = sent_dt + timedelta(days=2)
        except Exception:
            available_dt = None
        card_history.append({
            "sent_at": sent_at,
            "available_at": available_dt.isoformat() if available_dt else None,
            "amount": round(r["amount"], 2),
            "commissions_count": r["count"],
        })
    return {
        "has_referral_program": has_program,
        "referral_code": user.get("referral_code") if has_program else None,
        "commission_rate": AFFILIATE_COMMISSION_RATE,
        "referrals_count": referrals_count,
        "stats": stats,
        "account_balance": round((account_agg[0]["total"] if account_agg else 0), 2),
        "sent_to_card_total": round((sent_agg[0]["total"] if sent_agg else 0), 2),
        "card_history": card_history,
        "card_release_days": 2,
        "referral_enrollment": user.get("referral_enrollment"),
        "referral_enrollment_status": user.get("referral_enrollment_status"),
        "referral_rejected_reason": user.get("referral_rejected_reason"),
    }

@app.get("/api/users/me/commissions")
async def my_commissions(request: Request, page: int = 1, limit: int = 20, user: dict = Depends(get_current_user)):
    db = request.app.db
    # Iter 40: usuario nao ve comissoes em status pending_enrollment (geradas antes
    # dele se inscrever no programa). Quando ele ativa, sao promovidas para 'pending'.
    q = {"user_id": user["user_id"], "status": {"$ne": "pending_enrollment"}}
    total = await db.commissions.count_documents(q)
    comms = await db.commissions.find(q, {"_id": 0}).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    return {"commissions": comms, "total": total, "page": page}

# ==================== PAYMENT - MERCADO PAGO ====================

@app.get("/api/payments/config")
async def payments_config(request: Request):
    """Config publica do MP (env + public_key)."""
    return await payments_service.get_public_config(request.app.db)


@app.post("/api/payments/create/{order_id}")
async def create_payment(request: Request, order_id: str, user: dict = Depends(get_current_user)):
    db = request.app.db
    order = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Pedido nao encontrado")
    if order["user_id"] != user["user_id"]:
        raise HTTPException(status_code=403, detail="Acesso negado")
    if order.get("payment_status") == "paid":
        raise HTTPException(status_code=400, detail="Pedido ja pago")

    # Iter 38: Pedido com total <= 0 (totalmente pago via voucher) eh marcado pago direto,
    # sem criar preferencia no MercadoPago.
    order_total = float(order.get("total") or 0)
    if order_total <= 0.005:
        await mark_order_paid(db, order_id, payment_id=f"voucher_{uuid.uuid4().hex[:10]}", source="voucher")
        return {
            "order_id": order_id,
            "payment_id": None,
            "payment_url": None,
            "provider": "voucher",
            "paid": True,
        }

    if not await payments_service.is_mp_configured(db):
        # Sem MP configurado: continua mock (auto-aprovavel via /mock/confirm)
        payment_id = f"mock_{uuid.uuid4().hex[:12]}"
        await db.orders.update_one(
            {"order_id": order_id},
            {"$set": {"payment_provider": "mock", "payment_id": payment_id, "payment_url": None}},
        )
        return {"order_id": order_id, "payment_id": payment_id, "payment_url": None, "provider": "mock"}

    items_full = order.get("items", []) + [{"product_id": "shipping", "name": "Frete", "price": float(order.get("shipping_cost") or 0), "quantity": 1}]
    items_full = [it for it in items_full if (it.get("price") or 0) > 0]

    # Iter 38: Se houver voucher utilizado ou desconto de cupom, consolida os itens em
    # uma unica linha com o valor real a cobrar (order.total). Caso contrario o MP
    # cobraria o subtotal + frete sem considerar abatimentos.
    voucher_used = float(order.get("voucher_used") or 0)
    discount_amount = float(order.get("discount_amount") or 0)
    if (voucher_used > 0 or discount_amount > 0) and order_total > 0:
        items_full = [{
            "product_id": order.get("order_id"),
            "name": f"Pedido #{order.get('order_id')}",
            "price": round(order_total, 2),
            "quantity": 1,
        }]

    frontend_url = get_app_url()
    backend_url = os.environ.get("BACKEND_URL") or frontend_url

    try:
        pref = await payments_service.create_preference(db, order, user, items_full, frontend_url, backend_url)
    except Exception as e:
        logger.exception(f"Falha criando preferencia MP para {order_id}: {e}")
        raise HTTPException(status_code=502, detail=f"Erro MercadoPago: {e}")

    # Escolhe init_point: se environment=test, usa sandbox_init_point
    env = pref.get("environment")
    init_point = pref.get("sandbox_init_point") if env == "test" else pref.get("init_point")
    init_point = init_point or pref.get("init_point")

    await db.orders.update_one(
        {"order_id": order_id},
        {"$set": {
            "payment_provider": "mercadopago",
            "payment_id": pref["preference_id"],
            "payment_url": init_point,
            "mp_environment": env,
        }},
    )
    return {
        "order_id": order_id,
        "payment_id": pref["preference_id"],
        "payment_url": init_point,
        "provider": "mercadopago",
        "environment": env,
    }

@app.post("/api/payments/mock/confirm/{order_id}")
async def mock_confirm_payment(request: Request, order_id: str, user: dict = Depends(get_current_user)):
    """Confirma pagamento manualmente (mock). Disponivel em ambiente test ou quando MP nao configurado."""
    db = request.app.db
    env = await payments_service.get_mp_environment(db)
    if env == "production" and await payments_service.is_mp_configured(db):
        raise HTTPException(status_code=403, detail="Mock indisponivel em producao")
    order = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Pedido nao encontrado")
    if order["user_id"] != user["user_id"]:
        raise HTTPException(status_code=403, detail="Acesso negado")
    return await mark_order_paid(db, order_id, payment_id=order.get("payment_id") or f"mock_{uuid.uuid4().hex[:8]}", source="mock")


async def mark_order_paid(db, order_id: str, payment_id: Optional[str] = None, source: str = "mp"):
    """Helper centralizado: marca pedido como pago, cria nota, paga commissions, registra pontos, dispara email."""
    order = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    if not order:
        return None
    if order.get("payment_status") == "paid":
        return order  # idempotente
    await db.orders.update_one(
        {"order_id": order_id},
        {"$set": {"payment_status": "paid", "order_status": "paid", "paid_at": now_iso(), "payment_id": payment_id, "paid_via": source}},
    )
    # Nota fiscal
    fresh = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    if fresh and not fresh.get("invoice_number"):
        inv = await issue_invoice(db, fresh)
        await db.orders.update_one(
            {"order_id": order_id},
            {"$set": {"invoice_number": inv["number"], "invoice_issued_at": inv["issued_at"]}},
        )
    # Commissions -> paid
    await db.commissions.update_many(
        {"order_id": order_id, "status": "pending"},
        {"$set": {"status": "paid", "paid_at": now_iso()}},
    )
    # Pontos
    await register_points_from_order(db, order_id)
    # Email
    final = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    order_user = await db.users.find_one({"user_id": final.get("user_id")}, {"_id": 0, "password_hash": 0}) if final else None
    if final and order_user and order_user.get("email"):
        asyncio.create_task(email_service.trigger(db, "order_paid", order_user["email"], order_ctx(final, order_user)))
    # Fatura detalhada para email configurado no admin
    if final and order_user:
        asyncio.create_task(_send_admin_invoice_if_configured(db, final, order_user))
    return final


async def register_points_from_order(db, order_id: str):
    """Registra logs de pontos para cada item com points_value > 0."""
    o = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    if not o:
        return
    # Idempotencia: se ja registrou, ignora
    existing = await db.points_log.find_one({"order_id": order_id})
    if existing:
        return
    user = await db.users.find_one({"user_id": o["user_id"]}, {"_id": 0, "password_hash": 0})
    if not user:
        return
    logs = []
    for it in o.get("items", []):
        pv = float(it.get("points_value") or 0)
        if pv <= 0:
            continue
        qty = int(it.get("quantity") or 1)
        total_points = round(pv * qty, 2)
        logs.append({
            "log_id": gen_id("pts_"),
            "user_id": user["user_id"],
            "user_name": user.get("name"),
            "user_email": user.get("email"),
            "user_external_id": user.get("external_id"),
            "order_id": order_id,
            "product_id": it.get("product_id"),
            "product_name": it.get("name"),
            "quantity": qty,
            "points_per_unit": pv,
            "points_total": total_points,
            "registered_at": now_iso(),
            "applied_externally": False,
        })
    if logs:
        await db.points_log.insert_many(logs)

        # Iter 37: Pontos espelhados para o sponsor direto (1a geracao) se ele for Equipe 1.
        # Cada produto comprado pelo cliente final tambem gera os mesmos pontos no sponsor
        # de Equipe 1, para ser enviado a Maxx no nome do leader.
        sponsor_id = user.get("sponsor_id")
        if sponsor_id:
            sponsor = await db.users.find_one({"user_id": sponsor_id}, {"_id": 0, "password_hash": 0})
            if sponsor and sponsor.get("network_type") == NETWORK_1:
                mirrored = []
                for it in o.get("items", []):
                    pv = float(it.get("points_value") or 0)
                    if pv <= 0:
                        continue
                    qty = int(it.get("quantity") or 1)
                    total_points = round(pv * qty, 2)
                    mirrored.append({
                        "log_id": gen_id("pts_"),
                        "user_id": sponsor["user_id"],
                        "user_name": sponsor.get("name"),
                        "user_email": sponsor.get("email"),
                        "user_external_id": sponsor.get("external_id"),
                        "order_id": order_id,
                        "product_id": it.get("product_id"),
                        "product_name": it.get("name"),
                        "quantity": qty,
                        "points_per_unit": pv,
                        "points_total": total_points,
                        "registered_at": now_iso(),
                        "applied_externally": False,
                        # Marca a origem para auditoria
                        "source": "team1_indicated",
                        "indicated_user_id": user["user_id"],
                        "indicated_user_name": user.get("name"),
                    })
                if mirrored:
                    await db.points_log.insert_many(mirrored)
                    logs = logs + mirrored  # consolida para envio a Maxx

        # Realtime sync para Maxx (best-effort, nao bloqueia)
        try:
            await maxx_service.trigger_realtime(db, [l["log_id"] for l in logs])
        except Exception as e:
            logger.warning(f"trigger_realtime maxx falhou: {e}")

@app.post("/api/payments/webhook/mercadopago")
async def mp_webhook(request: Request):
    """Webhook MercadoPago: valida assinatura, busca payment, marca pedido como pago."""
    db = request.app.db
    raw = await request.body()
    x_signature = request.headers.get("x-signature")
    x_request_id = request.headers.get("x-request-id")
    try:
        body = await request.json()
    except Exception:
        body = {}
    data_id = (body.get("data") or {}).get("id") or request.query_params.get("data.id") or request.query_params.get("id")

    valid = await payments_service.verify_webhook_signature(db, raw, x_signature, x_request_id, str(data_id) if data_id else None)
    log_entry = {
        "log_id": gen_id("mphook_"),
        "received_at": now_iso(),
        "type": body.get("type") or body.get("topic") or request.query_params.get("topic"),
        "data_id": data_id,
        "valid_signature": valid,
        "raw_body": body,
        "query": dict(request.query_params),
    }
    if not valid:
        log_entry["error"] = "invalid_signature"
        await db.payment_webhook_logs.insert_one(log_entry)
        raise HTTPException(status_code=401, detail="Invalid signature")

    notif_type = body.get("type") or body.get("topic") or request.query_params.get("topic")
    if notif_type == "payment" and data_id:
        details = await payments_service.get_payment_details(db, str(data_id))
        log_entry["payment_details"] = {
            "id": details.get("id") if details else None,
            "status": details.get("status") if details else None,
            "external_reference": details.get("external_reference") if details else None,
        }
        if details:
            order_id = details.get("external_reference")
            status_mp = details.get("status")
            if order_id and status_mp == "approved":
                await mark_order_paid(db, order_id, payment_id=str(details.get("id")), source="mercadopago")
                log_entry["action"] = "marked_paid"
            elif order_id and status_mp in ("rejected", "cancelled"):
                await db.orders.update_one({"order_id": order_id}, {"$set": {"payment_status": status_mp}})
                log_entry["action"] = f"set_{status_mp}"

    await db.payment_webhook_logs.insert_one(log_entry)
    return {"received": True}

# ==================== SETTINGS (ADMIN) ====================

@app.get("/api/admin/settings")
async def get_admin_settings(request: Request, user: dict = Depends(require_super_admin())):
    return await get_settings(request.app.db)

@app.put("/api/admin/settings")
async def update_admin_settings(request: Request, user: dict = Depends(require_super_admin())):
    db = request.app.db
    body = await request.json()
    allowed_keys = {
        "affiliate_commission_rate",
        "network1_generations", "network2_generations",
        "propaganda_threshold_referrals", "propaganda_threshold_period_days",
        "withdrawal_enabled", "withdrawal_min_amount", "withdrawal_release_days",
        "company_name", "company_cnpj", "company_address", "company_city",
        "company_state", "company_zip", "company_phone", "company_email",
        "invoice_prefix",
        "email_enabled", "resend_api_key", "email_from", "email_admin_recipients",
        "order_invoice_email_to",
        "email_trigger_order_created", "email_trigger_order_paid",
        "email_trigger_order_shipped", "email_trigger_order_delivered",
        "email_trigger_commission_earned", "email_trigger_admin_new_candidate",
        "email_trigger_admin_new_order", "email_trigger_welcome",
    }
    update = {k: v for k, v in body.items() if k in allowed_keys}
    # Sanitizar generations (garantir lista de 6 floats)
    for key in ("network1_generations", "network2_generations"):
        if key in update:
            arr = update[key]
            if not isinstance(arr, list):
                raise HTTPException(status_code=400, detail=f"{key} deve ser uma lista")
            arr = [float(x or 0) for x in arr[:6]]
            while len(arr) < 6:
                arr.append(0.0)
            update[key] = arr
    update["updated_at"] = now_iso()
    await db.settings.update_one({"_id": "global"}, {"$set": update}, upsert=True)
    return await get_settings(db)

# ==================== MY NETWORK (USER) ====================

@app.get("/api/users/me/network")
async def my_network(request: Request, user: dict = Depends(get_current_user)):
    """Retorna a rede MMN do usuario (ate 6 geracoes abaixo) com stats."""
    db = request.app.db
    settings = await get_settings(db)
    network_type = user.get("network_type", NETWORK_CUSTOMER)
    gens_pct = settings.get(f"{'network1' if network_type == NETWORK_1 else 'network2'}_generations", []) if network_type in (NETWORK_1, NETWORK_2) else []

    generations = []
    if network_type in (NETWORK_1, NETWORK_2):
        current_ids = [user["user_id"]]
        seen = {user["user_id"]}
        for gen in range(1, 7):
            # Iter 31: BFS unificado via sponsor_id OU network_sponsor_id (dedup por user_id).
            # Reflete a nova regra de comissao MMN que sobe por ambas as cadeias.
            level_users = await db.users.find(
                {"$or": [
                    {"sponsor_id": {"$in": current_ids}},
                    {"network_sponsor_id": {"$in": current_ids}},
                ], "user_id": {"$nin": list(seen)}},
                {"_id": 0, "password_hash": 0}
            ).to_list(2000)
            # Stats de comissao da geracao
            level_user_ids = [u["user_id"] for u in level_users]
            comm_agg = await db.commissions.aggregate([
                {"$match": {"user_id": user["user_id"], "generation": gen, "network_type": network_type}},
                {"$group": {"_id": "$status", "total": {"$sum": "$amount"}, "count": {"$sum": 1}}}
            ]).to_list(10)
            stats = {"pending": 0, "paid": 0, "count": 0}
            for s in comm_agg:
                stats[s["_id"] or "pending"] = round(s["total"], 2)
                stats["count"] += s["count"]
            # Lista resumida dos membros para exibir nominalmente
            members_list = [
                {
                    "user_id": m.get("user_id"),
                    "name": m.get("name"),
                    "email": m.get("email"),
                    "external_id": m.get("external_id"),
                    "network_type": m.get("network_type") or NETWORK_CUSTOMER,
                    "created_at": m.get("created_at"),
                    "referral_program_active": bool(m.get("referral_program_active")),
                }
                for m in level_users
            ]
            generations.append({
                "generation": gen,
                "rate_pct": gens_pct[gen - 1] if gen - 1 < len(gens_pct) else 0,
                "members_count": len(level_users),
                "members": members_list,
                "pending": stats.get("pending", 0),
                "paid": stats.get("paid", 0),
                "total_commissions": stats.get("count", 0),
            })
            current_ids = level_user_ids
            seen.update(level_user_ids)
            if not current_ids:
                # Sem mais descendentes - preencher zeros ate 6
                for g in range(gen + 1, 7):
                    generations.append({
                        "generation": g, "rate_pct": gens_pct[g - 1] if g - 1 < len(gens_pct) else 0,
                        "members_count": 0, "members": [], "pending": 0, "paid": 0, "total_commissions": 0,
                    })
                break

    # Totais gerais do usuario (afiliado + MMN)
    totals_agg = await db.commissions.aggregate([
        {"$match": {"user_id": user["user_id"]}},
        {"$group": {"_id": "$status", "total": {"$sum": "$amount"}}}
    ]).to_list(10)
    totals = {"pending": 0, "paid": 0, "cancelled": 0}
    for s in totals_agg:
        totals[s["_id"] or "pending"] = round(s["total"], 2)

    # Iter 35: separar por origem (afiliado / Equipe 1 / Equipe 2)
    by_source_agg = await db.commissions.aggregate([
        {"$match": {"user_id": user["user_id"]}},
        {"$group": {
            "_id": {"type": "$type", "network_type": "$network_type", "status": "$status"},
            "total": {"$sum": "$amount"},
            "count": {"$sum": 1},
        }},
    ]).to_list(50)
    by_source = {
        "affiliate": {"pending": 0, "paid": 0, "count": 0},
        "network_1": {"pending": 0, "paid": 0, "count": 0},
        "network_2": {"pending": 0, "paid": 0, "count": 0},
    }
    for s in by_source_agg:
        key = "affiliate" if s["_id"].get("type") == "affiliate" else (s["_id"].get("network_type") or "network_1")
        if key not in by_source:
            by_source[key] = {"pending": 0, "paid": 0, "count": 0}
        st = s["_id"].get("status") or "pending"
        if st in ("pending", "paid"):
            by_source[key][st] = round(by_source[key].get(st, 0) + (s["total"] or 0), 2)
        by_source[key]["count"] += s["count"]

    return {
        "network_type": network_type,
        "referral_code": user.get("referral_code"),
        "generations": generations,
        "totals": totals,
        "by_source": by_source,
        "commission_rate_affiliate": settings["affiliate_commission_rate"],
    }

# ==================== ADMIN: USERS BY NETWORK ====================

@app.get("/api/admin/users-by-network")
async def admin_users_by_network(request: Request, network_type: str, search: Optional[str] = None, page: int = 1, limit: int = 50, user: dict = Depends(require_admin())):
    db = request.app.db
    if network_type not in (NETWORK_CUSTOMER, NETWORK_1, NETWORK_2):
        raise HTTPException(status_code=400, detail="network_type invalido")
    q = {"network_type": network_type}
    if search:
        q["$or"] = [{"name": {"$regex": search, "$options": "i"}}, {"email": {"$regex": search, "$options": "i"}}, {"external_id": search}]
    total = await db.users.count_documents(q)
    users = await db.users.find(q, {"_id": 0, "password_hash": 0}).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    return {"users": users, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit), "limit": limit}

@app.get("/api/admin/users/{user_id}/tree")
async def admin_user_tree(request: Request, user_id: str, user: dict = Depends(require_admin())):
    """Retorna ate 6 geracoes abaixo do usuario, respeitando sua rede."""
    db = request.app.db
    root = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    if not root:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    network_type = root.get("network_type", NETWORK_CUSTOMER)
    generations = []
    current_ids = [user_id]
    for gen in range(1, 7):
        level = await db.users.find(
            {"network_sponsor_id": {"$in": current_ids}, "network_type": network_type},
            {"_id": 0, "password_hash": 0}
        ).to_list(2000)
        generations.append({"generation": gen, "users": level})
        current_ids = [u["user_id"] for u in level]
        if not current_ids:
            break
    return {"root": root, "generations": generations}

# ==================== ADMIN: PROMOTE TO PROPAGANDISTA (NETWORK_2) ====================

@app.post("/api/admin/users/{user_id}/promote-to-propagandista")
async def promote_to_propagandista(request: Request, user_id: str, user: dict = Depends(require_admin())):
    db = request.app.db
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    if target.get("network_type") != NETWORK_CUSTOMER:
        raise HTTPException(status_code=400, detail="Somente clientes podem ser promovidos a Propagandista")
    # network_sponsor_id = sponsor_id atual (se existir E for da rede ou for outro Propagandista)
    network_sponsor_id = None
    if target.get("sponsor_id"):
        sp = await db.users.find_one({"user_id": target["sponsor_id"]}, {"_id": 0})
        if sp and sp.get("network_type") == NETWORK_2:
            network_sponsor_id = sp["user_id"]
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {
            "network_type": NETWORK_2,
            "network_sponsor_id": network_sponsor_id,
            "promoted_at": now_iso(),
        }},
    )
    return await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})

@app.post("/api/admin/users/{user_id}/revoke-network")
async def revoke_network(request: Request, user_id: str, user: dict = Depends(require_admin())):
    """Reverte usuario network_2 para customer (so permite em network_2)."""
    db = request.app.db
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    if target.get("network_type") != NETWORK_2:
        raise HTTPException(status_code=400, detail="So eh possivel reverter Propagandistas (network_2)")
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"network_type": NETWORK_CUSTOMER, "network_sponsor_id": None, "revoked_at": now_iso()}},
    )
    return {"message": "Rede removida"}

# ==================== ADMIN: CANDIDATES "EM ALTA" PARA PROPAGANDISTA ====================

@app.get("/api/admin/propaganda-candidates")
async def propaganda_candidates(request: Request, user: dict = Depends(require_admin())):
    """Lista clientes (network_type=customer) com volume de indicacoes acima do threshold configurado."""
    db = request.app.db
    settings = await get_settings(db)
    threshold = settings.get("propaganda_threshold_referrals", 5)
    days = settings.get("propaganda_threshold_period_days", 30)
    since = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()

    # Contar indicados criados na janela por sponsor_id
    pipeline = [
        {"$match": {"created_at": {"$gte": since}, "sponsor_id": {"$ne": None}}},
        {"$group": {"_id": "$sponsor_id", "count": {"$sum": 1}}},
        {"$match": {"count": {"$gte": threshold}}},
        {"$sort": {"count": -1}},
        {"$limit": 100},
    ]
    agg = await db.users.aggregate(pipeline).to_list(100)
    candidate_ids = [a["_id"] for a in agg]
    candidates = await db.users.find(
        {"user_id": {"$in": candidate_ids}, "network_type": NETWORK_CUSTOMER},
        {"_id": 0, "password_hash": 0}
    ).to_list(100)
    # Enriquecer com count
    counts = {a["_id"]: a["count"] for a in agg}
    for c in candidates:
        c["referrals_in_period"] = counts.get(c["user_id"], 0)
    candidates.sort(key=lambda x: x["referrals_in_period"], reverse=True)
    return {"candidates": candidates, "threshold": threshold, "period_days": days}

# ==================== ADMIN: IMPORT NETWORK 1 (EXCEL/CSV) ====================

class Network1ImportRow(BaseModel):
    external_id: str
    name: str
    email: EmailStr
    leader_external_id: Optional[str] = None
    phone: Optional[str] = None
    cpf: Optional[str] = None

class Network1ImportPayload(BaseModel):
    rows: List[Network1ImportRow]
    default_password: Optional[str] = "oxx@pharma"


def _clean_cpf(cpf: Optional[str]) -> Optional[str]:
    """Normaliza CPF removendo tudo que nao for digito."""
    if not cpf:
        return None
    digits = "".join(ch for ch in str(cpf) if ch.isdigit())
    return digits or None


def _clean_phone(phone: Optional[str]) -> Optional[str]:
    """Normaliza telefone para apenas digitos. Remove DDI 55 quando presente."""
    if not phone:
        return None
    digits = "".join(ch for ch in str(phone) if ch.isdigit())
    if not digits:
        return None
    # Remove DDI 55 do BR para normalizar (ex: 5511999999999 -> 11999999999)
    if len(digits) > 11 and digits.startswith("55"):
        digits = digits[2:]
    return digits or None


async def find_existing_user_for_sync(db, external_id: Optional[str], email: Optional[str], cpf: Optional[str]) -> Optional[Dict]:
    """Prioriza match por external_id (ja vinculado) -> CPF normalizado -> email.
    Isso permite vincular automaticamente usuarios que se cadastraram antes da integracao Maxx.
    """
    if external_id:
        doc = await db.users.find_one({"external_id": external_id})
        if doc:
            return doc
    clean = _clean_cpf(cpf)
    if clean:
        # Match tentando variar formato: digits puros OU com mascara
        doc = await db.users.find_one({"$or": [{"cpf": clean}, {"cpf_digits": clean}]})
        if doc:
            return doc
        # Fallback: varre docs que contenham os digitos (usuarios antigos podem ter CPF com mascara)
        import re
        doc = await db.users.find_one({"cpf": {"$regex": re.escape(clean)}})
        if doc:
            return doc
    if email:
        return await db.users.find_one({"email": str(email).lower()})
    return None

async def resolve_leader_links(db, rows, id_map: dict) -> dict:
    """Resolve leader_external_id -> network_sponsor_id para os usuarios processados
    e tambem para qualquer usuario ja persistido com leader_external_id pendente.

    Retorna stats: {sponsors_mapped, sponsors_pending, leader_external_persisted}.

    Cada `row` deve ter atributos `external_id` e `leader_external_id` (Pydantic ou dict-like).
    `id_map` mapeia external_id -> user_id (usuario sendo sincronizado neste batch).
    """
    sponsors_mapped = 0
    sponsors_pending = 0
    persisted = 0

    def _attr(r, name):
        if isinstance(r, dict):
            return r.get(name)
        return getattr(r, name, None)

    # Passo A: persistir leader_external_id em CADA usuario processado (mesmo sem resolver),
    # para podermos resolve-lo depois quando o lider for importado.
    for row in rows:
        ext_id = _attr(row, "external_id")
        leader_ext = _attr(row, "leader_external_id")
        uid = id_map.get(ext_id)
        if not uid:
            continue
        # Sempre persistir o leader_external_id (mesmo se vazio: limpa link antigo)
        await db.users.update_one(
            {"user_id": uid},
            {"$set": {"leader_external_id": leader_ext or None}},
        )
        persisted += 1

    # Passo B: tentar resolver para os usuarios deste batch
    for row in rows:
        ext_id = _attr(row, "external_id")
        leader_ext = _attr(row, "leader_external_id")
        uid = id_map.get(ext_id)
        if not uid or not leader_ext:
            continue
        leader_uid = id_map.get(leader_ext)
        if not leader_uid:
            leader_doc = await db.users.find_one({"external_id": leader_ext}, {"_id": 0, "user_id": 1})
            if leader_doc:
                leader_uid = leader_doc["user_id"]
        if leader_uid:
            await db.users.update_one(
                {"user_id": uid},
                {"$set": {"network_sponsor_id": leader_uid}},
            )
            sponsors_mapped += 1
        else:
            sponsors_pending += 1

    # Passo C: para CADA novo external_id criado/atualizado neste batch,
    # tentar resolver outros usuarios na base que estavam aguardando este lider.
    for ext_id, uid in id_map.items():
        if not ext_id:
            continue
        result = await db.users.update_many(
            {
                "leader_external_id": ext_id,
                "$or": [
                    {"network_sponsor_id": None},
                    {"network_sponsor_id": {"$exists": False}},
                    {"network_sponsor_id": ""},
                ],
            },
            {"$set": {"network_sponsor_id": uid}},
        )
        sponsors_mapped += (result.modified_count or 0)

    return {
        "sponsors_mapped": sponsors_mapped,
        "sponsors_pending": sponsors_pending,
        "leader_external_persisted": persisted,
    }


@app.post("/api/admin/network1/import")
async def import_network1(request: Request, data: Network1ImportPayload, user: dict = Depends(require_admin())):
    """Importa/atualiza usuarios da Rede 1 a partir de uma lista (vinda de upload Excel/CSV).

    Passo 1: upsert de todos os usuarios (sem network_sponsor_id ainda).
    Passo 2: mapeia external_id -> user_id e define network_sponsor_id.
    """
    db = request.app.db
    default_password = data.default_password or "oxx@pharma"
    pw_hash = hash_pw(default_password)

    # Passo 1: upsert usuarios
    stats = {"created": 0, "updated": 0, "linked_by_cpf": 0, "total": len(data.rows), "errors": []}
    id_map = {}  # external_id -> user_id
    for row in data.rows:
        try:
            existing = await find_existing_user_for_sync(db, row.external_id, str(row.email), row.cpf)
            if existing:
                match_cpf = _clean_cpf(row.cpf)
                existing_ext = existing.get("external_id")
                linked_by_cpf_flag = bool(match_cpf and not existing_ext and match_cpf == _clean_cpf(existing.get("cpf")))
                update_fields = {
                    "external_id": row.external_id,
                    "name": row.name,
                    "phone": row.phone,
                    "network_type": NETWORK_1,
                    "leader_external_id": row.leader_external_id or None,
                }
                if row.cpf:
                    update_fields["cpf"] = row.cpf
                    if match_cpf:
                        update_fields["cpf_digits"] = match_cpf
                # Atualiza email SOMENTE se o existente nao tem email (ou se eh o mesmo) - evita sobrescrever login do usuario
                if not existing.get("email") or existing.get("email") == str(row.email).lower():
                    update_fields["email"] = str(row.email).lower()
                await db.users.update_one(
                    {"user_id": existing["user_id"]},
                    {"$set": update_fields},
                )
                id_map[row.external_id] = existing["user_id"]
                stats["updated"] += 1
                if linked_by_cpf_flag:
                    stats["linked_by_cpf"] += 1
            else:
                uid = gen_id("user_")
                cpf_clean = _clean_cpf(row.cpf)
                await db.users.insert_one({
                    "user_id": uid, "email": row.email.lower(),
                    "password_hash": pw_hash, "name": row.name,
                    "phone": row.phone, "role": "customer", "access_level": 99,
                    "status": "active", "addresses": [],
                    "sponsor_id": None, "sponsor_code": None,
                    "network_type": NETWORK_1,
                    "network_sponsor_id": None,
                    "external_id": row.external_id,
                    "leader_external_id": row.leader_external_id or None,
                    "cpf": row.cpf or None,
                    "cpf_digits": cpf_clean,
                    "must_set_password": True,
                    "referral_program_active": False,
                    "created_at": now_iso(),
                })
                id_map[row.external_id] = uid
                stats["created"] += 1
        except Exception as e:
            stats["errors"].append({"external_id": row.external_id, "error": str(e)})

    # Passo 2: resolver lideres (helper centralizado)
    link_stats = await resolve_leader_links(db, data.rows, id_map)
    stats.update(link_stats)
    return stats

# ==================== ADMIN: COMMISSIONS REPORT (BENEFIT CARD) ====================

@app.get("/api/admin/commissions-by-generation")
async def admin_commissions_by_generation(request: Request, status: Optional[str] = None, start: Optional[str] = None, end: Optional[str] = None, page: int = 1, limit: int = 20, user: dict = Depends(require_admin())):
    """Iter 39: Relatorio visual por geracao.

    Retorna:
      - summary_by_generation: [{generation, label, network_type, total_amount, count, avg_rate_pct, beneficiaries}]
      - recent_orders: ultimos 30 pedidos com a cadeia completa de comissoes geradas
        (para validar visualmente se cada geracao recebeu o %% correto)
      - totals: {total_amount, total_count, orders_with_commission}
    """
    db = request.app.db
    match = {}
    if status:
        match["status"] = status
    if start:
        match.setdefault("created_at", {})["$gte"] = start
    if end:
        match.setdefault("created_at", {})["$lte"] = end

    # Agrega por (generation, type, network_type)
    pipeline = [
        {"$match": match},
        {"$group": {
            "_id": {
                "generation": "$generation",
                "type": "$type",
                "network_type": "$network_type",
            },
            "total_amount": {"$sum": "$amount"},
            "count": {"$sum": 1},
            "avg_rate": {"$avg": "$rate"},
            "beneficiaries": {"$addToSet": "$user_id"},
        }},
        {"$sort": {"_id.generation": 1, "_id.network_type": 1}},
    ]
    agg = await db.commissions.aggregate(pipeline).to_list(100)
    summary = []
    total_amount_all = 0.0
    total_count_all = 0
    for a in agg:
        gen = a["_id"].get("generation", 0)
        typ = a["_id"].get("type")
        net = a["_id"].get("network_type")
        if typ == "affiliate":
            label = "Afiliado Direto (cliente indicador)"
        else:
            net_label = "Equipe 1" if net == "network_1" else ("Equipe 2" if net == "network_2" else (net or "—"))
            label = f"{gen}ª geração · {net_label}"
        amt = round(a["total_amount"] or 0, 2)
        total_amount_all += amt
        total_count_all += a["count"]
        summary.append({
            "generation": gen,
            "type": typ,
            "network_type": net,
            "label": label,
            "total_amount": amt,
            "count": a["count"],
            "avg_rate_pct": round((a["avg_rate"] or 0) * 100, 3),
            "beneficiaries_count": len(a.get("beneficiaries") or []),
        })

    # Iter 41: pagina os pedidos com comissao
    recent_order_ids = await db.commissions.distinct("order_id", match)
    total_orders = len(recent_order_ids)
    skip = max(0, (page - 1) * limit)
    if recent_order_ids:
        orders = await db.orders.find(
            {"order_id": {"$in": recent_order_ids}},
            {"_id": 0, "order_id": 1, "user_id": 1, "subtotal": 1, "total": 1,
             "created_at": 1, "payment_status": 1, "order_number": 1, "invoice_number": 1}
        ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    else:
        orders = []

    # Para cada pedido, busca todas as comissoes + hidrata usuarios
    order_ids = [o["order_id"] for o in orders]
    comms_cursor = db.commissions.find(
        {"order_id": {"$in": order_ids}, **match},
        {"_id": 0}
    ).sort("generation", 1)
    comms_list = await comms_cursor.to_list(5000)
    # Hidrata nomes/emails
    user_ids = list({c.get("user_id") for c in comms_list if c.get("user_id")} |
                    {o.get("user_id") for o in orders if o.get("user_id")})
    users_map = {}
    if user_ids:
        for u in await db.users.find({"user_id": {"$in": user_ids}}, {"_id": 0, "user_id": 1, "name": 1, "email": 1, "network_type": 1, "referral_code": 1}).to_list(len(user_ids)):
            users_map[u["user_id"]] = u

    recent_orders = []
    for o in orders:
        chain = [c for c in comms_list if c["order_id"] == o["order_id"]]
        # Ordena: afiliado (gen 0, type=affiliate) primeiro, depois demais por geracao
        chain.sort(key=lambda c: (
            0 if (c.get("type") == "affiliate") else (c.get("generation") or 99),
            c.get("generation") or 99,
            c.get("network_type") or "",
        ))
        chain_rows = []
        total_chain = 0.0
        for c in chain:
            u = users_map.get(c.get("user_id")) or {}
            amt = round(c.get("amount") or 0, 2)
            total_chain += amt
            chain_rows.append({
                "commission_id": c.get("commission_id"),
                "generation": c.get("generation"),
                "type": c.get("type"),
                "network_type": c.get("network_type"),
                "beneficiary_id": c.get("user_id"),
                "beneficiary_name": u.get("name") or "—",
                "beneficiary_email": u.get("email"),
                "beneficiary_network": u.get("network_type"),
                "amount": amt,
                "rate_pct": round((c.get("rate") or 0) * 100, 3),
                "status": c.get("status"),
                "retroactive": bool(c.get("retroactive")),
                "program_active_at_creation": c.get("program_active_at_creation"),
                "beneficiary_enrolled_now": bool(u.get("referral_program_active")),
            })
        customer = users_map.get(o.get("user_id")) or {}
        recent_orders.append({
            "order_id": o.get("order_id"),
            "order_number": o.get("order_number"),
            "invoice_number": o.get("invoice_number"),
            "created_at": o.get("created_at"),
            "payment_status": o.get("payment_status"),
            "subtotal": round(o.get("subtotal") or 0, 2),
            "total": round(o.get("total") or 0, 2),
            "customer_id": o.get("user_id"),
            "customer_name": customer.get("name") or "—",
            "customer_email": customer.get("email"),
            "commission_total": round(total_chain, 2),
            "chain": chain_rows,
        })

    return {
        "summary_by_generation": summary,
        "totals": {
            "total_amount": round(total_amount_all, 2),
            "total_count": total_count_all,
            "orders_with_commission": total_orders,
        },
        "recent_orders": recent_orders,
        "page": page,
        "limit": limit,
        "total_orders": total_orders,
        "pages": max(1, (total_orders + limit - 1) // limit),
        "filter": {"status": status, "start": start, "end": end},
    }


@app.get("/api/admin/commissions-report")
async def admin_commissions_report(request: Request, status: str = "paid", start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(require_admin())):
    """Relatorio agregado por usuario para envio a empresa de cartao de beneficios.
    Inclui breakdown por origem (Iter 35): Indicacao (afiliado), Equipe 1, Equipe 2.
    """
    db = request.app.db
    match = {"status": status}
    if start:
        match.setdefault("created_at", {})["$gte"] = start
    if end:
        match.setdefault("created_at", {})["$lte"] = end
    # Agrupa por (user_id, origem) onde origem = (type + network_type)
    pipeline = [
        {"$match": match},
        {"$group": {
            "_id": {
                "user_id": "$user_id",
                "type": "$type",
                "network_type": "$network_type",
            },
            "total": {"$sum": "$amount"},
            "count": {"$sum": 1},
        }},
    ]
    agg = await db.commissions.aggregate(pipeline).to_list(20000)

    # Combina por user_id em uma estrutura final
    by_user = {}
    for a in agg:
        uid = a["_id"]["user_id"]
        if not uid:
            continue
        row = by_user.setdefault(uid, {
            "user_id": uid, "amount": 0.0, "commissions_count": 0,
            "by_source": {
                "affiliate": {"amount": 0.0, "count": 0},
                "network_1": {"amount": 0.0, "count": 0},
                "network_2": {"amount": 0.0, "count": 0},
            },
        })
        amt = round(a["total"] or 0, 2)
        cnt = a["count"]
        row["amount"] = round(row["amount"] + amt, 2)
        row["commissions_count"] += cnt
        if a["_id"].get("type") == "affiliate":
            row["by_source"]["affiliate"]["amount"] = round(row["by_source"]["affiliate"]["amount"] + amt, 2)
            row["by_source"]["affiliate"]["count"] += cnt
        else:
            net = a["_id"].get("network_type") or "network_1"
            if net not in row["by_source"]:
                row["by_source"][net] = {"amount": 0.0, "count": 0}
            row["by_source"][net]["amount"] = round(row["by_source"][net]["amount"] + amt, 2)
            row["by_source"][net]["count"] += cnt

    # Hidrata dados dos usuarios
    uids = list(by_user.keys())
    users = await db.users.find({"user_id": {"$in": uids}}, {"_id": 0, "password_hash": 0}).to_list(len(uids))
    umap = {u["user_id"]: u for u in users}
    rows = []
    for uid, r in by_user.items():
        u = umap.get(uid, {})
        bs = r["by_source"]
        rows.append({
            "user_id": uid,
            "cpf": u.get("cpf"),
            "name": u.get("name"),
            "email": u.get("email"),
            "pix_key": u.get("pix_key"),
            "amount": r["amount"],
            "commissions_count": r["commissions_count"],
            "amount_affiliate": bs["affiliate"]["amount"],
            "count_affiliate": bs["affiliate"]["count"],
            "amount_network_1": bs["network_1"]["amount"],
            "count_network_1": bs["network_1"]["count"],
            "amount_network_2": bs["network_2"]["amount"],
            "count_network_2": bs["network_2"]["count"],
        })
    rows.sort(key=lambda r: -r["amount"])

    # Totais agregados (KPIs do relatorio)
    totals = {
        "total": round(sum(r["amount"] for r in rows), 2),
        "by_source": {
            "affiliate": round(sum(r["amount_affiliate"] for r in rows), 2),
            "network_1": round(sum(r["amount_network_1"] for r in rows), 2),
            "network_2": round(sum(r["amount_network_2"] for r in rows), 2),
        },
        "users_count": len(rows),
        "commissions_count": sum(r["commissions_count"] for r in rows),
    }
    return {"rows": rows, "status": status, "period": {"start": start, "end": end}, "totals": totals}


_COMMISSIONS_COLS = [
    {"key": "user_id", "label": "user_id", "width": 22},
    {"key": "cpf", "label": "CPF", "width": 16},
    {"key": "name", "label": "Nome", "width": 28},
    {"key": "email", "label": "Email", "width": 30},
    {"key": "pix_key", "label": "Chave PIX", "width": 24},
    {"key": "amount", "label": "Valor total (R$)", "type": "money", "width": 16},
    {"key": "amount_affiliate", "label": "Indicação (R$)", "type": "money", "width": 16},
    {"key": "amount_network_1", "label": "Equipe 1 (R$)", "type": "money", "width": 14},
    {"key": "amount_network_2", "label": "Equipe 2 (R$)", "type": "money", "width": 14},
    {"key": "commissions_count", "label": "Qtd. comissões", "type": "int", "width": 14},
]


async def _commissions_rows(db, status: str, start: Optional[str], end: Optional[str]):
    """Linhas para export, agora incluindo amount_affiliate, amount_network_1, amount_network_2."""
    match = {"status": status}
    if start:
        match.setdefault("created_at", {})["$gte"] = start
    if end:
        match.setdefault("created_at", {})["$lte"] = end
    pipeline = [
        {"$match": match},
        {"$group": {
            "_id": {"user_id": "$user_id", "type": "$type", "network_type": "$network_type"},
            "total": {"$sum": "$amount"},
            "count": {"$sum": 1},
        }},
    ]
    agg = await db.commissions.aggregate(pipeline).to_list(20000)
    by_user = {}
    for a in agg:
        uid = a["_id"]["user_id"]
        if not uid:
            continue
        row = by_user.setdefault(uid, {
            "user_id": uid, "amount": 0.0, "commissions_count": 0,
            "amount_affiliate": 0.0, "amount_network_1": 0.0, "amount_network_2": 0.0,
        })
        amt = round(a["total"] or 0, 2)
        row["amount"] = round(row["amount"] + amt, 2)
        row["commissions_count"] += a["count"]
        if a["_id"].get("type") == "affiliate":
            row["amount_affiliate"] = round(row["amount_affiliate"] + amt, 2)
        else:
            net = a["_id"].get("network_type") or "network_1"
            if net == "network_1":
                row["amount_network_1"] = round(row["amount_network_1"] + amt, 2)
            elif net == "network_2":
                row["amount_network_2"] = round(row["amount_network_2"] + amt, 2)
    uids = list(by_user.keys())
    users = await db.users.find({"user_id": {"$in": uids}}, {"_id": 0, "password_hash": 0}).to_list(len(uids))
    umap = {u["user_id"]: u for u in users}
    out = []
    for uid, r in by_user.items():
        u = umap.get(uid, {})
        out.append({
            "user_id": uid,
            "cpf": u.get("cpf"),
            "name": u.get("name"),
            "email": u.get("email"),
            "pix_key": u.get("pix_key"),
            "amount": r["amount"],
            "amount_affiliate": r["amount_affiliate"],
            "amount_network_1": r["amount_network_1"],
            "amount_network_2": r["amount_network_2"],
            "commissions_count": r["commissions_count"],
        })
    out.sort(key=lambda x: -x["amount"])
    return out


@app.get("/api/admin/commissions-report/export.csv")
async def admin_commissions_report_csv(request: Request, status: str = "paid", start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(require_admin())):
    import export_utils
    rows = await _commissions_rows(request.app.db, status, start, end)
    return export_utils.csv_response(export_utils.make_csv(rows, _COMMISSIONS_COLS), f"comissoes-{status}.csv")


@app.get("/api/admin/commissions-report/export.xlsx")
async def admin_commissions_report_xlsx(request: Request, status: str = "paid", start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(require_admin())):
    import export_utils
    rows = await _commissions_rows(request.app.db, status, start, end)
    return export_utils.xlsx_response(export_utils.make_xlsx(rows, _COMMISSIONS_COLS, sheet_name="Comissoes"), f"comissoes-{status}.xlsx")

# ==================== INVOICES (FATURAMENTO INTERNO) ====================

async def issue_invoice(db, order: dict):
    """Incrementa contador de notas e retorna {number, issued_at}. Nao persiste no order aqui."""
    settings = await get_settings(db)
    prefix = settings.get("invoice_prefix") or "OXX"
    # Atomico: incrementar contador
    result = await db.settings.find_one_and_update(
        {"_id": "global"},
        {"$inc": {"invoice_counter": 1}},
        return_document=True,
        upsert=True,
    )
    counter = (result or {}).get("invoice_counter", 1)
    number = f"{prefix}-{str(counter).zfill(6)}"
    return {"number": number, "issued_at": now_iso(), "counter": counter}

async def build_invoice_data(db, order: dict):
    """Monta o objeto completo de nota (empresa + pedido) para exibicao/impressao."""
    settings = await get_settings(db)
    company = {
        "name": settings.get("company_name"),
        "cnpj": settings.get("company_cnpj"),
        "address": settings.get("company_address"),
        "city": settings.get("company_city"),
        "state": settings.get("company_state"),
        "zip": settings.get("company_zip"),
        "phone": settings.get("company_phone"),
        "email": settings.get("company_email"),
    }
    # Busca CPF do usuario se existir
    buyer = await db.users.find_one({"user_id": order.get("user_id")}, {"_id": 0, "password_hash": 0}) or {}
    return {
        "invoice_number": order.get("invoice_number"),
        "invoice_issued_at": order.get("invoice_issued_at"),
        "order": order,
        "company": company,
        "buyer": {
            "name": order.get("customer_name") or buyer.get("name"),
            "email": order.get("customer_email") or buyer.get("email"),
            "cpf": buyer.get("cpf"),
            "phone": buyer.get("phone"),
        },
    }

@app.get("/api/orders/{order_id}/invoice")
async def get_invoice(request: Request, order_id: str, user: dict = Depends(get_current_user)):
    """Usuario/admin consulta dados da nota do pedido."""
    db = request.app.db
    o = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    if not o:
        raise HTTPException(status_code=404, detail="Pedido nao encontrado")
    is_admin = user.get("role") == "admin" or user.get("access_level", 99) <= 1
    if not is_admin and o.get("user_id") != user["user_id"]:
        raise HTTPException(status_code=403, detail="Acesso negado")
    if not o.get("invoice_number"):
        raise HTTPException(status_code=404, detail="Nota ainda nao emitida. Pedido precisa estar pago.")
    return await build_invoice_data(db, o)

@app.post("/api/admin/orders/{order_id}/issue-invoice")
async def admin_issue_invoice(request: Request, order_id: str, user: dict = Depends(require_admin())):
    """Admin emite manualmente (caso pedido ja esteja pago sem nota, ou precise re-emitir)."""
    db = request.app.db
    o = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    if not o:
        raise HTTPException(status_code=404, detail="Pedido nao encontrado")
    if o.get("order_status") not in ("paid", "shipped", "delivered"):
        raise HTTPException(status_code=400, detail="Pedido precisa estar pago para emitir nota")
    if o.get("invoice_number"):
        raise HTTPException(status_code=400, detail=f"Nota ja emitida: {o['invoice_number']}")
    inv = await issue_invoice(db, o)
    await db.orders.update_one(
        {"order_id": order_id},
        {"$set": {"invoice_number": inv["number"], "invoice_issued_at": inv["issued_at"]}},
    )
    o = await db.orders.find_one({"order_id": order_id}, {"_id": 0})
    return await build_invoice_data(db, o)

@app.get("/api/admin/invoices")
async def admin_list_invoices(request: Request, search: Optional[str] = None, page: int = 1, limit: int = 50, user: dict = Depends(require_admin())):
    """Lista pedidos com nota emitida (faturamento interno)."""
    db = request.app.db
    q = {"invoice_number": {"$ne": None}}
    if search:
        q["$or"] = [
            {"invoice_number": {"$regex": search, "$options": "i"}},
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"customer_email": {"$regex": search, "$options": "i"}},
        ]
    total = await db.orders.count_documents(q)
    orders = await db.orders.find(q, {"_id": 0}).sort("invoice_issued_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    # Agregado faturado
    agg = await db.orders.aggregate([
        {"$match": {"invoice_number": {"$ne": None}}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}, "subtotal": {"$sum": "$subtotal"}}},
    ]).to_list(1)
    totals = {"total": round((agg[0]["total"] if agg else 0), 2), "subtotal": round((agg[0]["subtotal"] if agg else 0), 2), "count": total}
    return {"invoices": orders, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit), "totals": totals}


_INVOICE_COLS = [
    {"key": "invoice_number", "label": "Nº Nota", "width": 16},
    {"key": "invoice_issued_at", "label": "Emitida em", "width": 22},
    {"key": "order_id", "label": "Pedido", "width": 22},
    {"key": "customer_name", "label": "Cliente", "width": 28},
    {"key": "customer_email", "label": "Email", "width": 30},
    {"key": "subtotal", "label": "Subtotal", "type": "money", "width": 14},
    {"key": "shipping_cost", "label": "Frete", "type": "money", "width": 12},
    {"key": "discount_amount", "label": "Desconto", "type": "money", "width": 12},
    {"key": "total", "label": "Total", "type": "money", "width": 14},
    {"key": "payment_method", "label": "Pagamento", "width": 14},
]


async def _invoices_rows(db, search: Optional[str]):
    q = {"invoice_number": {"$ne": None}}
    if search:
        q["$or"] = [
            {"invoice_number": {"$regex": search, "$options": "i"}},
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"customer_email": {"$regex": search, "$options": "i"}},
        ]
    return await db.orders.find(q, {"_id": 0}).sort("invoice_issued_at", -1).to_list(10000)


@app.get("/api/admin/invoices/export.csv")
async def admin_invoices_export_csv(request: Request, search: Optional[str] = None, user: dict = Depends(require_admin())):
    import export_utils
    rows = await _invoices_rows(request.app.db, search)
    return export_utils.csv_response(export_utils.make_csv(rows, _INVOICE_COLS), "faturamento.csv")


@app.get("/api/admin/invoices/export.xlsx")
async def admin_invoices_export_xlsx(request: Request, search: Optional[str] = None, user: dict = Depends(require_admin())):
    import export_utils
    rows = await _invoices_rows(request.app.db, search)
    return export_utils.xlsx_response(export_utils.make_xlsx(rows, _INVOICE_COLS, sheet_name="Faturamento"), "faturamento.xlsx")

# ==================== WITHDRAWALS (SAQUES) ====================

async def compute_balance(db, user_id: str):
    """Calcula saldo disponivel, em quarentena, total pago, total sacado."""
    settings = await get_settings(db)
    release_days = int(settings.get("withdrawal_release_days") or 0)
    cutoff = (datetime.now(timezone.utc) - timedelta(days=release_days)).isoformat()

    # Comissoes pagas (status=paid) ainda nao vinculadas a saque
    # Considerar quarentena: paid_at <= cutoff => available; > cutoff => quarentine
    available_agg = await db.commissions.aggregate([
        {"$match": {"user_id": user_id, "status": "paid", "withdrawal_id": {"$in": [None, ""]}, "paid_at": {"$lte": cutoff}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    quarantine_agg = await db.commissions.aggregate([
        {"$match": {"user_id": user_id, "status": "paid", "withdrawal_id": {"$in": [None, ""]}, "paid_at": {"$gt": cutoff}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    # Saques
    withdrawn_agg = await db.withdrawals.aggregate([
        {"$match": {"user_id": user_id, "status": "paid_out"}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    reserved_agg = await db.withdrawals.aggregate([
        {"$match": {"user_id": user_id, "status": {"$in": ["pending", "approved"]}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    pending_commissions_agg = await db.commissions.aggregate([
        {"$match": {"user_id": user_id, "status": "pending"}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)

    return {
        "available": round((available_agg[0]["total"] if available_agg else 0), 2),
        "quarantine": round((quarantine_agg[0]["total"] if quarantine_agg else 0), 2),
        "pending_commissions": round((pending_commissions_agg[0]["total"] if pending_commissions_agg else 0), 2),
        "reserved_in_withdrawals": round((reserved_agg[0]["total"] if reserved_agg else 0), 2),
        "total_withdrawn": round((withdrawn_agg[0]["total"] if withdrawn_agg else 0), 2),
        "withdrawal_enabled": bool(settings.get("withdrawal_enabled")),
        "withdrawal_min_amount": float(settings.get("withdrawal_min_amount") or 0),
        "withdrawal_release_days": release_days,
    }

@app.get("/api/users/me/balance")
async def my_balance(request: Request, user: dict = Depends(get_current_user)):
    return await compute_balance(request.app.db, user["user_id"])

@app.post("/api/withdrawals")
async def create_withdrawal(request: Request, data: WithdrawalCreate, user: dict = Depends(get_current_user)):
    db = request.app.db
    settings = await get_settings(db)
    if not settings.get("withdrawal_enabled"):
        raise HTTPException(status_code=400, detail="Saques estao desativados no momento")

    min_amt = float(settings.get("withdrawal_min_amount") or 0)
    if data.amount < min_amt:
        raise HTTPException(status_code=400, detail=f"Valor minimo de saque: R$ {min_amt:.2f}")

    balance = await compute_balance(db, user["user_id"])
    if data.amount > balance["available"]:
        raise HTTPException(status_code=400, detail=f"Saldo disponivel insuficiente (R$ {balance['available']:.2f})")

    # Selecionar comissoes elegiveis (FIFO pelo paid_at) ate cobrir o valor
    release_days = int(settings.get("withdrawal_release_days") or 0)
    cutoff = (datetime.now(timezone.utc) - timedelta(days=release_days)).isoformat()
    candidates = await db.commissions.find(
        {"user_id": user["user_id"], "status": "paid", "withdrawal_id": {"$in": [None, ""]}, "paid_at": {"$lte": cutoff}},
        {"_id": 0}
    ).sort("paid_at", 1).to_list(10000)
    selected = []
    accum = 0
    for c in candidates:
        if accum >= data.amount:
            break
        selected.append(c["commission_id"])
        accum += c["amount"]
    if accum < data.amount:
        raise HTTPException(status_code=400, detail=f"Saldo disponivel insuficiente (R$ {balance['available']:.2f})")

    wid = gen_id("wd_")
    withdrawal = {
        "withdrawal_id": wid,
        "user_id": user["user_id"],
        "user_name": user.get("name"),
        "user_email": user.get("email"),
        "amount": round(float(data.amount), 2),
        "selected_amount": round(accum, 2),
        "commission_ids": selected,
        "pix_key": data.pix_key,
        "pix_key_type": data.pix_key_type,
        "pix_name": data.pix_name,
        "pix_cpf": data.pix_cpf,
        "notes": data.notes,
        "status": "pending",  # pending -> approved -> paid_out; ou rejected
        "created_at": now_iso(),
        "decided_at": None,
        "decided_by": None,
        "admin_notes": None,
        "paid_at": None,
    }
    await db.withdrawals.insert_one(withdrawal)
    # Marca comissoes como reservadas via withdrawal_id
    await db.commissions.update_many(
        {"commission_id": {"$in": selected}},
        {"$set": {"withdrawal_id": wid}},
    )
    # Atualiza perfil do usuario com dados PIX (conveniencia)
    await db.users.update_one(
        {"user_id": user["user_id"]},
        {"$set": {
            "pix_key": data.pix_key, "pix_key_type": data.pix_key_type,
            "cpf": data.pix_cpf,
        }},
    )
    return await db.withdrawals.find_one({"withdrawal_id": wid}, {"_id": 0})

@app.get("/api/users/me/withdrawals")
async def my_withdrawals(request: Request, page: int = 1, limit: int = 20, user: dict = Depends(get_current_user)):
    db = request.app.db
    q = {"user_id": user["user_id"]}
    total = await db.withdrawals.count_documents(q)
    items = await db.withdrawals.find(q, {"_id": 0}).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    return {"withdrawals": items, "total": total, "page": page}

@app.post("/api/users/me/withdrawals/{wid}/cancel")
async def cancel_withdrawal(request: Request, wid: str, user: dict = Depends(get_current_user)):
    db = request.app.db
    w = await db.withdrawals.find_one({"withdrawal_id": wid}, {"_id": 0})
    if not w:
        raise HTTPException(status_code=404, detail="Solicitacao nao encontrada")
    if w["user_id"] != user["user_id"]:
        raise HTTPException(status_code=403, detail="Acesso negado")
    if w["status"] != "pending":
        raise HTTPException(status_code=400, detail="Apenas solicitacoes pendentes podem ser canceladas")
    await db.withdrawals.update_one({"withdrawal_id": wid}, {"$set": {"status": "cancelled", "decided_at": now_iso()}})
    await db.commissions.update_many({"withdrawal_id": wid}, {"$set": {"withdrawal_id": None}})
    return {"message": "Solicitacao cancelada"}

# === ADMIN WITHDRAWALS ===

@app.get("/api/admin/withdrawals")
async def admin_list_withdrawals(request: Request, status: Optional[str] = None, search: Optional[str] = None, page: int = 1, limit: int = 50, user: dict = Depends(require_admin())):
    db = request.app.db
    q = {}
    if status:
        q["status"] = status
    if search:
        q["$or"] = [{"user_name": {"$regex": search, "$options": "i"}}, {"user_email": {"$regex": search, "$options": "i"}}, {"pix_cpf": search}]
    total = await db.withdrawals.count_documents(q)
    items = await db.withdrawals.find(q, {"_id": 0}).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    # Agregados
    agg = await db.withdrawals.aggregate([{"$group": {"_id": "$status", "total": {"$sum": "$amount"}, "count": {"$sum": 1}}}]).to_list(10)
    summary = {s["_id"]: {"total": round(s["total"], 2), "count": s["count"]} for s in agg}
    return {"withdrawals": items, "total": total, "page": page, "summary": summary}

@app.put("/api/admin/withdrawals/{wid}/approve")
async def admin_approve_withdrawal(request: Request, wid: str, user: dict = Depends(require_admin())):
    db = request.app.db
    w = await db.withdrawals.find_one({"withdrawal_id": wid})
    if not w:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    if w["status"] != "pending":
        raise HTTPException(status_code=400, detail="Status atual nao permite aprovacao")
    await db.withdrawals.update_one(
        {"withdrawal_id": wid},
        {"$set": {"status": "approved", "decided_at": now_iso(), "decided_by": user["user_id"]}},
    )
    return await db.withdrawals.find_one({"withdrawal_id": wid}, {"_id": 0})

@app.put("/api/admin/withdrawals/{wid}/reject")
async def admin_reject_withdrawal(request: Request, wid: str, user: dict = Depends(require_admin())):
    db = request.app.db
    body = await request.json()
    reason = (body or {}).get("reason", "")
    w = await db.withdrawals.find_one({"withdrawal_id": wid})
    if not w:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    if w["status"] not in ("pending", "approved"):
        raise HTTPException(status_code=400, detail="Status atual nao permite rejeicao")
    await db.withdrawals.update_one(
        {"withdrawal_id": wid},
        {"$set": {"status": "rejected", "decided_at": now_iso(), "decided_by": user["user_id"], "admin_notes": reason}},
    )
    # Devolve comissoes
    await db.commissions.update_many({"withdrawal_id": wid}, {"$set": {"withdrawal_id": None}})
    return await db.withdrawals.find_one({"withdrawal_id": wid}, {"_id": 0})

@app.put("/api/admin/withdrawals/{wid}/mark-paid")
async def admin_mark_paid(request: Request, wid: str, user: dict = Depends(require_admin())):
    db = request.app.db
    w = await db.withdrawals.find_one({"withdrawal_id": wid})
    if not w:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    if w["status"] not in ("approved", "pending"):
        raise HTTPException(status_code=400, detail="Status atual nao permite pagamento")
    now = now_iso()
    await db.withdrawals.update_one(
        {"withdrawal_id": wid},
        {"$set": {"status": "paid_out", "paid_at": now, "decided_by": user["user_id"], "decided_at": now}},
    )
    # Marcar comissoes vinculadas como paid_out
    await db.commissions.update_many(
        {"withdrawal_id": wid},
        {"$set": {"status": "paid_out", "paid_out_at": now}},
    )
    return await db.withdrawals.find_one({"withdrawal_id": wid}, {"_id": 0})

@app.get("/api/admin/withdrawals/export")
async def admin_export_withdrawals(request: Request, status: str = "approved", user: dict = Depends(require_admin())):
    """Exporta solicitacoes de saque (formato pronto para empresa de cartao de beneficios)."""
    db = request.app.db
    items = await db.withdrawals.find({"status": status}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    rows = [{
        "withdrawal_id": w["withdrawal_id"],
        "cpf": w.get("pix_cpf"),
        "name": w.get("pix_name") or w.get("user_name"),
        "email": w.get("user_email"),
        "pix_key_type": w.get("pix_key_type"),
        "pix_key": w.get("pix_key"),
        "amount": w["amount"],
        "created_at": w["created_at"],
    } for w in items]
    return {"rows": rows, "status": status, "count": len(rows)}


def _withdrawals_rows(items):
    return [{
        "withdrawal_id": w["withdrawal_id"],
        "cpf": w.get("pix_cpf"),
        "name": w.get("pix_name") or w.get("user_name"),
        "email": w.get("user_email"),
        "pix_key_type": w.get("pix_key_type"),
        "pix_key": w.get("pix_key"),
        "amount": float(w.get("amount", 0) or 0),
        "created_at": w["created_at"],
    } for w in items]


@app.get("/api/admin/withdrawals/export.csv")
async def admin_export_withdrawals_csv(request: Request, status: str = "approved", user: dict = Depends(require_admin())):
    from fastapi.responses import Response as FastAPIResponse
    import io as _io
    import csv as _csv
    db = request.app.db
    items = await db.withdrawals.find({"status": status}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    rows = _withdrawals_rows(items)
    buf = _io.StringIO()
    w = _csv.writer(buf)
    w.writerow(["withdrawal_id", "cpf", "name", "email", "pix_key_type", "pix_key", "amount", "created_at"])
    for r in rows:
        w.writerow([r["withdrawal_id"], r["cpf"], r["name"], r["email"], r["pix_key_type"], r["pix_key"], f"{r['amount']:.2f}", r["created_at"]])
    return FastAPIResponse(
        content=buf.getvalue().encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="saques-{status}.csv"'},
    )


@app.get("/api/admin/withdrawals/export.xlsx")
async def admin_export_withdrawals_xlsx(request: Request, status: str = "approved", user: dict = Depends(require_admin())):
    from fastapi.responses import Response as FastAPIResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io as _io
    db = request.app.db
    items = await db.withdrawals.find({"status": status}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    rows = _withdrawals_rows(items)

    wb = Workbook()
    ws = wb.active
    ws.title = "Saques"
    headers = ["ID", "CPF", "Nome", "Email", "Tipo Chave PIX", "Chave PIX", "Valor (R$)", "Criado em"]
    ws.append(headers)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="E8731A")
    for col_i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_i)
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r["withdrawal_id"], r["cpf"], r["name"], r["email"], r["pix_key_type"], r["pix_key"], r["amount"], r["created_at"]])
    widths = [22, 16, 28, 30, 16, 30, 14, 22]
    for i, wd in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = wd

    buf = _io.BytesIO()
    wb.save(buf)
    return FastAPIResponse(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="saques-{status}.xlsx"'},
    )

# ==================== EMAIL TEMPLATES (ADMIN) ====================

class EmailTemplateIn(BaseModel):
    slug: str
    name: str
    subject: str
    body_html: str
    body_text: Optional[str] = ""
    active: bool = True

class EmailBroadcast(BaseModel):
    subject: str
    body_html: str
    body_text: Optional[str] = ""
    # target: "all" | "network_1" | "network_2" | "customer" | "admin" | "user_ids"
    target: str = "all"
    user_ids: Optional[List[str]] = None
    emails: Optional[List[EmailStr]] = None  # envio direto para emails especificos

class EmailTestIn(BaseModel):
    to: EmailStr
    subject: str = "Teste de envio - OxxPharma"
    body_html: str = "<p>Este e um teste de envio via Resend.</p>"

@app.get("/api/admin/email-templates")
async def list_email_templates(request: Request, user: dict = Depends(require_admin())):
    db = request.app.db
    items = await db.email_templates.find({}, {"_id": 0}).sort("slug", 1).to_list(200)
    return {"templates": items}

@app.post("/api/admin/email-templates")
async def create_email_template(request: Request, data: EmailTemplateIn, user: dict = Depends(require_admin())):
    db = request.app.db
    if await db.email_templates.find_one({"slug": data.slug}):
        raise HTTPException(status_code=400, detail="slug ja existe")
    doc = {
        "template_id": gen_id("tpl_"),
        **data.model_dump(),
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.email_templates.insert_one(doc)
    return await db.email_templates.find_one({"template_id": doc["template_id"]}, {"_id": 0})

@app.put("/api/admin/email-templates/{template_id}")
async def update_email_template(request: Request, template_id: str, user: dict = Depends(require_admin())):
    db = request.app.db
    body = await request.json()
    allowed = {"name", "subject", "body_html", "body_text", "active", "slug"}
    update = {k: v for k, v in body.items() if k in allowed}
    update["updated_at"] = now_iso()
    await db.email_templates.update_one({"template_id": template_id}, {"$set": update})
    return await db.email_templates.find_one({"template_id": template_id}, {"_id": 0})

@app.delete("/api/admin/email-templates/{template_id}")
async def delete_email_template(request: Request, template_id: str, user: dict = Depends(require_admin())):
    db = request.app.db
    await db.email_templates.delete_one({"template_id": template_id})
    return {"message": "Removido"}

@app.post("/api/admin/email-templates/{template_id}/reset")
async def reset_email_template(request: Request, template_id: str, user: dict = Depends(require_admin())):
    """Restaura template padrao se for um gatilho default."""
    db = request.app.db
    doc = await db.email_templates.find_one({"template_id": template_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Template nao encontrado")
    default = next((t for t in email_service.DEFAULT_TEMPLATES if t["slug"] == doc["slug"]), None)
    if not default:
        raise HTTPException(status_code=400, detail="Este template nao possui versao padrao")
    await db.email_templates.update_one(
        {"template_id": template_id},
        {"$set": {**default, "updated_at": now_iso()}},
    )
    return await db.email_templates.find_one({"template_id": template_id}, {"_id": 0})

@app.post("/api/admin/email-test")
async def send_test_email(request: Request, data: EmailTestIn, user: dict = Depends(require_admin())):
    db = request.app.db
    result = await email_service.send_email(db, data.to, data.subject, data.body_html, meta={"type": "test", "by": user["user_id"]})
    return result

@app.post("/api/admin/email-broadcast")
async def email_broadcast(request: Request, data: EmailBroadcast, user: dict = Depends(require_admin())):
    db = request.app.db
    # Resolver destinatarios
    emails: List[str] = []
    if data.emails:
        emails.extend([str(e) for e in data.emails])
    if data.target == "user_ids" and data.user_ids:
        users = await db.users.find({"user_id": {"$in": data.user_ids}}, {"_id": 0, "email": 1}).to_list(len(data.user_ids))
        emails.extend([u.get("email") for u in users if u.get("email")])
    elif data.target in ("customer", "network_1", "network_2"):
        users = await db.users.find({"network_type": data.target}, {"_id": 0, "email": 1}).to_list(10000)
        emails.extend([u.get("email") for u in users if u.get("email")])
    elif data.target == "admin":
        users = await db.users.find({"role": "admin"}, {"_id": 0, "email": 1}).to_list(100)
        emails.extend([u.get("email") for u in users if u.get("email")])
    elif data.target == "all":
        users = await db.users.find({"status": "active"}, {"_id": 0, "email": 1}).to_list(50000)
        emails.extend([u.get("email") for u in users if u.get("email")])
    # Dedup
    emails = list({e.lower() for e in emails if e})
    if not emails:
        raise HTTPException(status_code=400, detail="Nenhum destinatario encontrado")

    # Dispara em lote (sequencial para nao saturar; Resend limita a ~10 req/s com key free)
    sent = 0
    failed = 0
    for em in emails:
        r = await email_service.send_email(db, em, data.subject, data.body_html, text=data.body_text, meta={"type": "broadcast", "target": data.target, "by": user["user_id"]})
        if r.get("sent"):
            sent += 1
        else:
            failed += 1
    return {"sent": sent, "failed": failed, "total": len(emails)}

@app.get("/api/admin/email-logs")
async def list_email_logs(request: Request, page: int = 1, limit: int = 50, user: dict = Depends(require_admin())):
    db = request.app.db
    total = await db.email_logs.count_documents({})
    logs = await db.email_logs.find({}, {"_id": 0}).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    return {"logs": logs, "total": total, "page": page}

# ==================== EXTERNAL SYNC WEBHOOK (REDE 1) ====================

class ExternalSyncUser(BaseModel):
    external_id: str
    name: str
    email: EmailStr
    leader_external_id: Optional[str] = None
    phone: Optional[str] = None
    cpf: Optional[str] = None
    voucher: Optional[float] = None  # Iter 36: saldo pre-pago em R$ vindo da Maxx

class ExternalSyncPayload(BaseModel):
    action: str = "upsert"   # "upsert" | "delete"
    users: List[ExternalSyncUser]
    default_password: Optional[str] = "oxx@pharma"

@app.post("/api/external/network1/sync")
async def external_network1_sync(request: Request, data: ExternalSyncPayload, x_webhook_token: Optional[str] = Header(None)):
    """Webhook inbound para o sistema externo sincronizar usuarios da Rede 1.

    Autenticacao: header X-Webhook-Token deve coincidir com settings.external_webhook_token.
    Actions: 'upsert' cria/atualiza, 'delete' remove network (nao apaga user).
    """
    db = request.app.db
    settings = await get_settings(db)
    expected = (settings.get("external_webhook_token") or "").strip()
    # Captura metadados da request para o log detalhado
    client_ip = request.client.host if request.client else None
    user_agent = request.headers.get("user-agent")
    raw_payload = data.model_dump() if hasattr(data, "model_dump") else None

    if not expected or not x_webhook_token or x_webhook_token != expected:
        # Log tentativa nao autorizada (preserva payload para auditoria)
        await db.webhook_logs.insert_one({
            "log_id": gen_id("whk_"), "source": "network1_sync",
            "authorized": False, "action": data.action,
            "users_count": len(data.users), "created_at": now_iso(),
            "payload": raw_payload,
            "client_ip": client_ip, "user_agent": user_agent,
            "token_provided": bool(x_webhook_token),
        })
        raise HTTPException(status_code=401, detail="Token invalido")

    stats = {"created": 0, "updated": 0, "deleted": 0, "linked_by_cpf": 0, "errors": []}
    id_map = {}
    if data.action == "upsert":
        pw_hash = hash_pw(data.default_password or "oxx@pharma")
        for row in data.users:
            try:
                existing = await find_existing_user_for_sync(db, row.external_id, str(row.email), row.cpf)
                if existing:
                    match_cpf = _clean_cpf(row.cpf)
                    existing_ext = existing.get("external_id")
                    linked_by_cpf_flag = bool(match_cpf and not existing_ext and match_cpf == _clean_cpf(existing.get("cpf")))
                    update_fields = {
                        "external_id": row.external_id,
                        "name": row.name,
                        "phone": row.phone,
                        "network_type": NETWORK_1,
                        "leader_external_id": row.leader_external_id or None,
                    }
                    if row.cpf:
                        update_fields["cpf"] = row.cpf
                        if match_cpf:
                            update_fields["cpf_digits"] = match_cpf
                    if not existing.get("email") or existing.get("email") == str(row.email).lower():
                        update_fields["email"] = str(row.email).lower()
                    # Iter 36: Voucher (saldo pre-pago)
                    # Acumula incrementos: cada chamada com voucher>0 ADICIONA ao saldo,
                    # cada chamada com voucher=0 nao zera o saldo (mantem o que tem).
                    # Usa unique key (external_id + voucher_token) para evitar duplicacao.
                    if row.voucher is not None and row.voucher > 0:
                        # Soma ao saldo existente
                        await db.users.update_one(
                            {"user_id": existing["user_id"]},
                            {"$inc": {"voucher_balance": float(row.voucher)},
                             "$push": {"voucher_history": {
                                 "delta": float(row.voucher),
                                 "source": "maxx_sync",
                                 "external_id": row.external_id,
                                 "received_at": now_iso(),
                             }}},
                        )
                    await db.users.update_one(
                        {"user_id": existing["user_id"]},
                        {"$set": update_fields},
                    )
                    id_map[row.external_id] = existing["user_id"]
                    stats["updated"] += 1
                    if linked_by_cpf_flag:
                        stats["linked_by_cpf"] += 1
                else:
                    uid = gen_id("user_")
                    cpf_clean = _clean_cpf(row.cpf)
                    voucher_init = float(row.voucher) if (row.voucher and row.voucher > 0) else 0.0
                    await db.users.insert_one({
                        "user_id": uid, "email": row.email.lower(),
                        "password_hash": pw_hash, "name": row.name,
                        "phone": row.phone, "role": "customer", "access_level": 99,
                        "status": "active", "addresses": [],
                        "sponsor_id": None, "sponsor_code": None,
                        "network_type": NETWORK_1,
                        "network_sponsor_id": None,
                        "external_id": row.external_id,
                        "leader_external_id": row.leader_external_id or None,
                        "cpf": row.cpf or None,
                        "cpf_digits": cpf_clean,
                        "voucher_balance": voucher_init,
                        "voucher_history": ([{
                            "delta": voucher_init,
                            "source": "maxx_sync",
                            "external_id": row.external_id,
                            "received_at": now_iso(),
                        }] if voucher_init > 0 else []),
                        "must_set_password": True,  # primeiro acesso obriga a definir senha
                        "referral_program_active": False,
                        "created_at": now_iso(),
                    })
                    id_map[row.external_id] = uid
                    stats["created"] += 1
            except Exception as e:
                stats["errors"].append({"external_id": row.external_id, "error": str(e)})
        # Resolver lideres (helper centralizado: persiste leader_external_id +
        # tenta resolver para este batch e tambem para usuarios pendentes na base)
        link_stats = await resolve_leader_links(db, data.users, id_map)
        stats.update(link_stats)
    elif data.action == "delete":
        for row in data.users:
            existing = await db.users.find_one({"external_id": row.external_id})
            if existing:
                await db.users.update_one(
                    {"user_id": existing["user_id"]},
                    {"$set": {"network_type": NETWORK_CUSTOMER, "network_sponsor_id": None, "revoked_at": now_iso()}},
                )
                stats["deleted"] += 1
    else:
        raise HTTPException(status_code=400, detail=f"action '{data.action}' invalida")

    await db.webhook_logs.insert_one({
        "log_id": gen_id("whk_"), "source": "network1_sync",
        "authorized": True, "action": data.action,
        "users_count": len(data.users), "stats": stats, "created_at": now_iso(),
        "payload": raw_payload,
        "client_ip": client_ip, "user_agent": user_agent,
        "id_map": id_map,
    })
    return stats

@app.get("/api/admin/webhook-logs")
async def list_webhook_logs(request: Request, page: int = 1, limit: int = 50, user: dict = Depends(require_super_admin())):
    db = request.app.db
    # Lista nao retorna o payload completo (economiza banda) - so resumo
    logs = await db.webhook_logs.find(
        {},
        {"_id": 0, "payload": 0},
    ).sort("created_at", -1).skip((page - 1) * limit).limit(limit).to_list(limit)
    total = await db.webhook_logs.count_documents({})
    return {"logs": logs, "total": total, "page": page, "limit": limit, "pages": max(1, (total + limit - 1) // limit)}


@app.get("/api/admin/webhook-logs/{log_id}")
async def get_webhook_log_detail(request: Request, log_id: str, user: dict = Depends(require_super_admin())):
    """Retorna o log completo com payload, headers e mapeamento de IDs."""
    db = request.app.db
    log = await db.webhook_logs.find_one({"log_id": log_id}, {"_id": 0})
    if not log:
        raise HTTPException(status_code=404, detail="Log nao encontrado")
    return log


@app.post("/api/admin/webhook-token/regenerate")
async def regenerate_webhook_token(request: Request, user: dict = Depends(require_super_admin())):
    db = request.app.db
    token = gen_referral_code() + gen_referral_code()
    await db.settings.update_one({"_id": "global"}, {"$set": {"external_webhook_token": token, "updated_at": now_iso()}})
    return {"external_webhook_token": token}

@app.get("/api/admin/points-report/export.xlsx")
async def admin_points_report_xlsx(request: Request, start: Optional[str] = None, end: Optional[str] = None,
                                    user_id: Optional[str] = None, user: dict = Depends(require_admin())):
    """Exporta relatorio de pontos em XLSX (Data/Hora, ID, Nome, Pontos totais)."""
    from fastapi.responses import Response as FastAPIResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    db = request.app.db
    q = {}
    if start:
        q.setdefault("registered_at", {})["$gte"] = start
    if end:
        q.setdefault("registered_at", {})["$lte"] = end
    if user_id:
        q["user_id"] = user_id
    logs = await db.points_log.find(q, {"_id": 0}).sort("registered_at", -1).to_list(100000)

    # Agrupa por user (Data/Hora=ultimo registro do user, ID, Nome, Pontos totais)
    # Conforme pedido: layout simples com colunas Data/Hora, ID, Nome, Pontos totais
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio de Pontos"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="E8731A", end_color="E8731A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Data/Hora", "ID", "Nome", "Pontos totais"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Dados (1 linha por registro de ponto)
    for row_idx, l in enumerate(logs, start=2):
        # Formata Data/Hora pt-BR
        dt = l.get("registered_at", "")
        try:
            dt_obj = datetime.fromisoformat(dt.replace("Z", "+00:00"))
            dt_fmt = dt_obj.strftime("%d/%m/%Y %H:%M:%S")
        except Exception:
            dt_fmt = dt
        ws.cell(row=row_idx, column=1, value=dt_fmt).border = border
        ws.cell(row=row_idx, column=2, value=l.get("user_external_id") or l.get("user_id") or "").border = border
        ws.cell(row=row_idx, column=3, value=l.get("user_name") or "").border = border
        c4 = ws.cell(row=row_idx, column=4, value=float(l.get("points_total") or 0))
        c4.number_format = "#,##0.00"
        c4.border = border

    # Auto-ajusta largura
    widths = [22, 28, 32, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    # Filtro auto
    if logs:
        ws.auto_filter.ref = f"A1:D{len(logs) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FastAPIResponse(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="relatorio-pontos.xlsx"'},
    )


# ==================== CORREIOS / FRETE ====================

@app.get("/api/admin/correios-config")
async def admin_correios_config(request: Request, user: dict = Depends(require_admin())):
    return await correios_service.get_config(request.app.db)


@app.put("/api/admin/correios-config")
async def admin_update_correios_config(request: Request, user: dict = Depends(require_admin())):
    body = await request.json() or {}
    try:
        cfg = await correios_service.update_config(request.app.db, body)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return cfg


@app.get("/api/admin/correios-logs")
async def admin_correios_logs(request: Request, page: int = 1, limit: int = 50, user: dict = Depends(require_admin())):
    db = request.app.db
    total = await db.correios_logs.count_documents({})
    logs = await db.correios_logs.find({}, {"_id": 0}).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    return {"logs": logs, "total": total, "page": page}


# ============ Melhor Envio ============
@app.get("/api/admin/melhorenvio/config")
async def me_admin_get_config(request: Request, user: dict = Depends(require_super_admin())):
    db = request.app.db
    cfg = await melhorenvio_service.get_config(db)
    connected = await melhorenvio_service.is_connected(db)
    tokens = await melhorenvio_service.get_tokens(db)
    # Nao retornar o client_secret integral
    masked_secret = ""
    if cfg.get("client_secret"):
        s = cfg["client_secret"]
        masked_secret = (s[:4] + "•" * max(len(s) - 8, 4) + s[-4:]) if len(s) >= 8 else "••••"
    frontend_origin = request.headers.get("origin") or ""
    # URL de callback derivada do Host atual (funciona em preview + produção)
    host = request.headers.get("host", "")
    scheme = request.headers.get("x-forwarded-proto", "https")
    default_callback = f"{scheme}://{host}/api/admin/melhorenvio/callback" if host else "/api/admin/melhorenvio/callback"
    callback_url = cfg.get("redirect_uri") or default_callback
    return {
        **cfg,
        "client_secret": masked_secret,
        "has_client_secret": bool(cfg.get("client_secret")),
        "connected": connected,
        "token_expires_at": tokens.get("expires_at"),
        "token_last_refresh_at": tokens.get("last_refresh_at"),
        "suggested_callback_url": callback_url,
        "frontend_origin": frontend_origin,
    }


@app.put("/api/admin/melhorenvio/config")
async def me_admin_put_config(request: Request, user: dict = Depends(require_super_admin())):
    db = request.app.db
    body = await request.json() or {}
    # Client secret vazio/mascarado nao deve sobrescrever - proteger
    if not body.get("client_secret") or "•" in str(body.get("client_secret")):
        body.pop("client_secret", None)
    cfg = await melhorenvio_service.update_config(db, body)
    # Mascarar no retorno
    s = cfg.get("client_secret") or ""
    cfg["client_secret"] = (s[:4] + "•" * max(len(s) - 8, 4) + s[-4:]) if len(s) >= 8 else ("••••" if s else "")
    cfg["has_client_secret"] = bool(s)
    return cfg


@app.post("/api/admin/melhorenvio/authorize-url")
async def me_admin_authorize_url(request: Request, user: dict = Depends(require_super_admin())):
    """Gera URL para o admin autorizar a integracao."""
    db = request.app.db
    cfg = await melhorenvio_service.get_config(db)
    if not cfg.get("client_id") or not cfg.get("redirect_uri"):
        raise HTTPException(status_code=400, detail="Configure client_id e redirect_uri primeiro")
    state = gen_id("state_")
    # Salvar state para validar no callback
    await db.app_credentials.update_one(
        {"_id": "melhorenvio_oauth_state"},
        {"$set": {"state": state, "created_at": now_iso(), "created_by": user["user_id"]}},
        upsert=True,
    )
    url = melhorenvio_service.build_authorize_url(cfg, state)
    return {"authorize_url": url, "state": state}


@app.get("/api/admin/melhorenvio/callback")
async def me_admin_callback(request: Request, code: Optional[str] = None, state: Optional[str] = None, error: Optional[str] = None, error_description: Optional[str] = None):
    """Callback OAuth2. O usuario chega aqui vindo do Melhor Envio.
    Como o admin pode estar autenticado via JWT no browser, redirecionamos para a tela
    de admin com os resultados como query params.
    """
    db = request.app.db
    frontend_url = request.headers.get("referer") or os.environ.get("FRONTEND_URL", "")
    # Tentar derivar frontend URL do Origin/Host
    if not frontend_url:
        host = request.headers.get("host", "")
        scheme = request.headers.get("x-forwarded-proto", "https")
        frontend_url = f"{scheme}://{host}"
    redirect_back = f"{frontend_url.rstrip('/')}/backoffice/melhor-envio"

    if error:
        return RedirectResponse(url=f"{redirect_back}?me_error={error}&me_desc={(error_description or '')[:200]}")

    # Validar state
    state_doc = await db.app_credentials.find_one({"_id": "melhorenvio_oauth_state"}) or {}
    if not code or not state or state != state_doc.get("state"):
        return RedirectResponse(url=f"{redirect_back}?me_error=invalid_state")

    try:
        await melhorenvio_service.exchange_code(db, code)
        await db.app_credentials.delete_one({"_id": "melhorenvio_oauth_state"})
        return RedirectResponse(url=f"{redirect_back}?me_success=1")
    except Exception as e:
        logger.exception("me callback exchange falhou")
        return RedirectResponse(url=f"{redirect_back}?me_error=exchange_failed&me_desc={str(e)[:200]}")


@app.post("/api/admin/melhorenvio/disconnect")
async def me_admin_disconnect(request: Request, user: dict = Depends(require_super_admin())):
    db = request.app.db
    await melhorenvio_service.clear_tokens(db)
    return {"success": True}


@app.post("/api/admin/melhorenvio/refresh")
async def me_admin_refresh(request: Request, user: dict = Depends(require_super_admin())):
    db = request.app.db
    try:
        await melhorenvio_service.refresh_access_token(db)
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/admin/melhorenvio/test-calculate")
async def me_admin_test(request: Request, user: dict = Depends(require_super_admin())):
    """Teste rapido: {cep_origin, cep_destination, weight_kg, length_cm, width_cm, height_cm, insurance_value}."""
    db = request.app.db
    body = await request.json() or {}
    items = [{
        "weight_kg": float(body.get("weight_kg") or 0.5),
        "length_cm": float(body.get("length_cm") or 16),
        "width_cm": float(body.get("width_cm") or 11),
        "height_cm": float(body.get("height_cm") or 2),
        "insurance_value": float(body.get("insurance_value") or 50),
        "quantity": 1,
    }]
    return await melhorenvio_service.calculate_shipping(
        db,
        cep_destination=body.get("cep_destination") or "",
        items=items,
        cep_origin=body.get("cep_origin"),
        insurance_value=float(body.get("insurance_value") or 0),
    )


@app.get("/api/admin/melhorenvio/logs")
async def me_admin_logs(request: Request, page: int = 1, limit: int = 50, user: dict = Depends(require_super_admin())):
    db = request.app.db
    total = await db.melhorenvio_logs.count_documents({})
    logs = await db.melhorenvio_logs.find({}, {"_id": 0}).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    return {"logs": logs, "total": total, "page": page}


@app.post("/api/shipping/calculate")
async def public_calculate_shipping(request: Request):
    """Endpoint publico para calcular frete: {cep_destination, items:[{product_id,quantity}]}.

    Para usuarios autenticados, podemos buscar o carrinho automaticamente.
    """
    db = request.app.db
    body = await request.json() or {}
    cep = body.get("cep_destination") or body.get("cep") or ""
    items_in = body.get("items") or []
    declared_value = float(body.get("declared_value") or 0)

    full_items = []
    for it in items_in:
        pid = it.get("product_id")
        qty = int(it.get("quantity") or 1)
        if pid:
            prod = await db.products.find_one({"product_id": pid}, {"_id": 0})
            if prod:
                full_items.append({
                    "weight": prod.get("weight") or 0.3,
                    "length_cm": prod.get("length_cm"),
                    "width_cm": prod.get("width_cm"),
                    "height_cm": prod.get("height_cm"),
                    "quantity": qty,
                })
        else:
            full_items.append({
                "weight": float(it.get("weight") or 0.3),
                "length_cm": it.get("length_cm"),
                "width_cm": it.get("width_cm"),
                "height_cm": it.get("height_cm"),
                "quantity": qty,
            })

    # Se nao veio items, tenta usar carrinho do usuario logado
    if not full_items:
        token = request.headers.get("authorization", "").replace("Bearer ", "")
        if token:
            try:
                payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
                uid = payload.get("sub")
                cart = await db.carts.find_one({"user_id": uid}, {"_id": 0})
                if cart:
                    for ci in cart.get("items", []):
                        prod = await db.products.find_one({"product_id": ci["product_id"]}, {"_id": 0})
                        if prod:
                            full_items.append({
                                "weight": prod.get("weight") or 0.3,
                                "length_cm": prod.get("length_cm"),
                                "width_cm": prod.get("width_cm"),
                                "height_cm": prod.get("height_cm"),
                                "quantity": ci["quantity"],
                            })
            except Exception:
                pass

    result = None
    settings = await _get_site_settings(db)
    provider = (settings.get("shipping_provider") or "correios").lower()

    # Roteamento por provider
    if provider == "melhorenvio":
        me_items = [{
            "weight_kg": it.get("weight") or 0.3,
            "length_cm": it.get("length_cm"),
            "width_cm": it.get("width_cm"),
            "height_cm": it.get("height_cm"),
            "quantity": it.get("quantity", 1),
            "insurance_value": (declared_value / max(len(full_items), 1)) if declared_value else 0,
        } for it in full_items]
        result = await melhorenvio_service.calculate_shipping(db, cep, me_items, insurance_value=declared_value)
        # Fallback para correios se Melhor Envio falhar
        if result.get("error") or not result.get("options"):
            if settings.get("shipping_fallback_to_correios"):
                result = await correios_service.calculate_freight(db, cep, full_items, declared_value)
    else:
        result = await correios_service.calculate_freight(db, cep, full_items, declared_value)

    # Aplica frete grátis se configurado em site_settings
    fs_mode = (settings.get("free_shipping_mode") or "off").lower()
    fs_min = float(settings.get("free_shipping_min_subtotal") or 0)
    fs_label = settings.get("free_shipping_label") or "Frete grátis"
    fs_audiences = settings.get("free_shipping_audiences") or []
    subtotal = float(body.get("subtotal") or declared_value or 0)

    # Descobrir user logado (se houver) para checar audiencia
    current_user_doc = None
    _tok = request.headers.get("authorization", "").replace("Bearer ", "")
    if _tok:
        try:
            payload = jwt.decode(_tok, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            current_user_doc = await db.users.find_one({"user_id": payload.get("sub")}, {"_id": 0, "network_type": 1, "category_ids": 1})
        except Exception:
            pass

    def _user_matches_audiences(u, audiences):
        if not audiences:
            return False
        if not u:
            return False
        ntype = u.get("network_type") or "customer"
        if ntype in audiences:
            return True
        user_cats = u.get("category_ids") or []
        for tok in audiences:
            if isinstance(tok, str) and tok.startswith("cat:") and tok[4:] in user_cats:
                return True
        return False

    apply_free = False
    if fs_mode == "all":
        apply_free = True
    elif fs_mode == "above" and subtotal >= fs_min and fs_min > 0:
        apply_free = True
    elif fs_mode == "audiences":
        # Somente para audiencias selecionadas
        if _user_matches_audiences(current_user_doc, fs_audiences):
            # Se fs_min > 0, exige minimo; senao libera direto
            apply_free = (fs_min <= 0) or (subtotal >= fs_min)

    if apply_free and isinstance(result, dict):
        # Zera price em todos os options retornados (e marca como gratis)
        for opt in (result.get("options") or []):
            opt["original_price"] = float(opt.get("price") or 0)
            opt["price"] = 0.0
            opt["free_shipping"] = True
            opt["free_shipping_label"] = fs_label
        result["free_shipping_applied"] = True
        result["free_shipping_label"] = fs_label
    else:
        result["free_shipping_applied"] = False
        if fs_mode == "above" and fs_min > 0:
            result["free_shipping_threshold"] = fs_min
            result["free_shipping_remaining"] = max(0.0, fs_min - subtotal)

    return result


@app.post("/api/admin/correios-test")
async def admin_correios_test(request: Request, user: dict = Depends(require_admin())):
    """Testa configuracao atual com CEP + peso fornecido."""
    body = await request.json() or {}
    cep = body.get("cep_destination") or "01310100"
    weight = float(body.get("weight") or 0.5)
    items = [{"weight": weight, "quantity": 1}]
    result = await correios_service.calculate_freight(request.app.db, cep, items, 0)
    return result


@app.post("/api/admin/correios-test-auth")
async def admin_correios_test_auth(request: Request, user: dict = Depends(require_admin())):
    """Testa apenas autenticacao com Correios CWS (sem calcular frete)."""
    return await correios_service.test_credentials(request.app.db)


# ==================== MAXX MMN INTEGRATION ====================

def _mask_secret(v: str) -> str:
    if not v: return ""
    s = str(v)
    if len(s) <= 6: return "•" * len(s)
    return f"{s[:3]}{'•' * (len(s) - 6)}{s[-3:]}"


@app.get("/api/admin/maxx-config")
async def admin_maxx_config(request: Request, user: dict = Depends(require_super_admin())):
    cfg = await maxx_service.get_config(request.app.db)
    # Mascara segredo para nao vazar em screenshots/print
    raw = cfg.get("maxx_auth_value") or ""
    cfg["maxx_auth_value"] = _mask_secret(raw)
    cfg["maxx_auth_value_length"] = len(raw)
    cfg["maxx_auth_value_configured"] = bool(raw)
    return cfg


@app.put("/api/admin/maxx-config")
async def admin_update_maxx_config(request: Request, user: dict = Depends(require_super_admin())):
    body = await request.json() or {}
    try:
        cfg = await maxx_service.update_config(request.app.db, body)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    # Retorna mascarado tambem no PUT para UX consistente
    raw = cfg.get("maxx_auth_value") or ""
    cfg["maxx_auth_value"] = _mask_secret(raw)
    cfg["maxx_auth_value_length"] = len(raw)
    cfg["maxx_auth_value_configured"] = bool(raw)
    return cfg


@app.post("/api/admin/maxx-sync-points")
async def admin_maxx_sync(request: Request, user: dict = Depends(require_super_admin())):
    """Dispara envio manual de TODOS os pontos pendentes para o Maxx."""
    return await maxx_service.send_pending_batch(request.app.db, kind="manual")


@app.post("/api/admin/maxx-test-send")
async def admin_maxx_test_send(request: Request, user: dict = Depends(require_super_admin())):
    """Envia um ponto de TESTE para a API Maxx referenciando um usuario real,
    SEM persistir no points_log. Body: {user_id, points_value?, product_name?}.
    """
    db = request.app.db
    body = await request.json() or {}
    uid = (body.get("user_id") or "").strip()
    if not uid:
        raise HTTPException(status_code=400, detail="user_id e obrigatorio")
    return await maxx_service.send_test_for_user(
        db, uid,
        points_value=float(body.get("points_value") or 1.0),
        product_name=str(body.get("product_name") or "[TESTE] Integracao API"),
    )


@app.get("/api/admin/maxx-test-users")
async def admin_maxx_test_users(request: Request, q: str = "", limit: int = 30, user: dict = Depends(require_super_admin())):
    """Busca usuarios para o seletor do teste Maxx. Filtra por nome/email/cpf/external_id."""
    db = request.app.db
    query = {}
    if q:
        import re
        rx = re.compile(re.escape(q), re.IGNORECASE)
        query = {"$or": [{"name": rx}, {"email": rx}, {"cpf": rx}, {"external_id": rx}]}
    users = await db.users.find(
        query,
        {"_id": 0, "user_id": 1, "name": 1, "email": 1, "external_id": 1, "network_type": 1, "cpf": 1},
    ).sort("name", 1).limit(min(limit, 100)).to_list(min(limit, 100))
    return {"users": users}


@app.post("/api/admin/maxx-sync-points/{log_id}")
async def admin_maxx_sync_one(request: Request, log_id: str, user: dict = Depends(require_super_admin())):
    """Reenviar um registro especifico."""
    db = request.app.db
    p = await db.points_log.find_one({"log_id": log_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Log nao encontrado")
    return await maxx_service.send_points(db, [p], kind="manual")


@app.get("/api/admin/maxx-pending-by-user")
async def admin_maxx_pending_by_user(request: Request, user: dict = Depends(require_super_admin())):
    """Agrega points_log pendentes (sent_to_maxx=false AND applied_externally=false)
    por usuario. Usado para o admin enviar em massa por pessoa usuarios que foram
    vinculados via sync CPF apos ja terem gerado pontos."""
    db = request.app.db
    pipeline = [
        {"$match": {
            "$and": [
                {"$or": [{"sent_to_maxx": {"$ne": True}}, {"sent_to_maxx": {"$exists": False}}]},
                {"$or": [{"applied_externally": {"$ne": True}}, {"applied_externally": {"$exists": False}}]},
            ]
        }},
        {"$group": {
            "_id": "$user_id",
            "user_name": {"$first": "$user_name"},
            "user_email": {"$first": "$user_email"},
            "user_external_id": {"$first": "$user_external_id"},
            "points_total": {"$sum": "$points_total"},
            "records_count": {"$sum": 1},
            "oldest_at": {"$min": "$registered_at"},
            "latest_at": {"$max": "$registered_at"},
        }},
        {"$sort": {"latest_at": -1}},
    ]
    rows = await db.points_log.aggregate(pipeline).to_list(5000)
    out = []
    for r in rows:
        uid = r.pop("_id", None)
        has_external = bool(r.get("user_external_id"))
        # Buscar dados frescos do user (pode ter sido vinculado depois)
        udoc = await db.users.find_one({"user_id": uid}, {"_id": 0, "user_id": 1, "name": 1, "email": 1, "external_id": 1, "cpf": 1, "network_type": 1}) or {}
        out.append({
            "user_id": uid,
            "user_name": udoc.get("name") or r.get("user_name"),
            "user_email": udoc.get("email") or r.get("user_email"),
            "user_external_id": udoc.get("external_id"),
            "cpf": udoc.get("cpf"),
            "network_type": udoc.get("network_type"),
            "linked": bool(udoc.get("external_id")),
            "previous_external_id_on_log": has_external,
            "points_total": round(float(r.get("points_total") or 0), 2),
            "records_count": r.get("records_count") or 0,
            "oldest_at": r.get("oldest_at"),
            "latest_at": r.get("latest_at"),
        })
    return {"users": out, "total": len(out)}


@app.post("/api/admin/maxx-sync-user/{user_id}")
async def admin_maxx_sync_user(request: Request, user_id: str, user: dict = Depends(require_super_admin())):
    """Envia ao Maxx todos os points_log pendentes de um unico usuario.
    Antes de enviar, atualiza user_external_id/user_name/user_email nos logs
    com os dados atuais do user (uteis para users vinculados depois via CPF)."""
    db = request.app.db
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    if not target.get("external_id"):
        raise HTTPException(status_code=400, detail="Usuario sem external_id. Vincule primeiro (sync Maxx ou edicao manual).")

    # Atualiza os points_log com os dados mais recentes do user antes de enviar
    await db.points_log.update_many(
        {"user_id": user_id, "sent_to_maxx": {"$ne": True}},
        {"$set": {
            "user_external_id": target.get("external_id"),
            "user_name": target.get("name"),
            "user_email": target.get("email"),
        }},
    )

    pending = await db.points_log.find(
        {"user_id": user_id,
         "$or": [{"sent_to_maxx": {"$ne": True}}, {"sent_to_maxx": {"$exists": False}}],
         "applied_externally": {"$ne": True}},
        {"_id": 0},
    ).to_list(10000)
    if not pending:
        return {"success": True, "sent_count": 0, "skipped": True, "reason": "sem pontos pendentes"}
    return await maxx_service.send_points(db, pending, kind="manual_by_user")


@app.get("/api/users/me/points")
async def my_points_history(request: Request, user: dict = Depends(get_current_user)):
    """Historico de pontos do usuario logado, com totais e status (pendente vs enviado).
    NAO menciona o nome 'Maxx' - expoe apenas 'Pendente' / 'Enviado ao programa'."""
    db = request.app.db
    uid = user["user_id"]
    logs = await db.points_log.find({"user_id": uid}, {
        "_id": 0,
        "log_id": 1, "registered_at": 1, "order_id": 1,
        "product_name": 1, "quantity": 1, "points_unit": 1, "points_total": 1,
        "sent_to_maxx": 1, "sent_to_maxx_at": 1,
        "applied_externally": 1, "applied_at": 1,
    }).sort("registered_at", -1).to_list(1000)

    sent_total = 0.0
    pending_total = 0.0
    for l in logs:
        is_sent = bool(l.get("sent_to_maxx") or l.get("applied_externally"))
        l["status"] = "sent" if is_sent else "pending"
        l["status_label"] = "Enviado ao programa" if is_sent else "Pendente"
        # Data efetiva de envio (qualquer que seja a fonte do 'applied')
        l["effective_sent_at"] = l.get("sent_to_maxx_at") or l.get("applied_at")
        pts = float(l.get("points_total") or 0)
        if is_sent:
            sent_total += pts
        else:
            pending_total += pts
        # Remove flags tecnicas do retorno publico
        l.pop("sent_to_maxx", None)
        l.pop("applied_externally", None)

    return {
        "total_points": round(sent_total + pending_total, 2),
        "sent_total": round(sent_total, 2),
        "pending_total": round(pending_total, 2),
        "records_count": len(logs),
        "logs": logs,
    }



@app.get("/api/admin/maxx-logs")
async def admin_maxx_logs(request: Request, page: int = 1, limit: int = 50, user: dict = Depends(require_super_admin())):
    db = request.app.db
    total = await db.maxx_logs.count_documents({})
    logs = await db.maxx_logs.find({}, {"_id": 0}).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    return {"logs": logs, "total": total, "page": page}


# ==================== SITE SETTINGS (APARENCIA DA LOJA) ====================

DEFAULT_SITE_SETTINGS = {
    "store_name": "OxxPharma",
    "tagline": "Sua farmácia online de confiança",
    "logo_url": "",
    "logo_dark_url": "",
    "favicon_url": "",
    "brand_primary_color": "#E8731A",
    "brand_secondary_color": "#1F2937",
    "hero_title": "Saúde e bem-estar em sua casa",
    "hero_subtitle": "Os melhores produtos farmacêuticos com entrega rápida.",
    "hero_image_url": "",
    "hero_cta_label": "Comprar agora",
    "hero_cta_link": "/produtos",
    "hero_overlay_opacity": 0.4,
    "social_instagram": "",
    "social_facebook": "",
    "social_youtube": "",
    "social_whatsapp": "",
    "footer_about": "Cuidando da sua saúde com qualidade desde 2026.",
    "footer_contact_email": "contato@oxxpharma.com",
    "footer_contact_phone": "(11) 9 9999-9999",
    "footer_address": "Av Paulista, 1000 - São Paulo/SP",
    "footer_pages": [
        {"label": "Sobre nós", "slug": "sobre"},
        {"label": "Política de Privacidade", "slug": "politica-de-privacidade"},
        {"label": "Termos de Uso", "slug": "termos"},
    ],
    "announcement_bar_enabled": False,
    "announcement_bar_text": "",
    "announcement_bar_link": "",
    "announcement_bar_bg_color": "#E8731A",
    "trust_bar_enabled": True,
    "trust_items": [
        {"icon": "Truck", "title": "Entrega rápida", "desc": "Receba em todo Brasil"},
        {"icon": "ShieldCheck", "title": "Compra segura", "desc": "Dados 100% protegidos"},
        {"icon": "CreditCard", "title": "Parcele em até 6x", "desc": "Cartão, PIX ou boleto"},
    ],
    # ============ Programa de indicação editável ============
    "referral_program_name": "Cartão de Benefícios",
    "referral_menu_label": "Indique e ganhe benefícios",
    "referral_box_badge": "NOVO PROGRAMA",
    "referral_box_title": "",  # vazio = usa "Cartão de Benefícios {store_name}"
    "referral_box_description": "Indique amigos e receba suas comissões direto no <b>seu cartão</b>. Adira agora ao programa, gere seu link personalizado e comece a ganhar em cada compra indicada.",
    "referral_box_cta_label": "Aderir ao programa de indicação",
    "referral_box_features": [
        {"icon": "Gift", "title": "Cartão de Benefícios", "desc": "Receba suas comissões em um cartão de benefícios exclusivo."},
        {"icon": "Rocket", "title": "Link exclusivo", "desc": "Compartilhe seu código nas redes sociais."},
        {"icon": "Clock", "title": "Envio diário", "desc": "Todo dia às 23:59 (horário de Brasília) seu saldo é enviado pro cartão."},
    ],
    "referral_box_image_url": "",
    "referral_box_image_width": "320px",
    "referral_box_image_rotation": "-8",
    "referral_box_image_translate_x": "12",
    "referral_box_image_translate_y": "-50",
    "referral_box_image_float": True,
    # ============ Frete grátis ============
    # mode: 'off' | 'all' | 'above' | 'audiences'
    "free_shipping_mode": "off",
    "free_shipping_min_subtotal": 199.0,
    "free_shipping_label": "Frete grátis",
    # tokens aceitos: 'customer' | 'network_1' | 'network_2' | 'cat:{category_id}'
    "free_shipping_audiences": [],
    # ============ Provider de envio ============
    # 'correios' (CWS API - padrao atual) | 'melhorenvio' (multi-transportadora OAuth2)
    "shipping_provider": "correios",
    "shipping_fallback_to_correios": False,
    # ============ Exibição de pontos por produto na loja ============
    # mode: 'none' (ninguém) | 'all' (todos incluindo visitantes) | 'selected'
    # quando 'selected', usa points_visibility_audiences:
    #   tokens aceitos: 'customer', 'network_1', 'network_2', 'cat:{category_id}'
    "points_visibility_mode": "none",
    "points_visibility_audiences": [],
    "points_visibility_label": "pontos",
    "logo_sizes": {
        "store_header":  {"height": 40, "max_width": 180},
        "store_footer":  {"height": 36, "max_width": 160},
        "admin_sidebar": {"height": 36, "max_width": 160},
        "admin_topbar":  {"height": 28, "max_width": 140},
        "auth_pages":    {"height": 48, "max_width": 200},
        "invoice":       {"height": 56, "max_width": 220},
        "email_header":  {"height": 48, "max_width": 200},
        "email_footer":  {"height": 32, "max_width": 140},
    },
}


async def _get_site_settings(db) -> dict:
    s = await db.settings.find_one({"_id": "site"}) or {}
    out = {**DEFAULT_SITE_SETTINGS}
    for k, v in s.items():
        if k != "_id":
            out[k] = v
    return out


@app.get("/api/site-settings")
async def public_site_settings(request: Request):
    return await _get_site_settings(request.app.db)


@app.put("/api/admin/site-settings")
async def admin_site_settings(request: Request, user: dict = Depends(require_super_admin())):
    body = await request.json() or {}
    body["updated_at"] = now_iso()
    db = request.app.db
    await db.settings.update_one({"_id": "site"}, {"$set": body}, upsert=True)
    return await _get_site_settings(db)


@app.post("/api/admin/upload-image")
async def admin_upload_image(request: Request, user: dict = Depends(require_admin())):
    """Upload simples de imagem (base64). Retorna URL data:."""
    body = await request.json() or {}
    data_url = body.get("data") or ""
    if not data_url.startswith("data:"):
        raise HTTPException(status_code=400, detail="Formato invalido (use data URL)")
    # Persiste em coleção uploads para poder referenciar depois (opcional)
    db = request.app.db
    upload_id = "img_" + os.urandom(8).hex()
    await db.uploads.insert_one({
        "upload_id": upload_id,
        "data": data_url,
        "name": body.get("name") or upload_id,
        "uploaded_by": user["user_id"],
        "created_at": now_iso(),
    })
    return {"upload_id": upload_id, "url": data_url}


# ==================== CMS PAGES (Editor visual) ====================

@app.get("/api/pages/{slug}")
async def public_get_page(request: Request, slug: str):
    db = request.app.db
    p = await db.cms_pages.find_one({"slug": slug, "published": True}, {"_id": 0, "components_json": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Pagina nao encontrada")
    return p


@app.get("/api/admin/pages")
async def admin_list_pages(request: Request, user: dict = Depends(require_admin())):
    db = request.app.db
    pages = await db.cms_pages.find({}, {"_id": 0, "html": 0, "css": 0, "components_json": 0}).sort("updated_at", -1).to_list(200)
    return {"pages": pages}


@app.get("/api/admin/pages/{page_id}")
async def admin_get_page(request: Request, page_id: str, user: dict = Depends(require_admin())):
    db = request.app.db
    p = await db.cms_pages.find_one({"page_id": page_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Pagina nao encontrada")
    return p


@app.post("/api/admin/pages")
async def admin_create_page(request: Request, user: dict = Depends(require_admin())):
    body = await request.json() or {}
    slug = (body.get("slug") or "").strip().lower().replace(" ", "-")
    title = body.get("title") or slug
    if not slug:
        raise HTTPException(status_code=400, detail="slug obrigatorio")
    db = request.app.db
    if await db.cms_pages.find_one({"slug": slug}):
        raise HTTPException(status_code=400, detail="slug ja existe")
    page_id = "page_" + os.urandom(6).hex()
    doc = {
        "page_id": page_id, "slug": slug, "title": title,
        "html": body.get("html") or "<div><h1>Nova Página</h1><p>Comece a editar...</p></div>",
        "css": body.get("css") or "",
        "components_json": body.get("components_json"),
        "published": bool(body.get("published", True)),
        "meta_description": body.get("meta_description") or "",
        "created_at": now_iso(), "updated_at": now_iso(),
        "created_by": user["user_id"],
    }
    await db.cms_pages.insert_one(doc)
    return await db.cms_pages.find_one({"page_id": page_id}, {"_id": 0})


@app.put("/api/admin/pages/{page_id}")
async def admin_update_page(request: Request, page_id: str, user: dict = Depends(require_admin())):
    body = await request.json() or {}
    db = request.app.db
    target = await db.cms_pages.find_one({"page_id": page_id})
    if not target:
        raise HTTPException(status_code=404, detail="Pagina nao encontrada")
    update = {k: v for k, v in body.items() if k in {"title", "html", "css", "components_json", "published", "meta_description"}}
    # Slug pode mudar mas valida unicidade
    if body.get("slug"):
        new_slug = body["slug"].strip().lower().replace(" ", "-")
        if new_slug != target.get("slug"):
            if await db.cms_pages.find_one({"slug": new_slug, "page_id": {"$ne": page_id}}):
                raise HTTPException(status_code=400, detail="slug ja existe")
            update["slug"] = new_slug
    update["updated_at"] = now_iso()
    await db.cms_pages.update_one({"page_id": page_id}, {"$set": update})
    return await db.cms_pages.find_one({"page_id": page_id}, {"_id": 0})


@app.delete("/api/admin/pages/{page_id}")
async def admin_delete_page(request: Request, page_id: str, user: dict = Depends(require_admin())):
    db = request.app.db
    res = await db.cms_pages.delete_one({"page_id": page_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pagina nao encontrada")
    return {"ok": True}


@app.get("/api/health")
async def health():
    return {"status": "ok", "app": "OxxPharma"}

# ==================== ADMIN USERS - GESTAO COMPLETA ====================

@app.post("/api/admin/users")
async def admin_create_user(request: Request, user: dict = Depends(require_admin())):
    """Admin cria usuario manualmente sem senha. Usuario recebe email de primeiro acesso para definir a propria senha."""
    db = request.app.db
    body = await request.json() or {}

    email = (body.get("email") or "").strip().lower()
    name = (body.get("name") or "").strip()
    if not email:
        raise HTTPException(status_code=400, detail="Email obrigatorio")
    if not name:
        raise HTTPException(status_code=400, detail="Nome obrigatorio")
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email ja cadastrado")

    role = body.get("role") or "customer"
    # Valida papel solicitado: somente super_admin pode criar admin/super_admin
    creator_role = _role_of(user)
    if role in ("admin", "super_admin") and creator_role != "super_admin":
        raise HTTPException(status_code=403, detail="Apenas Super Admin pode criar Admin ou Super Admin")
    if role not in ("customer", "comercial", "financeiro", "admin", "super_admin"):
        raise HTTPException(status_code=400, detail="Papel invalido")
    network_type = body.get("network_type") or NETWORK_CUSTOMER

    # Sponsor por codigo (opcional)
    sponsor_id = body.get("sponsor_id") or None
    sponsor_code_norm = None
    if body.get("sponsor_code"):
        sp_code = str(body["sponsor_code"]).strip().upper()
        sp = await db.users.find_one({"referral_code": sp_code})
        if sp:
            sponsor_id = sp["user_id"]
            sponsor_code_norm = sp.get("referral_code")

    # Senha temporaria nao-utilizavel (must_set_password=True força fluxo de primeiro acesso)
    placeholder_pw = uuid.uuid4().hex + uuid.uuid4().hex

    user_doc = {
        "user_id": gen_id("user_"),
        "email": email,
        "password_hash": hash_pw(placeholder_pw),
        "must_set_password": True,
        "name": name,
        "phone": (body.get("phone") or "").strip() or None,
        "cpf": (body.get("cpf") or "").strip() or None,
        "cpf_digits": _clean_cpf(body.get("cpf")),
        "external_id": (body.get("external_id") or "").strip() or None,
        "role": role,
        "access_level": 0 if role in ("super_admin", "admin") else (1 if role in ("comercial", "financeiro") else int(body.get("access_level") or 99)),
        "status": body.get("status") or "active",
        "addresses": body.get("addresses") or [],
        "pix_key": body.get("pix_key") or None,
        "pix_key_type": body.get("pix_key_type") or None,
        "referral_code": None,
        "referral_program_active": False,
        "referral_enrollment": None,
        "referral_enrolled_at": None,
        "sponsor_id": sponsor_id,
        "sponsor_code": sponsor_code_norm or (body.get("sponsor_code") or None),
        "network_type": network_type,
        "network_sponsor_id": body.get("network_sponsor_id") or None,
        "category_ids": list(body.get("category_ids") or []),
        "created_at": now_iso(),
        "created_by_admin": user["user_id"],
    }
    await db.users.insert_one(user_doc)

    # Dispara fluxo de primeiro acesso (token + email)
    token = uuid.uuid4().hex + uuid.uuid4().hex
    expires = (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()
    await db.password_reset_tokens.insert_one({
        "token": token,
        "user_id": user_doc["user_id"],
        "email": email,
        "expires_at": expires,
        "used": False,
        "type": "first_access",
        "created_at": now_iso(),
        "created_by_admin": user["user_id"],
    })
    app_url = get_app_url()
    reset_link = f"{app_url}/primeiro-acesso?token={token}"
    if body.get("send_first_access", True):
        asyncio.create_task(email_service.trigger(db, "first_access", email, {
            "user": user_doc, "reset_link": reset_link,
        }))

    fresh = await db.users.find_one({"user_id": user_doc["user_id"]}, {"_id": 0, "password_hash": 0})
    return {"user": fresh, "reset_link": reset_link}


@app.put("/api/admin/users/{user_id}")
async def admin_update_user(request: Request, user_id: str, user: dict = Depends(require_admin())):
    """Atualiza qualquer campo do usuario (exceto password e id)."""
    db = request.app.db
    body = await request.json() or {}
    target = await db.users.find_one({"user_id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    # Campos editaveis
    allowed = {
        "name", "email", "phone", "cpf", "status", "role", "access_level",
        "network_type", "network_sponsor_id", "sponsor_id", "sponsor_code",
        "external_id", "leader_external_id", "addresses", "pix_key", "pix_key_type",
        "referral_program_active", "must_set_password",
    }
    update = {}
    for k, v in body.items():
        if k in allowed:
            update[k] = v
    # Iter 38: somente super_admin pode promover alguem a admin/super_admin
    if "role" in update:
        new_role = update["role"]
        if new_role not in ("customer", "comercial", "financeiro", "admin", "super_admin"):
            raise HTTPException(status_code=400, detail="Papel invalido")
        current_role = target.get("role") or "customer"
        editor_role = _role_of(user)
        is_super = editor_role == "super_admin"
        if new_role != current_role:
            if new_role in ("admin", "super_admin") and not is_super:
                raise HTTPException(status_code=403, detail="Apenas Super Admin pode promover a Admin/Super Admin")
            # Nao-super_admin tambem nao pode rebaixar um admin/super_admin
            if current_role in ("admin", "super_admin") and not is_super:
                raise HTTPException(status_code=403, detail="Apenas Super Admin pode alterar o papel de um Admin")
        # Sincroniza access_level com o papel (backoffice=0/1, cliente=99)
        if new_role in ("super_admin", "admin"):
            update["access_level"] = 0
        elif new_role in ("comercial", "financeiro"):
            update["access_level"] = 1
        else:
            update["access_level"] = 99
    # Validar email unico
    if "email" in update and update["email"]:
        update["email"] = update["email"].lower()
        existing = await db.users.find_one({"email": update["email"], "user_id": {"$ne": user_id}})
        if existing:
            raise HTTPException(status_code=400, detail="Email ja em uso")
    # Sponsor code -> sponsor_id
    if "sponsor_code" in update and update["sponsor_code"]:
        sp = await db.users.find_one({"referral_code": update["sponsor_code"].strip().upper()})
        if sp:
            update["sponsor_id"] = sp["user_id"]
            update["sponsor_code"] = sp.get("referral_code")
    # Se o leader_external_id mudou e network_sponsor_id nao foi enviado explicitamente,
    # tentar resolver automaticamente para o user_id correspondente.
    if "leader_external_id" in update and "network_sponsor_id" not in update:
        leader_ext = (update.get("leader_external_id") or "").strip() if isinstance(update.get("leader_external_id"), str) else update.get("leader_external_id")
        update["leader_external_id"] = leader_ext or None
        if leader_ext:
            leader_doc = await db.users.find_one({"external_id": leader_ext}, {"_id": 0, "user_id": 1})
            update["network_sponsor_id"] = leader_doc["user_id"] if leader_doc else None
        else:
            update["network_sponsor_id"] = None
    # Se o CPF foi alterado, recalcula cpf_digits para manter o indice de match correto
    if "cpf" in update:
        update["cpf_digits"] = _clean_cpf(update.get("cpf"))
    update["updated_at"] = now_iso()
    await db.users.update_one({"user_id": user_id}, {"$set": update})
    fresh = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    return fresh


@app.post("/api/admin/users/{user_id}/set-role")
async def admin_set_user_role(request: Request, user_id: str, payload: dict, user: dict = Depends(require_super_admin())):
    """Define/altera a role administrativa de um usuario.
    Apenas super_admin pode usar. Roles permitidas: customer, comercial, financeiro, admin, super_admin.
    """
    db = request.app.db
    new_role = (payload.get("role") or "").strip()
    if new_role not in ("customer", "comercial", "financeiro", "admin", "super_admin"):
        raise HTTPException(status_code=400, detail="Role invalida")
    target = await db.users.find_one({"user_id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    # Protecao: nao permite super_admin rebaixar a si mesmo (evita ficar sem super_admin)
    if target.get("user_id") == user["user_id"] and new_role != "super_admin":
        raise HTTPException(status_code=400, detail="Voce nao pode rebaixar a propria conta")
    update = {"role": new_role, "updated_at": now_iso()}
    # Alinha access_level pra compat com codigo legado
    if new_role in ("super_admin", "admin"):
        update["access_level"] = 1
    elif new_role in ("comercial", "financeiro"):
        update["access_level"] = 2
    else:
        update["access_level"] = 10
    await db.users.update_one({"user_id": user_id}, {"$set": update})
    return {"ok": True, "user_id": user_id, "role": new_role}


# ==================== IMPERSONATION (#6) ====================
#
# Fluxo: super_admin/admin/comercial lista usuarios, clica em "Entrar como"
# -> gera token especial com flag impersonated=True + impersonated_by=<admin_id>
# -> frontend guarda o token original em sessionStorage.impersonation_return_token
# -> quando termina, chama /api/auth/impersonate/stop -> retorna o token original

IMPERSONATE_ALLOWED_ROLES = {"super_admin", "admin", "comercial"}


@app.post("/api/admin/users/{user_id}/impersonate")
async def admin_impersonate_user(request: Request, user_id: str, user: dict = Depends(get_current_user)):
    """Gera token JWT para acessar a conta do usuario-alvo.
    Apenas super_admin, admin e comercial podem usar."""
    r = _role_of(user)
    if r not in IMPERSONATE_ALLOWED_ROLES:
        raise HTTPException(status_code=403, detail="Seu papel nao permite entrar em outras contas")
    db = request.app.db
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    if target.get("status") in ("cancelled", "deleted"):
        raise HTTPException(status_code=400, detail="Conta alvo desativada")
    # Protecao: nao se pode impersonar outro admin/super_admin (so customers)
    target_role = target.get("role") or "customer"
    if target_role in ADMIN_ROLES and r != "super_admin":
        raise HTTPException(status_code=403, detail="Apenas super_admin pode impersonar outro membro da equipe")
    # Cria token especial
    payload = {
        "sub": target["user_id"],
        "email": target.get("email"),
        "role": target_role,
        "impersonated": True,
        "impersonator_user_id": user.get("user_id"),
        "impersonator_email": user.get("email"),
        "impersonator_role": r,
        "exp": datetime.now(timezone.utc) + timedelta(hours=8),
    }
    token = jwt.encode(payload, JWT_SECRET, algorithm="HS256")
    # Audit log
    try:
        await db.impersonation_audit_log.insert_one({
            "event_id": gen_id("imp_"),
            "impersonator_user_id": user.get("user_id"),
            "impersonator_email": user.get("email"),
            "impersonator_role": r,
            "target_user_id": target["user_id"],
            "target_email": target.get("email"),
            "target_role": target_role,
            "action": "start",
            "created_at": now_iso(),
        })
    except Exception as e:
        logger.warning(f"impersonation audit insert failed: {e}")
    target.pop("password_hash", None)
    return {
        "token": token,
        "user": target,
        "impersonator": {"user_id": user.get("user_id"), "email": user.get("email"), "role": r},
    }


@app.post("/api/auth/impersonate/stop")
async def stop_impersonation(request: Request, user: dict = Depends(get_current_user)):
    """Endpoint para registrar o fim da impersonation (o frontend retoma o token original).
    O token impersonado so vive no cliente; aqui so logamos o fim."""
    db = request.app.db
    payload = getattr(request.state, "jwt_payload", None)
    try:
        # Recupera do header pra saber se era impersonado
        auth = request.headers.get("Authorization", "")
        if auth.lower().startswith("bearer "):
            tok = auth.split(" ", 1)[1]
            data = jwt.decode(tok, JWT_SECRET, algorithms=["HS256"])
            if data.get("impersonated"):
                await db.impersonation_audit_log.insert_one({
                    "event_id": gen_id("imp_"),
                    "impersonator_user_id": data.get("impersonator_user_id"),
                    "impersonator_email": data.get("impersonator_email"),
                    "target_user_id": data.get("sub"),
                    "target_email": data.get("email"),
                    "action": "stop",
                    "created_at": now_iso(),
                })
    except Exception as e:
        logger.warning(f"stop impersonation log failed: {e}")
    return {"ok": True}


@app.get("/api/admin/impersonation-audit-log")
async def admin_impersonation_audit_log(request: Request, limit: int = 100, user: dict = Depends(require_role("financeiro"))):
    """Historico de impersonations (para auditoria)."""
    db = request.app.db
    rows = await db.impersonation_audit_log.find({}, {"_id": 0}).sort("created_at", -1).limit(int(limit)).to_list(int(limit))
    return {"items": rows, "total": len(rows)}


@app.get("/api/users/me/voucher")
async def me_voucher(request: Request, user: dict = Depends(get_current_user)):
    """Saldo + historico de voucher do usuario logado."""
    db = request.app.db
    u = await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0, "voucher_balance": 1, "voucher_history": 1})
    return {
        "balance": round(u.get("voucher_balance") or 0, 2) if u else 0,
        "history": (u.get("voucher_history") or [])[-50:] if u else [],
    }


@app.post("/api/admin/users/{user_id}/voucher-adjust")
async def admin_voucher_adjust(request: Request, user_id: str, payload: dict, user: dict = Depends(require_admin())):
    """Admin/Comercial/Financeiro ajusta o saldo de voucher (credit ou debit).
    Body: {"delta": float, "note": str}
    """
    if _role_of(user) not in ("super_admin", "admin", "financeiro", "comercial"):
        raise HTTPException(status_code=403, detail="Acesso negado")
    db = request.app.db
    try:
        delta = float(payload.get("delta") or 0)
    except Exception:
        raise HTTPException(status_code=400, detail="delta invalido")
    if delta == 0:
        raise HTTPException(status_code=400, detail="delta nao pode ser zero")
    note = (payload.get("note") or "").strip()
    target = await db.users.find_one({"user_id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    new_balance = round((target.get("voucher_balance") or 0) + delta, 2)
    if new_balance < 0:
        raise HTTPException(status_code=400, detail=f"Saldo ficaria negativo (R$ {new_balance:.2f})")
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"voucher_balance": new_balance},
         "$push": {"voucher_history": {
             "delta": delta,
             "source": "admin_adjust",
             "note": note,
             "performed_by": user.get("email"),
             "received_at": now_iso(),
         }}},
    )
    return {"ok": True, "balance": new_balance, "delta": delta}


@app.get("/api/admin/users/{user_id}/voucher")
async def admin_user_voucher(request: Request, user_id: str, user: dict = Depends(require_admin())):
    db = request.app.db
    u = await db.users.find_one({"user_id": user_id}, {"_id": 0, "voucher_balance": 1, "voucher_history": 1, "name": 1, "email": 1})
    if not u:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    return {
        "balance": round(u.get("voucher_balance") or 0, 2),
        "history": u.get("voucher_history") or [],
        "name": u.get("name"),
        "email": u.get("email"),
    }


@app.delete("/api/admin/users/{user_id}")
async def admin_delete_user(request: Request, user_id: str, user: dict = Depends(require_admin())):
    """Hard delete: remove user + commissions + carrinho. Pedidos sao mantidos por integridade fiscal."""
    db = request.app.db
    target = await db.users.find_one({"user_id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    if target.get("role") == "admin" and target.get("user_id") == user["user_id"]:
        raise HTTPException(status_code=400, detail="Voce nao pode deletar a propria conta")
    # Restaurar dependentes downstream (descendentes na rede)
    await db.users.update_many({"network_sponsor_id": user_id}, {"$set": {"network_sponsor_id": None}})
    await db.users.update_many({"sponsor_id": user_id}, {"$set": {"sponsor_id": None, "sponsor_code": None}})
    # Deletar dados
    await db.users.delete_one({"user_id": user_id})
    await db.carts.delete_many({"user_id": user_id})
    await db.commissions.delete_many({"user_id": user_id})
    return {"ok": True, "deleted_user_id": user_id}


@app.post("/api/admin/users/{user_id}/toggle-status")
async def admin_toggle_user_status(request: Request, user_id: str, user: dict = Depends(require_admin())):
    db = request.app.db
    target = await db.users.find_one({"user_id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    new_status = "inactive" if target.get("status") == "active" else "active"
    await db.users.update_one({"user_id": user_id}, {"$set": {"status": new_status, "updated_at": now_iso()}})
    return {"ok": True, "status": new_status}


@app.post("/api/admin/users/{user_id}/send-password-reset")
async def admin_send_password_reset(request: Request, user_id: str, user: dict = Depends(require_admin())):
    db = request.app.db
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    token = uuid.uuid4().hex + uuid.uuid4().hex
    expires = (datetime.now(timezone.utc) + timedelta(minutes=60)).isoformat()
    await db.password_reset_tokens.insert_one({
        "token": token, "user_id": user_id, "email": target.get("email"),
        "expires_at": expires, "used": False, "type": "reset",
        "created_at": now_iso(), "created_by_admin": user["user_id"],
    })
    app_url = get_app_url()
    reset_link = f"{app_url}/redefinir-senha?token={token}"
    asyncio.create_task(email_service.trigger(db, "password_reset", target["email"], {"user": target, "reset_link": reset_link}))
    return {"ok": True, "reset_link": reset_link}


@app.post("/api/admin/users/{user_id}/send-first-access")
async def admin_send_first_access(request: Request, user_id: str, user: dict = Depends(require_admin())):
    """Marca user como must_set_password e envia email com link."""
    db = request.app.db
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    token = uuid.uuid4().hex + uuid.uuid4().hex
    expires = (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()
    await db.password_reset_tokens.insert_one({
        "token": token, "user_id": user_id, "email": target.get("email"),
        "expires_at": expires, "used": False, "type": "first_access",
        "created_at": now_iso(), "created_by_admin": user["user_id"],
    })
    await db.users.update_one({"user_id": user_id}, {"$set": {"must_set_password": True}})
    app_url = get_app_url()
    reset_link = f"{app_url}/primeiro-acesso?token={token}"
    asyncio.create_task(email_service.trigger(db, "first_access", target["email"], {"user": target, "reset_link": reset_link}))
    return {"ok": True, "reset_link": reset_link}


# ==================== AUTH - PASSWORD RESET (PUBLIC) ====================

@app.post("/api/auth/password-reset/request")
async def password_reset_request(request: Request):
    """Usuario solicita reset por email. Sempre retorna sucesso para nao vazar quais emails existem."""
    body = await request.json() or {}
    email = (body.get("email") or "").lower().strip()
    if not email:
        raise HTTPException(status_code=400, detail="Email obrigatorio")
    db = request.app.db
    target = await db.users.find_one({"email": email}, {"_id": 0, "password_hash": 0})
    if target:
        token = uuid.uuid4().hex + uuid.uuid4().hex
        expires = (datetime.now(timezone.utc) + timedelta(minutes=60)).isoformat()
        await db.password_reset_tokens.insert_one({
            "token": token, "user_id": target["user_id"], "email": email,
            "expires_at": expires, "used": False, "type": "reset",
            "created_at": now_iso(),
        })
        app_url = get_app_url()
        reset_link = f"{app_url}/redefinir-senha?token={token}"
        asyncio.create_task(email_service.trigger(db, "password_reset", email, {"user": target, "reset_link": reset_link}))
    return {"ok": True}


@app.get("/api/auth/password-reset/validate")
async def password_reset_validate(request: Request, token: str):
    db = request.app.db
    t = await db.password_reset_tokens.find_one({"token": token, "used": False}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=400, detail="Token invalido ou ja utilizado")
    if t.get("expires_at", "") < now_iso():
        raise HTTPException(status_code=400, detail="Token expirado")
    u = await db.users.find_one({"user_id": t["user_id"]}, {"_id": 0, "password_hash": 0})
    return {"ok": True, "email": t["email"], "name": u.get("name") if u else None, "type": t.get("type", "reset")}


@app.post("/api/auth/password-reset/confirm")
async def password_reset_confirm(request: Request):
    body = await request.json() or {}
    token = body.get("token")
    new_password = body.get("password")
    if not token or not new_password or len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Token e senha (>=6) obrigatorios")
    db = request.app.db
    t = await db.password_reset_tokens.find_one({"token": token, "used": False})
    if not t:
        raise HTTPException(status_code=400, detail="Token invalido")
    if t.get("expires_at", "") < now_iso():
        raise HTTPException(status_code=400, detail="Token expirado")
    pw_hash = hash_pw(new_password)
    await db.users.update_one(
        {"user_id": t["user_id"]},
        {"$set": {"password_hash": pw_hash, "must_set_password": False, "password_changed_at": now_iso()}},
    )
    await db.password_reset_tokens.update_one({"token": token}, {"$set": {"used": True, "used_at": now_iso()}})
    return {"ok": True}


@app.post("/api/auth/first-access/request")
async def first_access_request(request: Request):
    """Usuario importado solicita link de primeiro acesso por email."""
    body = await request.json() or {}
    email = (body.get("email") or "").lower().strip()
    if not email:
        raise HTTPException(status_code=400, detail="Email obrigatorio")
    db = request.app.db
    target = await db.users.find_one({"email": email}, {"_id": 0, "password_hash": 0})
    if target:
        token = uuid.uuid4().hex + uuid.uuid4().hex
        expires = (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()
        await db.password_reset_tokens.insert_one({
            "token": token, "user_id": target["user_id"], "email": email,
            "expires_at": expires, "used": False, "type": "first_access",
            "created_at": now_iso(),
        })
        app_url = get_app_url()
        reset_link = f"{app_url}/primeiro-acesso?token={token}"
        asyncio.create_task(email_service.trigger(db, "first_access", email, {"user": target, "reset_link": reset_link}))
    return {"ok": True}


# ==================== POINTS REPORT ====================

@app.get("/api/admin/points-report")
async def admin_points_report(request: Request, start: Optional[str] = None, end: Optional[str] = None,
                              user_id: Optional[str] = None, applied: Optional[bool] = None,
                              page: int = 1, limit: int = 100, user: dict = Depends(require_admin())):
    db = request.app.db
    q = {}
    if start:
        q.setdefault("registered_at", {})["$gte"] = start
    if end:
        q.setdefault("registered_at", {})["$lte"] = end
    if user_id:
        q["user_id"] = user_id
    if applied is not None:
        q["applied_externally"] = applied
    total = await db.points_log.count_documents(q)
    logs = await db.points_log.find(q, {"_id": 0}).sort("registered_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    summary = await db.points_log.aggregate([
        {"$match": q},
        {"$group": {"_id": None, "total_points": {"$sum": "$points_total"}, "count": {"$sum": 1}}}
    ]).to_list(1)
    return {
        "logs": logs,
        "total": total,
        "page": page,
        "pages": max(1, (total + limit - 1) // limit),
        "limit": limit,
        "summary": {
            "total_points": round((summary[0]["total_points"] if summary else 0), 2),
            "count": summary[0]["count"] if summary else 0,
        },
    }


@app.get("/api/admin/points-report/export.csv")
async def admin_points_report_csv(request: Request, start: Optional[str] = None, end: Optional[str] = None,
                                   user_id: Optional[str] = None, user: dict = Depends(require_admin())):
    from fastapi.responses import Response as FastAPIResponse
    import io, csv
    db = request.app.db
    q = {}
    if start:
        q.setdefault("registered_at", {})["$gte"] = start
    if end:
        q.setdefault("registered_at", {})["$lte"] = end
    if user_id:
        q["user_id"] = user_id
    logs = await db.points_log.find(q, {"_id": 0}).sort("registered_at", -1).to_list(100000)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["data_hora", "user_id", "external_id", "nome", "email", "produto", "qtd", "pontos_unidade", "pontos_total", "pedido", "aplicado_externamente"])
    for l in logs:
        w.writerow([
            l.get("registered_at"), l.get("user_id"), l.get("user_external_id") or "",
            l.get("user_name"), l.get("user_email"), l.get("product_name"),
            l.get("quantity"), l.get("points_per_unit"), l.get("points_total"),
            l.get("order_id"), "Sim" if l.get("applied_externally") else "Nao",
        ])
    return FastAPIResponse(
        content=buf.getvalue().encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="relatorio-pontos.csv"'},
    )


@app.post("/api/admin/points-report/mark-applied")
async def admin_points_mark_applied(request: Request, user: dict = Depends(require_admin())):
    """Marca um conjunto de pontos como aplicado externamente (admin clicou apos aplicar)."""
    body = await request.json() or {}
    log_ids = body.get("log_ids") or []
    if not log_ids:
        raise HTTPException(status_code=400, detail="log_ids obrigatorio")
    db = request.app.db
    res = await db.points_log.update_many(
        {"log_id": {"$in": log_ids}},
        {"$set": {"applied_externally": True, "applied_at": now_iso(), "applied_by": user["user_id"]}},
    )
    return {"ok": True, "modified": res.modified_count}


# ==================== PAYMENTS - ADMIN ====================

@app.get("/api/admin/payments-config")
async def admin_payments_config(request: Request, user: dict = Depends(require_admin())):
    return await payments_service.get_admin_config(request.app.db)


@app.put("/api/admin/payments-config")
async def admin_set_payments_config(request: Request, user: dict = Depends(require_admin())):
    body = await request.json() or {}
    try:
        cfg = await payments_service.update_credentials(request.app.db, body)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return cfg


@app.get("/api/admin/payments-webhook-logs")
async def admin_payments_webhook_logs(request: Request, page: int = 1, limit: int = 50, user: dict = Depends(require_admin())):
    db = request.app.db
    total = await db.payment_webhook_logs.count_documents({})
    logs = await db.payment_webhook_logs.find({}, {"_id": 0}).sort("received_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    return {"logs": logs, "total": total, "page": page}


# ==================== CARTAO DE BENEFICIOS / GIFT CARD ====================

def _validate_enrollment_payload(fields: List[Dict], data: Dict) -> Dict:
    """Valida dados submetidos contra os campos configurados. Retorna dict sanitizado."""
    clean = {}
    for f in fields or []:
        key = f.get("key")
        if not key:
            continue
        val = data.get(key)
        if f.get("required") and (val is None or str(val).strip() == ""):
            raise HTTPException(status_code=400, detail=f"Campo obrigatorio: {f.get('label') or key}")
        if val is not None:
            clean[key] = val
    return clean


@app.get("/api/admin/card-config")
async def get_admin_card_config(request: Request, user: dict = Depends(require_admin())):
    return await card_service.get_card_config(request.app.db)


@app.put("/api/admin/card-config")
async def update_admin_card_config(request: Request, user: dict = Depends(require_admin())):
    body = await request.json()
    cfg = await card_service.update_card_config(request.app.db, body or {})
    return cfg


@app.get("/api/public/card-enrollment-fields")
async def public_card_fields(request: Request):
    """Campos publicos (apenas meta) para montar o formulario no frontend do usuario."""
    cfg = await card_service.get_card_config(request.app.db)
    return {"fields": cfg.get("enrollment_fields", [])}


@app.post("/api/users/me/referral-enrollment")
async def submit_referral_enrollment(request: Request, user: dict = Depends(get_current_user)):
    db = request.app.db
    if user.get("referral_program_active") and user.get("referral_code"):
        raise HTTPException(status_code=400, detail="Usuario ja esta no programa de indicacao")
    if user.get("referral_enrollment_status") == "pending_approval":
        raise HTTPException(status_code=400, detail="Sua adesão já está em análise pelo administrador")
    body = await request.json()
    cfg = await card_service.get_card_config(db)
    clean = _validate_enrollment_payload(cfg.get("enrollment_fields", []), body or {})
    # Cria solicitacao em estado PENDING_APPROVAL (admin precisa aprovar)
    await db.users.update_one(
        {"user_id": user["user_id"]},
        {"$set": {
            "referral_program_active": False,
            "referral_enrollment": clean,
            "referral_enrollment_status": "pending_approval",
            "referral_enrollment_submitted_at": now_iso(),
        }, "$unset": {"referral_code": "", "referral_enrolled_at": "", "referral_rejected_reason": ""}},
    )
    u = await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0, "password_hash": 0})
    return {
        "ok": True,
        "status": "pending_approval",
        "message": "Sua adesão foi enviada e está aguardando aprovação do administrador.",
        "enrollment": clean,
        "user": u,
    }


@app.post("/api/admin/users/{user_id}/approve-referral-enrollment")
async def admin_approve_referral_enrollment(request: Request, user_id: str, user: dict = Depends(require_admin())):
    """Admin aprova solicitacao de adesao -> gera referral_code e ativa programa."""
    db = request.app.db
    target = await db.users.find_one({"user_id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    if target.get("referral_program_active") and target.get("referral_code"):
        return {"ok": True, "already_active": True, "referral_code": target.get("referral_code")}
    code = gen_referral_code()
    while await db.users.find_one({"referral_code": code}):
        code = gen_referral_code()
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {
            "referral_code": code,
            "referral_program_active": True,
            "referral_enrollment_status": "approved",
            "referral_enrolled_at": now_iso(),
            "referral_approved_by_admin": user["user_id"],
            "referral_approved_at": now_iso(),
        }},
    )
    # Iter 40: Promove comissoes pendentes de inscricao para pending (visiveis e pagaveis)
    promo_result = await db.commissions.update_many(
        {"user_id": user_id, "status": "pending_enrollment"},
        {"$set": {"status": "pending", "promoted_on_enrollment_at": now_iso()}},
    )
    if promo_result.modified_count > 0:
        logger.info(f"Iter 40: promovidas {promo_result.modified_count} comissoes pending_enrollment -> pending para {user_id}")
    # best-effort enviar para API do cartao
    try:
        clean = target.get("referral_enrollment") or {}
        await card_service.send_enrollment_to_card_api(db, {**target, "referral_code": code}, clean)
    except Exception as e:
        logger.warning(f"Falha enviando enrollment para API cartao: {e}")
    # email de notificacao
    app_url = get_app_url()
    asyncio.create_task(email_service.trigger(db, "referral_approved", target.get("email"), {
        "user": target, "referral_code": code, "referral_link": f"{app_url}/?ref={code}",
    }))
    return {"ok": True, "referral_code": code}


class _RejectReferralBody(BaseModel):
    reason: str = ""


# ============ Relatorio de aprovados no Programa de Beneficios ============
_REFERRAL_APPROVED_COLS = [
    {"key": "referral_code", "label": "Código", "type": "str", "width": 12},
    {"key": "name", "label": "Nome", "type": "str", "width": 28},
    {"key": "email", "label": "E-mail", "type": "str", "width": 28},
    {"key": "phone", "label": "Telefone", "type": "str", "width": 16},
    {"key": "cpf", "label": "CPF", "type": "str", "width": 16},
    {"key": "rg", "label": "RG", "type": "str", "width": 14},
    {"key": "birth_date", "label": "Data nascimento", "type": "str", "width": 14},
    {"key": "mother_name", "label": "Nome da mãe", "type": "str", "width": 28},
    {"key": "address_zip", "label": "CEP", "type": "str", "width": 10},
    {"key": "address_street", "label": "Rua", "type": "str", "width": 26},
    {"key": "address_number", "label": "Número", "type": "str", "width": 8},
    {"key": "address_complement", "label": "Complemento", "type": "str", "width": 16},
    {"key": "address_neighborhood", "label": "Bairro", "type": "str", "width": 20},
    {"key": "address_city", "label": "Cidade", "type": "str", "width": 20},
    {"key": "address_state", "label": "UF", "type": "str", "width": 4},
    {"key": "pix_key_type", "label": "Tipo PIX", "type": "str", "width": 10},
    {"key": "pix_key", "label": "Chave PIX", "type": "str", "width": 26},
    {"key": "sponsor_name", "label": "Patrocinador", "type": "str", "width": 22},
    {"key": "sponsor_code", "label": "Código patrocinador", "type": "str", "width": 14},
    {"key": "approved_at", "label": "Aprovado em", "type": "str", "width": 20},
    {"key": "approved_by", "label": "Aprovador", "type": "str", "width": 20},
]


def _fmt_cpf(v):
    s = "".join(ch for ch in str(v or "") if ch.isdigit())
    return f"{s[:3]}.{s[3:6]}.{s[6:9]}-{s[9:11]}" if len(s) == 11 else (v or "")


async def _referral_approved_rows(db, start_date: Optional[str] = None, end_date: Optional[str] = None):
    """Retorna lista de usuarios aprovados no Programa de Beneficios.
    start_date/end_date em formato YYYY-MM-DD (filtra por data de aprovacao local TZ America/Sao_Paulo)."""
    q = {"referral_program_active": True, "referral_enrollment_status": "approved"}
    if start_date:
        q["referral_approved_at"] = q.get("referral_approved_at", {})
        q["referral_approved_at"]["$gte"] = f"{start_date}T00:00:00"
    if end_date:
        q["referral_approved_at"] = q.get("referral_approved_at", {})
        # end date inclusivo: usamos end+1 dia na comparacao com string ISO
        q["referral_approved_at"]["$lte"] = f"{end_date}T23:59:59.999"
    users = await db.users.find(q, {"_id": 0, "password_hash": 0}).sort("referral_approved_at", -1).to_list(10000)
    # Enriquecer com sponsor
    sponsor_ids = list({u.get("sponsor_id") for u in users if u.get("sponsor_id")})
    sponsors = {}
    if sponsor_ids:
        async for sp in db.users.find({"user_id": {"$in": sponsor_ids}}, {"_id": 0, "user_id": 1, "name": 1, "referral_code": 1}):
            sponsors[sp["user_id"]] = sp
    # Enriquecer com nomes dos admins aprovadores
    admin_ids = list({u.get("referral_approved_by_admin") for u in users if u.get("referral_approved_by_admin")})
    admins = {}
    if admin_ids:
        async for ad in db.users.find({"user_id": {"$in": admin_ids}}, {"_id": 0, "user_id": 1, "name": 1, "email": 1}):
            admins[ad["user_id"]] = ad
    rows = []
    for u in users:
        enr = u.get("referral_enrollment") or {}
        addr = enr.get("address") or {}
        pix = enr.get("pix") or {}
        sp = sponsors.get(u.get("sponsor_id")) or {}
        ad = admins.get(u.get("referral_approved_by_admin")) or {}
        rows.append({
            "user_id": u.get("user_id"),
            "referral_code": u.get("referral_code") or "",
            "name": enr.get("full_name") or u.get("name") or "",
            "email": u.get("email") or "",
            "phone": enr.get("phone") or u.get("phone") or "",
            "cpf": _fmt_cpf(enr.get("cpf") or u.get("cpf")),
            "rg": enr.get("rg") or "",
            "birth_date": enr.get("birth_date") or "",
            "mother_name": enr.get("mother_name") or "",
            "address_zip": addr.get("zip_code") or "",
            "address_street": addr.get("street") or "",
            "address_number": addr.get("number") or "",
            "address_complement": addr.get("complement") or "",
            "address_neighborhood": addr.get("neighborhood") or "",
            "address_city": addr.get("city") or "",
            "address_state": addr.get("state") or "",
            "pix_key_type": pix.get("key_type") or u.get("pix_key_type") or "",
            "pix_key": pix.get("key") or u.get("pix_key") or "",
            "sponsor_name": sp.get("name") or "",
            "sponsor_code": sp.get("referral_code") or u.get("sponsor_code") or "",
            "approved_at": u.get("referral_approved_at") or u.get("referral_enrolled_at") or "",
            "approved_by": ad.get("name") or ad.get("email") or "",
        })
    return rows


@app.get("/api/admin/referral-approved/batches")
async def admin_referral_approved_batches(request: Request, user: dict = Depends(require_admin())):
    """Agrega aprovados agrupados por dia de aprovacao (TZ America/Sao_Paulo).
    Retorna [{date: 'YYYY-MM-DD', count, total_users_sample: [names...], first_at, last_at}].
    """
    db = request.app.db
    from collections import defaultdict
    all_rows = await _referral_approved_rows(db)
    groups = defaultdict(list)
    for r in all_rows:
        at = r.get("approved_at") or ""
        day = at[:10] if at else "sem-data"
        groups[day].append(r)
    out = []
    for day, rows in sorted(groups.items(), reverse=True):
        out.append({
            "date": day,
            "count": len(rows),
            "first_at": min((r["approved_at"] for r in rows if r.get("approved_at")), default=""),
            "last_at": max((r["approved_at"] for r in rows if r.get("approved_at")), default=""),
            "sample_names": [r["name"] for r in rows[:5]],
        })
    return {"batches": out, "total_approved": len(all_rows)}


@app.get("/api/admin/referral-approved/list")
async def admin_referral_approved_list(request: Request, start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(require_admin())):
    """Lista detalhada. Sem filtro = todos os aprovados."""
    rows = await _referral_approved_rows(request.app.db, start, end)
    return {"approvals": rows, "total": len(rows)}


@app.get("/api/admin/referral-approved/export.csv")
async def admin_referral_approved_csv(request: Request, start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(require_admin())):
    import export_utils
    rows = await _referral_approved_rows(request.app.db, start, end)
    suffix = f"-{start}_to_{end}" if start or end else ""
    return export_utils.csv_response(export_utils.make_csv(rows, _REFERRAL_APPROVED_COLS), f"programa-beneficios-aprovados{suffix}.csv")


@app.get("/api/admin/referral-approved/export.xlsx")
async def admin_referral_approved_xlsx(request: Request, start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(require_admin())):
    import export_utils
    rows = await _referral_approved_rows(request.app.db, start, end)
    suffix = f"-{start}_to_{end}" if start or end else ""
    return export_utils.xlsx_response(export_utils.make_xlsx(rows, _REFERRAL_APPROVED_COLS, sheet_name="Aprovados"), f"programa-beneficios-aprovados{suffix}.xlsx")


@app.post("/api/admin/users/{user_id}/reject-referral-enrollment")
async def admin_reject_referral_enrollment(request: Request, user_id: str, body: _RejectReferralBody, user: dict = Depends(require_admin())):
    db = request.app.db
    target = await db.users.find_one({"user_id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {
            "referral_program_active": False,
            "referral_enrollment_status": "rejected",
            "referral_rejected_reason": body.reason or "Solicitação não aprovada",
            "referral_rejected_by_admin": user["user_id"],
            "referral_rejected_at": now_iso(),
        }, "$unset": {"referral_code": ""}},
    )
    asyncio.create_task(email_service.trigger(db, "referral_rejected", target.get("email"), {
        "user": target, "reason": body.reason or "",
    }))
    return {"ok": True}


@app.get("/api/admin/referral-enrollments/pending")
async def admin_list_pending_referral_enrollments(request: Request, user: dict = Depends(require_admin())):
    db = request.app.db
    cur = db.users.find(
        {"referral_enrollment_status": "pending_approval"},
        {"_id": 0, "password_hash": 0},
    ).sort("referral_enrollment_submitted_at", -1)
    items = await cur.to_list(length=500)
    return {"items": items, "count": len(items)}


@app.post("/api/admin/users/{user_id}/activate-referral")
async def admin_activate_referral(request: Request, user_id: str, user: dict = Depends(require_admin())):
    """Admin ativa manualmente o programa de indicacao para um usuario, sem formulario."""
    db = request.app.db
    target = await db.users.find_one({"user_id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    if target.get("referral_program_active") and target.get("referral_code"):
        return {"ok": True, "already_active": True, "referral_code": target.get("referral_code")}
    code = gen_referral_code()
    while await db.users.find_one({"referral_code": code}):
        code = gen_referral_code()
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {
            "referral_code": code,
            "referral_program_active": True,
            "referral_enrolled_at": now_iso(),
            "referral_activated_by_admin": user["user_id"],
        }},
    )
    # Iter 40: promove comissoes pendentes de inscricao
    promo_result = await db.commissions.update_many(
        {"user_id": user_id, "status": "pending_enrollment"},
        {"$set": {"status": "pending", "promoted_on_enrollment_at": now_iso()}},
    )
    return {"ok": True, "referral_code": code, "promoted_commissions": promo_result.modified_count}


@app.post("/api/admin/users/{user_id}/deactivate-referral")
async def admin_deactivate_referral(request: Request, user_id: str, user: dict = Depends(require_admin())):
    db = request.app.db
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"referral_program_active": False}, "$unset": {"referral_code": ""}},
    )
    return {"ok": True}


@app.post("/api/admin/reset-all-referrals")
async def admin_reset_all_referrals(request: Request, user: dict = Depends(require_admin())):
    """Reseta TODOS os referral_codes (exceto admin), desativando o programa para todos."""
    db = request.app.db
    res = await db.users.update_many(
        {"role": {"$ne": "admin"}},
        {"$set": {"referral_program_active": False}, "$unset": {"referral_code": ""}},
    )
    return {"ok": True, "updated": res.modified_count}


@app.get("/api/admin/card-batches")
async def admin_list_card_batches(request: Request, page: int = 1, limit: int = 50, user: dict = Depends(require_admin())):
    db = request.app.db
    q = {"is_lock": {"$ne": True}}
    total = await db.card_batches.count_documents(q)
    batches = await db.card_batches.find(q, {"_id": 0}).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    return {"batches": batches, "total": total, "page": page}


@app.get("/api/admin/card-batches/{batch_id}")
async def admin_get_card_batch(request: Request, batch_id: str, user: dict = Depends(require_admin())):
    db = request.app.db
    b = await db.card_batches.find_one({"batch_id": batch_id}, {"_id": 0})
    if not b:
        raise HTTPException(status_code=404, detail="Batch nao encontrado")
    return b


@app.post("/api/admin/card-batches/run")
async def admin_run_card_batch(request: Request, user: dict = Depends(require_admin())):
    """Dispara manualmente a transferencia do dia."""
    db = request.app.db
    result = await card_service.run_daily_transfer(db, mode="manual", triggered_by=user["user_id"])
    return result


@app.post("/api/admin/card-batches/{batch_id}/mark-exported")
async def admin_mark_batch_exported(request: Request, batch_id: str, user: dict = Depends(require_admin())):
    db = request.app.db
    b = await card_service.mark_batch_exported(db, batch_id, mode="manual")
    if not b:
        raise HTTPException(status_code=404, detail="Batch nao encontrado")
    return b


@app.get("/api/admin/card-batches/{batch_id}/export.csv")
async def admin_export_batch_csv(request: Request, batch_id: str, user: dict = Depends(require_admin())):
    from fastapi.responses import Response as FastAPIResponse
    db = request.app.db
    b = await db.card_batches.find_one({"batch_id": batch_id}, {"_id": 0})
    if not b:
        raise HTTPException(status_code=404, detail="Batch nao encontrado")
    csv_bytes = card_service.batch_to_csv(b)
    return FastAPIResponse(
        content=csv_bytes,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{batch_id}.csv"'},
    )


@app.get("/api/admin/card-batches/{batch_id}/export.xlsx")
async def admin_export_batch_xlsx(request: Request, batch_id: str, user: dict = Depends(require_admin())):
    from fastapi.responses import Response as FastAPIResponse
    db = request.app.db
    b = await db.card_batches.find_one({"batch_id": batch_id}, {"_id": 0})
    if not b:
        raise HTTPException(status_code=404, detail="Batch nao encontrado")
    xlsx_bytes = card_service.batch_to_xlsx(b)
    return FastAPIResponse(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{batch_id}.xlsx"'},
    )


@app.get("/api/admin/card-logs")
async def admin_list_card_logs(request: Request, page: int = 1, limit: int = 50, user: dict = Depends(require_admin())):
    db = request.app.db
    total = await db.card_api_logs.count_documents({})
    logs = await db.card_api_logs.find({}, {"_id": 0}).sort("created_at", -1).skip((page-1)*limit).limit(limit).to_list(limit)
    return {"logs": logs, "total": total, "page": page}


@app.get("/api/users/me/card-balance")
async def my_card_balance(request: Request, user: dict = Depends(get_current_user)):
    """Saldo 'na conta' (paid ainda nao enviado) + historico detalhado de envios.

    Retorna:
      account_balance: total ainda na conta (paid, sent_to_card != True)
      sent_to_card_total: histórico total enviado
      pending_commissions: comissões pending (ainda não pagas)
      card_history: lista de envios agrupados por dia com data de disponibilidade D+2
        [{ sent_at, amount, available_at, batch_id }]
    """
    db = request.app.db
    account_agg = await db.commissions.aggregate([
        {"$match": {"user_id": user["user_id"], "status": "paid", "sent_to_card": {"$ne": True}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    sent_agg = await db.commissions.aggregate([
        {"$match": {"user_id": user["user_id"], "sent_to_card": True}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    pending_agg = await db.commissions.aggregate([
        {"$match": {"user_id": user["user_id"], "status": "pending"}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)

    # Historico de envios agrupado por batch_id (com data e D+2 de disponibilidade)
    history_pipeline = [
        {"$match": {"user_id": user["user_id"], "sent_to_card": True}},
        {"$group": {
            "_id": {"batch_id": "$card_batch_id", "sent_at": "$sent_to_card_at"},
            "amount": {"$sum": "$amount"},
            "count": {"$sum": 1},
        }},
        {"$sort": {"_id.sent_at": -1}},
        {"$limit": 100},
    ]
    raw = await db.commissions.aggregate(history_pipeline).to_list(100)
    card_history = []
    for r in raw:
        sent_at = r["_id"].get("sent_at")
        batch_id = r["_id"].get("batch_id")
        if not sent_at:
            continue
        try:
            sent_dt = datetime.fromisoformat(sent_at.replace("Z", "+00:00")) if isinstance(sent_at, str) else sent_at
            available_dt = sent_dt + timedelta(days=2)
        except Exception:
            available_dt = None
        card_history.append({
            "sent_at": sent_at,
            "available_at": available_dt.isoformat() if available_dt else None,
            "amount": round(r["amount"], 2),
            "commissions_count": r["count"],
            "batch_id": batch_id,
        })

    return {
        "account_balance": round((account_agg[0]["total"] if account_agg else 0), 2),
        "sent_to_card_total": round((sent_agg[0]["total"] if sent_agg else 0), 2),
        "pending_commissions": round((pending_agg[0]["total"] if pending_agg else 0), 2),
        "card_history": card_history,
        "card_release_days": 2,
    }



# ==================== USER CATEGORIES (admin) ====================

@app.get("/api/admin/user-categories")
async def admin_list_user_categories(request: Request, user: dict = Depends(require_admin())):
    return {"categories": await store_extras.list_user_categories(request.app.db)}


@app.post("/api/admin/user-categories")
async def admin_create_user_category(request: Request, payload: store_extras.UserCategoryIn, user: dict = Depends(require_admin())):
    return await store_extras.create_user_category(request.app.db, payload)


@app.put("/api/admin/user-categories/{category_id}")
async def admin_update_user_category(request: Request, category_id: str, payload: store_extras.UserCategoryIn, user: dict = Depends(require_admin())):
    return await store_extras.update_user_category(request.app.db, category_id, payload)


@app.delete("/api/admin/user-categories/{category_id}")
async def admin_delete_user_category(request: Request, category_id: str, user: dict = Depends(require_admin())):
    return await store_extras.delete_user_category(request.app.db, category_id)


class _UserCatsBody(BaseModel):
    category_ids: List[str] = []


@app.put("/api/admin/users/{user_id}/categories")
async def admin_set_user_categories(request: Request, user_id: str, body: _UserCatsBody, user: dict = Depends(require_admin())):
    return await store_extras.set_user_categories(request.app.db, user_id, body.category_ids)


# ==================== COUPONS ====================

@app.get("/api/admin/coupons")
async def admin_list_coupons(request: Request, user: dict = Depends(require_admin())):
    return {"coupons": await store_extras.list_coupons(request.app.db)}


@app.post("/api/admin/coupons")
async def admin_create_coupon(request: Request, payload: store_extras.CouponIn, user: dict = Depends(require_admin())):
    return await store_extras.create_coupon(request.app.db, payload)


@app.put("/api/admin/coupons/{coupon_id}")
async def admin_update_coupon(request: Request, coupon_id: str, payload: store_extras.CouponIn, user: dict = Depends(require_admin())):
    return await store_extras.update_coupon(request.app.db, coupon_id, payload)


@app.delete("/api/admin/coupons/{coupon_id}")
async def admin_delete_coupon(request: Request, coupon_id: str, user: dict = Depends(require_admin())):
    return await store_extras.delete_coupon(request.app.db, coupon_id)


class _CouponValidateBody(BaseModel):
    code: str
    subtotal: float


@app.post("/api/coupons/validate")
async def public_validate_coupon(request: Request, body: _CouponValidateBody):
    db = request.app.db
    user = await get_optional_user(request)
    return await store_extras.validate_coupon(db, body.code, body.subtotal, user)
