"""Gift Card / Cartao de Beneficios - Transferencia de comissoes.

Fluxo:
1. Compra gera commission (status=paid, in_account=True)
2. Diariamente no horario configurado, commissions com in_account=True sao agrupadas em card_batch.
3. Batch eh transmitido via API ou exportado via Excel/CSV.
4. commissions viram 'sent_to_card' e movem para "saldo enviado ao cartao".
"""
import asyncio
import json
import logging
import uuid
import io
import csv
from datetime import datetime, timezone
from typing import Optional, List, Dict, Any
import httpx
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import pytz

logger = logging.getLogger(__name__)

TZ_BR = pytz.timezone("America/Sao_Paulo")


def now_iso_utc():
    return datetime.now(timezone.utc).isoformat()


async def get_card_config(db) -> Dict:
    """Config do cartao fica em settings sob chave 'card_program_config'."""
    s = await db.settings.find_one({"_id": "card_config"})
    if not s:
        s = {
            "_id": "card_config",
            "enabled": False,
            "cron_hour": 23,
            "cron_minute": 59,
            # Campos do formulario de adesao (array de FieldDef)
            "enrollment_fields": [
                {"key": "cpf", "label": "CPF", "type": "text", "required": True, "mask": "cpf"},
                {"key": "full_name", "label": "Nome completo", "type": "text", "required": True},
                {"key": "birth_date", "label": "Data de nascimento", "type": "date", "required": True},
                {"key": "mother_name", "label": "Nome da mae", "type": "text", "required": True},
                {"key": "phone", "label": "Telefone celular", "type": "text", "required": True, "mask": "phone"},
            ],
            # API do cartao (adapter generico)
            "api_url": "",
            "api_method": "POST",  # POST | PUT
            "api_auth_type": "bearer",  # none | bearer | apikey | basic
            "api_auth_value": "",  # token / key / user:pass
            "api_auth_header_name": "Authorization",  # para apikey custom
            "api_extra_headers": "",  # JSON string: {"X-Client": "OxxPharma"}
            # Template de payload: {{batch}} eh o objeto completo
            "api_payload_template": "",  # JSON string com mustache; vazio = envia batch raw
            "api_timeout_seconds": 30,
            # Endpoint para cadastrar beneficiario (quando usuario adere)
            "enrollment_api_url": "",
            "enrollment_api_method": "POST",
            "enrollment_api_payload_template": "",
            # Iter 42o: Aprovacao automatica de adesoes ao Programa de Beneficios.
            # Quando ligado, o scheduler aprova pendencias cujo
            # submitted_at + auto_approve_delay_minutes >= now (sem precisar do admin).
            "auto_approve_enrollment": False,
            "auto_approve_delay_minutes": 60,
            "created_at": now_iso_utc(),
        }
        await db.settings.insert_one(s)
    return {k: v for k, v in s.items() if k != "_id"}


async def update_card_config(db, update: Dict) -> Dict:
    allowed = {
        "enabled", "cron_hour", "cron_minute", "enrollment_fields",
        "api_url", "api_method", "api_auth_type", "api_auth_value",
        "api_auth_header_name", "api_extra_headers", "api_payload_template",
        "api_timeout_seconds", "enrollment_api_url", "enrollment_api_method",
        "enrollment_api_payload_template",
        "auto_approve_enrollment", "auto_approve_delay_minutes",
    }
    payload = {k: v for k, v in update.items() if k in allowed}
    payload["updated_at"] = now_iso_utc()
    await db.settings.update_one({"_id": "card_config"}, {"$set": payload}, upsert=True)
    return await get_card_config(db)


def _build_headers(config: Dict) -> Dict[str, str]:
    headers = {"Content-Type": "application/json"}
    extra = config.get("api_extra_headers") or ""
    if extra.strip():
        try:
            headers.update(json.loads(extra))
        except Exception as e:
            logger.warning(f"api_extra_headers invalido: {e}")
    auth_type = (config.get("api_auth_type") or "none").lower()
    auth_value = config.get("api_auth_value") or ""
    if auth_type == "bearer" and auth_value:
        headers["Authorization"] = f"Bearer {auth_value}"
    elif auth_type == "apikey" and auth_value:
        headers[config.get("api_auth_header_name") or "X-API-Key"] = auth_value
    elif auth_type == "basic" and auth_value:
        import base64
        encoded = base64.b64encode(auth_value.encode()).decode()
        headers["Authorization"] = f"Basic {encoded}"
    return headers


def _apply_template(template: str, obj: Dict) -> Dict:
    """Se template fornecido, substitui. Senao retorna obj raw."""
    if not template or not template.strip():
        return obj
    try:
        # substituicao simples via json -> str -> replace -> json
        # apoia {{batch}} inteiro e {{batch.field}}
        rendered = template
        flat = json.dumps(obj, ensure_ascii=False, default=str)
        rendered = rendered.replace("{{batch_json}}", flat)
        return json.loads(rendered)
    except Exception as e:
        logger.warning(f"Erro aplicando template: {e}; usando payload raw")
        return obj


async def call_card_api(db, url: str, method: str, payload: Any, config: Dict, context: Dict = None) -> Dict:
    """Executa chamada HTTP contra API do cartao. Retorna dict com status/response."""
    headers = _build_headers(config)
    timeout = int(config.get("api_timeout_seconds") or 30)
    log_entry = {
        "log_id": f"cardapi_{uuid.uuid4().hex[:12]}",
        "url": url,
        "method": method,
        "request_headers": {k: ("***" if k.lower() in ("authorization", "x-api-key") else v) for k, v in headers.items()},
        "request_body": payload,
        "context": context or {},
        "created_at": now_iso_utc(),
    }
    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            r = await client.request(method.upper(), url, json=payload, headers=headers)
            log_entry.update({
                "status_code": r.status_code,
                "response_body": r.text[:5000],
                "success": 200 <= r.status_code < 300,
            })
    except Exception as e:
        logger.exception("Erro chamando API do cartao")
        log_entry.update({"success": False, "error": str(e)[:500]})
    await db.card_api_logs.insert_one(log_entry)
    return log_entry


async def send_enrollment_to_card_api(db, user: Dict, enrollment_data: Dict) -> Dict:
    """Envia dados de adesao para API do cartao (cria beneficiario)."""
    config = await get_card_config(db)
    url = (config.get("enrollment_api_url") or "").strip()
    if not url:
        return {"sent": False, "reason": "enrollment_api_not_configured"}
    body = {
        "user_id": user.get("user_id"),
        "email": user.get("email"),
        "name": user.get("name"),
        **enrollment_data,
    }
    payload = _apply_template(config.get("enrollment_api_payload_template", ""), {"user": user, "enrollment": enrollment_data, "body": body})
    log = await call_card_api(db, url, config.get("enrollment_api_method", "POST"), payload, config, context={"type": "enrollment", "user_id": user.get("user_id")})
    return {"sent": bool(log.get("success")), "log_id": log["log_id"], "status_code": log.get("status_code")}


async def build_daily_batch(db) -> Optional[Dict]:
    """Cria batch com todas as comissoes 'in_account' (status=paid, nao enviadas ainda)."""
    pipeline = [
        {"$match": {"status": "paid", "sent_to_card": {"$ne": True}}},
        {"$group": {
            "_id": "$user_id",
            "total": {"$sum": "$amount"},
            "count": {"$sum": 1},
            "commission_ids": {"$push": "$commission_id"},
        }},
    ]
    grouped = await db.commissions.aggregate(pipeline).to_list(100000)
    if not grouped:
        return None
    uids = [g["_id"] for g in grouped]
    users = await db.users.find({"user_id": {"$in": uids}}, {"_id": 0, "password_hash": 0}).to_list(len(uids))
    umap = {u["user_id"]: u for u in users}
    entries = []
    total = 0
    for g in grouped:
        u = umap.get(g["_id"], {})
        enroll = u.get("referral_enrollment") or {}
        entries.append({
            "user_id": g["_id"],
            "name": u.get("name"),
            "email": u.get("email"),
            "cpf": enroll.get("cpf") or u.get("cpf"),
            "phone": enroll.get("phone") or u.get("phone"),
            "enrollment": enroll,
            "amount": round(g["total"], 2),
            "commissions_count": g["count"],
            "commission_ids": g["commission_ids"],
        })
        total += g["total"]
    return {
        "batch_id": f"batch_{uuid.uuid4().hex[:12]}",
        "entries": entries,
        "total_amount": round(total, 2),
        "users_count": len(entries),
        "reference_date": datetime.now(TZ_BR).date().isoformat(),
        "created_at": now_iso_utc(),
        "status": "queued",  # queued -> sent_api | sent_manual | exported | failed
    }


async def mark_commissions_sent(db, commission_ids: List[str], batch_id: str):
    await db.commissions.update_many(
        {"commission_id": {"$in": commission_ids}},
        {"$set": {"sent_to_card": True, "sent_to_card_at": now_iso_utc(), "card_batch_id": batch_id}},
    )


async def run_daily_transfer(db, mode: str = "auto", triggered_by: Optional[str] = None) -> Dict:
    """Executa transferencia. mode: 'auto'|'manual'. Salva batch + tenta API."""
    batch = await build_daily_batch(db)
    if not batch:
        return {"ran": False, "reason": "no_pending_commissions"}
    batch["triggered_by"] = triggered_by or "scheduler"
    batch["mode"] = mode
    await db.card_batches.insert_one(batch)

    config = await get_card_config(db)
    url = (config.get("api_url") or "").strip()
    if url:
        payload = _apply_template(config.get("api_payload_template", ""), batch)
        log = await call_card_api(db, url, config.get("api_method", "POST"), payload, config, context={"type": "daily_batch", "batch_id": batch["batch_id"]})
        if log.get("success"):
            await db.card_batches.update_one({"batch_id": batch["batch_id"]}, {"$set": {"status": "sent_api", "api_log_id": log["log_id"], "sent_at": now_iso_utc()}})
            cids = [cid for e in batch["entries"] for cid in e["commission_ids"]]
            await mark_commissions_sent(db, cids, batch["batch_id"])
        else:
            await db.card_batches.update_one({"batch_id": batch["batch_id"]}, {"$set": {"status": "failed", "api_log_id": log["log_id"]}})
    # Se nao tem URL, ainda assim marcar como queued para export manual
    return {"ran": True, "batch_id": batch["batch_id"], "users_count": batch["users_count"], "total_amount": batch["total_amount"]}


async def mark_batch_exported(db, batch_id: str, mode: str = "manual"):
    b = await db.card_batches.find_one({"batch_id": batch_id})
    if not b:
        return None
    cids = [cid for e in b.get("entries", []) for cid in e["commission_ids"]]
    await mark_commissions_sent(db, cids, batch_id)
    await db.card_batches.update_one(
        {"batch_id": batch_id},
        {"$set": {"status": f"sent_{mode}", "sent_at": now_iso_utc()}},
    )
    return await db.card_batches.find_one({"batch_id": batch_id}, {"_id": 0})


def batch_to_csv(batch: Dict) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["user_id", "name", "email", "cpf", "phone", "amount", "commissions_count"])
    for e in batch.get("entries", []):
        writer.writerow([e.get("user_id"), e.get("name"), e.get("email"), e.get("cpf"), e.get("phone"), f"{e.get('amount', 0):.2f}", e.get("commissions_count")])
    return buf.getvalue().encode("utf-8-sig")


def batch_to_xlsx(batch: Dict) -> bytes:
    """Mesmas colunas do CSV, mas em XLSX com header em negrito."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io as _io

    wb = Workbook()
    ws = wb.active
    ws.title = "Lote"

    headers = ["user_id", "name", "email", "cpf", "phone", "amount", "commissions_count"]
    ws.append(headers)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="E8731A")
    for col_i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_i)
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")

    for e in batch.get("entries", []):
        ws.append([
            e.get("user_id"), e.get("name"), e.get("email"),
            e.get("cpf"), e.get("phone"),
            float(e.get("amount", 0) or 0),
            int(e.get("commissions_count", 0) or 0),
        ])
    # ajusta largura
    widths = [22, 28, 32, 16, 16, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ==================== SCHEDULER ====================

_scheduler: Optional[AsyncIOScheduler] = None


async def _scheduled_run(db_getter):
    db = db_getter()
    try:
        config = await get_card_config(db)
        if not config.get("enabled"):
            return
        logger.info("Scheduler: executando transferencia diaria")
        await run_daily_transfer(db, mode="auto", triggered_by="scheduler")
    except Exception as e:
        logger.exception(f"Erro no scheduler: {e}")

async def _auto_approve_pending_enrollments(db, config: Dict):
    """Iter 42o: aprova automaticamente adesoes ao Programa de Beneficios cujo
    submitted_at + auto_approve_delay_minutes <= now.

    Implementacao simples (sem importar circular do server.py): replica a logica
    de aprovacao mas inline aqui. Promove status -> approved, gera referral_code
    unico, ativa o programa, promove commissions pending_enrollment -> pending,
    e dispara email best-effort. Marca audit como approved_by_admin=null com
    approved_by_auto=True.
    """
    import string
    import secrets
    from datetime import datetime as _dt, timezone as _tz, timedelta as _td

    delay = int(config.get("auto_approve_delay_minutes") or 0)
    cutoff = _dt.now(_tz.utc) - _td(minutes=delay)

    # Busca pendencias (submitted_at em ISO; aceita string ISO ou datetime)
    pendings = await db.users.find(
        {"referral_enrollment_status": "pending_approval"},
        {"_id": 0, "user_id": 1, "email": 1, "name": 1,
         "referral_enrollment_submitted_at": 1, "referral_enrollment": 1},
    ).to_list(500)

    approved = 0
    for u in pendings:
        ts = u.get("referral_enrollment_submitted_at")
        if not ts:
            continue
        try:
            sub_dt = _dt.fromisoformat(str(ts).replace("Z", "+00:00"))
            if sub_dt.tzinfo is None:
                sub_dt = sub_dt.replace(tzinfo=_tz.utc)
        except Exception:
            continue
        if sub_dt > cutoff:
            continue  # ainda nao deu o delay
        # Gera referral_code unico
        alphabet = string.ascii_uppercase + string.digits
        code = "".join(secrets.choice(alphabet) for _ in range(6))
        while await db.users.find_one({"referral_code": code}):
            code = "".join(secrets.choice(alphabet) for _ in range(6))
        await db.users.update_one(
            {"user_id": u["user_id"], "referral_enrollment_status": "pending_approval"},
            {"$set": {
                "referral_code": code,
                "referral_program_active": True,
                "referral_enrollment_status": "approved",
                "referral_enrolled_at": now_iso_utc(),
                "referral_approved_by_auto": True,
                "referral_approved_at": now_iso_utc(),
            }},
        )
        # Promove comissoes pending_enrollment -> pending
        await db.commissions.update_many(
            {"user_id": u["user_id"], "status": "pending_enrollment"},
            {"$set": {"status": "pending", "promoted_on_enrollment_at": now_iso_utc()}},
        )
        approved += 1
    if approved > 0:
        logger.info(f"Iter 42o: auto-aprovou {approved} adesoes pendentes (delay={delay}min)")
    return approved




def start_scheduler(db_getter):
    """Inicia APScheduler lendo horario atual do config."""
    global _scheduler
    if _scheduler:
        return _scheduler
    _scheduler = AsyncIOScheduler(timezone=TZ_BR)

    async def reload_and_run():
        await _scheduled_run(db_getter)

    # Job fixo: a cada minuto verifica se horario bate
    async def tick():
        db = db_getter()
        try:
            config = await get_card_config(db)
            # Iter 42o: tenta auto-aprovar adesoes pendentes (independente do batch
            # diario do cartao — usa sua propria flag/delay)
            if config.get("auto_approve_enrollment"):
                try:
                    await _auto_approve_pending_enrollments(db, config)
                except Exception as e:
                    logger.exception(f"auto-approve enrollment falhou: {e}")
            if not config.get("enabled"):
                return
            now = datetime.now(TZ_BR)
            if now.hour == int(config.get("cron_hour", 23)) and now.minute == int(config.get("cron_minute", 59)):
                # Evita duplicar no mesmo minuto
                key = f"{now.date().isoformat()}_{now.hour}_{now.minute}"
                existing = await db.card_batches.find_one({"scheduler_key": key})
                if existing:
                    return
                await db.card_batches.insert_one({"batch_id": f"lock_{key}", "scheduler_key": key, "status": "queued", "created_at": now_iso_utc(), "is_lock": True})
                try:
                    await _scheduled_run(db_getter)
                finally:
                    await db.card_batches.delete_one({"batch_id": f"lock_{key}"})
        except Exception as e:
            logger.exception(f"Scheduler tick erro: {e}")

    _scheduler.add_job(tick, CronTrigger(second=0), id="card_tick", replace_existing=True)
    _scheduler.start()
    logger.info("Card scheduler iniciado (tick a cada minuto, TZ America/Sao_Paulo)")
    return _scheduler


def stop_scheduler():
    global _scheduler
    if _scheduler:
        _scheduler.shutdown(wait=False)
        _scheduler = None
